#!/usr/bin/env python
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import subprocess
import sys
import os
import re
import yaml
import threading
import logging
import json
import workout_editor
import calendar
from tkcalendar import Calendar, DateEntry
from datetime import datetime, timedelta

from planner.import_export import cmd_import_workouts, cmd_export_workouts, cmd_delete_workouts
from planner.schedule import cmd_schedule_workouts, cmd_unschedule_workouts
from planner.manage import cmd_list_scheduled, get_scheduled
from planner.garmin_client import cmd_login, GarminClient
from planner.license_manager import LicenseManager

# Disabilita verifica SSL per risolvere problemi di connessione
os.environ['PYTHONHTTPSVERIFY'] = '0'

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("GarminPlannerGUI")


# Determina la directory di base in modo diverso tra script e eseguibile
if getattr(sys, 'frozen', False):
    # Se in esecuzione come eseguibile compilato (PyInstaller)
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    # Se in esecuzione come script Python normale
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Definisci gli altri percorsi basati su SCRIPT_DIR
DEFAULT_OAUTH_FOLDER = os.path.join(SCRIPT_DIR, "oauth")
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")
CACHE_DIR = os.path.join(SCRIPT_DIR, "cache")
EXPORT_DIR = os.path.join(SCRIPT_DIR, "exported")
EXCEL_DIR = os.path.join(SCRIPT_DIR, "excel")
WORKOUTS_CACHE_FILE = os.path.join(CACHE_DIR, "workouts_cache.json")



class GarminPlannerGUI(tk.Tk):
    def __init__(self):
        """Inizializzazione dell'applicazione (con supporto per l'icona nella taskbar di Windows)"""
        super().__init__()
        self.withdraw()
        self.initialize_directories()
        
        # Status bar variable
        self.status_var = tk.StringVar(value="Pronto")

        # Inizializza il license manager
        self.license_manager = LicenseManager.get_instance(SCRIPT_DIR)
        self.features = ["basic"]  # Default features

        self.training_plan = tk.StringVar()
        self.race_day = tk.StringVar()
        self.day_selections = [tk.IntVar() for _ in range(7)]
        self.day_names = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
        
        # Verifica licenza - VERSIONE SEMPLIFICATA
        is_valid, message, features, expiry_date, username = self.license_manager.validate_license()
        if is_valid:
            self.features = features
            self.log(f"Licenza valida con funzionalità: {', '.join(features)}")
            
            # Se la licenza è valida ma sta per scadere (entro 30 giorni), mostra un avviso
            if expiry_date:
                try:
                    expiry = datetime.strptime(expiry_date, "%Y-%m-%d").date()
                    today = datetime.now().date()
                    days_left = (expiry - today).days
                    if 0 < days_left < 30:
                        # Nota: faremo questo dopo che l'interfaccia sarà inizializzata
                        self.show_expiry_warning = True
                        self.days_left = days_left
                    else:
                        self.show_expiry_warning = False
                except:
                    self.show_expiry_warning = False
            else:
                self.show_expiry_warning = False
        else:
            # SEMPLIFICATO: invece del dialogo di attivazione, mostra solo un messaggio
            messagebox.showinfo("Licenza non trovata", 
                             "Garmin Planner funzionerà in modalità limitata.\n\n"
                             "Per sbloccare tutte le funzionalità, posiziona un file license.dat valido "
                             "nella cartella dell'applicazione.")
            
            # Continua in modalità limitata
            self.features = ["basic"]
            self.log("Utilizzo dell'applicazione in modalità limitata")
        
        # Assicurati che la cartella cache esista
        if not os.path.exists(CACHE_DIR):
            os.makedirs(CACHE_DIR, exist_ok=True)
        
        self.title("Garmin Planner")
        self.geometry("800x850")
        
        # Variables
        # Carica l'ultima cartella OAuth usata dal file di configurazione, se esiste
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    config = json.load(f)
                    self.oauth_folder = tk.StringVar(value=config.get('oauth_folder', DEFAULT_OAUTH_FOLDER))
            else:
                self.oauth_folder = tk.StringVar(value=DEFAULT_OAUTH_FOLDER)
        except Exception as e:
            logger.error(f"Errore nel caricamento della configurazione: {str(e)}")
            self.oauth_folder = tk.StringVar(value=DEFAULT_OAUTH_FOLDER)
            
        self.log_level = tk.StringVar(value="INFO")
        self.log_output = []
        
        # Imposta l'icona personalizzata
        try:
            # Percorso dell'icona
            icon_path = os.path.join(SCRIPT_DIR, "assets", "garmin_planner_icon.ico")
            
            # Controlla se il file esiste in questo percorso
            if os.path.exists(icon_path):
                # Usa il percorso normale
                pass
            elif getattr(sys, 'frozen', False):
                # Se siamo in un eseguibile compilato, l'icona potrebbe essere in una posizione diversa
                # PyInstaller potrebbe mettere le risorse in una cartella temporanea
                base_path = getattr(sys, '_MEIPASS', SCRIPT_DIR)
                icon_path = os.path.join(base_path, "assets", "garmin_planner_icon.ico")
            
            # Per Windows: usa entrambi i metodi per massimizzare le possibilità di successo
            if os.name == 'nt':
                self.wm_iconbitmap(icon_path)
                self.iconbitmap(icon_path)
            # Per macOS o Linux
            else:
                icon_img = tk.PhotoImage(file=icon_path)
                self.iconphoto(True, icon_img)
                
            logger.info(f"Icona dell'applicazione caricata da {icon_path}")
        except Exception as e:
            # Se l'icona non può essere caricata, logga l'errore ma continua
            logger.error(f"Impossibile caricare l'icona: {str(e)}")
        
        # Crea uno stile personalizzato per i pulsanti di accento
        self.style = ttk.Style()
        self.style.configure("Accent.TButton", 
                            background="#0076c0",  # Colore di sfondo blu Garmin
                            foreground="white",    # Testo bianco
                            font=("", 10, "bold"))  # Font in grassetto

        # Configura anche gli stati di hover e premuto
        self.style.map("Accent.TButton",
                    background=[('active', '#005486')],  # Più scuro quando attivo/hover
                    foreground=[('active', 'white')])  # Testo rimane bianco
        
        
        # Create the notebook (tabs)
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create tabs
        self.create_login_tab()

        # Funzionalità disponibili nella versione BASIC
        self.create_import_tab()
        self.create_export_tab()

        # Funzionalità disponibili nella versione PRO
        self.create_excel_tools_tab()

        # Funzionalità disponibile solo nella versione PREMIUM
        workout_editor.add_workout_editor_tab(self.notebook, self)
        
        
        # Crea la tab Log per ultima
        self.create_log_tab()
      
        # Crea la tab About
        self.create_about_tab()

        # Status bar
        self.status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Questa è una soluzione alternativa che a volte è necessaria in Windows
        # per forzare l'aggiornamento dell'icona nella taskbar
        if os.name == 'nt':
            try:
                # Importa la libreria win32gui se disponibile
                import win32gui
                # Ottiene l'handle della finestra
                hwnd = self.winfo_id()
                # Forza l'aggiornamento
                win32gui.SendMessage(hwnd, 0x0080, 0, 0)
            except ImportError:
                logger.debug("win32gui non disponibile, impossibile forzare l'aggiornamento dell'icona")
        
        self.deiconify()

        # Se la licenza sta per scadere, mostra un avviso
        if hasattr(self, 'show_expiry_warning') and self.show_expiry_warning:
            self.after(1000, lambda: messagebox.showwarning(
                "Licenza in scadenza",
                f"La tua licenza scadrà tra {self.days_left} giorni.\n"
                f"Rinnova la licenza per continuare ad utilizzare tutte le funzionalità."
            ))

        lm = LicenseManager.get_instance()
        print(f"Features all'avvio: {lm.features}")

    def show_license_info(self):
        """Mostra informazioni sulla licenza"""
        # Ottieni informazioni sulla licenza
        is_valid, message, features, expiry_date, username = self.license_manager.validate_license()
        
        # Crea una semplice finestra di dialogo con le informazioni
        dialog = tk.Toplevel(self)
        dialog.title("Informazioni licenza")
        dialog.geometry("400x300")
        dialog.transient(self)
        dialog.grab_set()
        
        # Contenuto
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        if is_valid:
            ttk.Label(frame, text="Licenza attiva", style="Title.TLabel", font=("Helvetica", 14, "bold"),
                     foreground="green").pack(pady=(0, 10))
            
            # Dettagli licenza
            details_frame = ttk.Frame(frame)
            details_frame.pack(fill=tk.X, pady=10)
            
            # Username
            ttk.Label(details_frame, text="Utente:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
            ttk.Label(details_frame, text=username if username else "Non specificato").grid(
                row=0, column=1, sticky=tk.W, padx=5, pady=2)
            
            # Features
            ttk.Label(details_frame, text="Funzionalità:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
            ttk.Label(details_frame, text=", ".join(features)).grid(
                row=1, column=1, sticky=tk.W, padx=5, pady=2)
            
            # Expiry date
            ttk.Label(details_frame, text="Scadenza:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
            expiry_text = expiry_date if expiry_date else "Licenza perpetua"
            ttk.Label(details_frame, text=expiry_text).grid(
                row=2, column=1, sticky=tk.W, padx=5, pady=2)
        else:
            ttk.Label(frame, text="Licenza non trovata", style="Title.TLabel", font=("Helvetica", 14, "bold"),
                     foreground="red").pack(pady=(0, 10))
            
            ttk.Label(frame, text=message, wraplength=350).pack(pady=10)
            
            ttk.Label(frame, text="Per attivare Garmin Planner, posiziona un file license.dat valido "
                                "nella cartella dell'applicazione.", wraplength=350).pack(pady=10)
        
        # Pulsante chiudi
        ttk.Button(frame, text="Chiudi", command=dialog.destroy).pack(pady=20)


    def initialize_directories(self):
        """
        Verifica che tutte le cartelle necessarie esistano e le crea se necessario.
        """
        # Debug: mostra il percorso base per verificare
        logger.info(f"Directory base dell'applicazione: {SCRIPT_DIR}")
        logger.info(f"È un eseguibile compilato: {getattr(sys, 'frozen', False)}")
        
        # Lista di tutte le cartelle necessarie all'applicazione
        required_dirs = [
            CACHE_DIR,                                # Cartella per la cache
            os.path.join(SCRIPT_DIR, "oauth"),        # Cartella per i token OAuth
            os.path.join(SCRIPT_DIR, "training_plans"),  # Cartella per i piani di allenamento
            EXPORT_DIR,
            EXCEL_DIR,
        ]
        
        # Crea sottocartelle standard in training_plans
        training_subdirs = [
            "5K", "10K", "half_marathon", "marathon", "custom"
        ]
        
        for subdir in training_subdirs:
            required_dirs.append(os.path.join(SCRIPT_DIR, "training_plans", subdir))
        
        # Verifica e crea ogni cartella
        for directory in required_dirs:
            if not os.path.exists(directory):
                try:
                    os.makedirs(directory, exist_ok=True)
                    logger.info(f"Cartella creata: {directory}")
                except Exception as e:
                    logger.error(f"Impossibile creare la cartella {directory}: {str(e)}")
                    # Tenta di mostrare più informazioni sull'errore
                    logger.error(f"Dettagli errore: {type(e).__name__}, Permessi: {os.access(os.path.dirname(directory), os.W_OK)}")

    def run_command_with_live_output(self, cmd):
        """Emula l'esecuzione di un comando, ma usa chiamate dirette alle funzioni"""
        self.log(f"Emulazione del comando: {' '.join(cmd)}")
        
        try:
            # Identifica quale comando si sta cercando di eseguire
            if isinstance(cmd, list) and len(cmd) > 2:
                # Estrai l'operazione dal comando
                operation = None
                for i, arg in enumerate(cmd):
                    if arg in ["import", "export", "schedule", "unschedule", "list", "delete", "login"]:
                        operation = arg
                        break
                
                # Crea un oggetto args adatto all'operazione
                class Args:
                    pass
                
                args = Args()
                
                # Estrai i parametri comuni
                for i, arg in enumerate(cmd):
                    if arg == "--oauth-folder" and i+1 < len(cmd):
                        args.oauth_folder = cmd[i+1]
                    elif arg == "--log-level" and i+1 < len(cmd):
                        args.log_level = cmd[i+1]
                
                # Operazioni diverse richiedono parametri diversi
                if operation == "import":
                    for i, arg in enumerate(cmd):
                        if arg == "--workouts-file" and i+1 < len(cmd):
                            args.workouts_file = cmd[i+1]
                        elif arg == "--name-filter" and i+1 < len(cmd):
                            args.name_filter = cmd[i+1]
                    
                    args.replace = "--replace" in cmd
                    args.dry_run = "--dry-run" in cmd
                    args.treadmill = "--treadmill" in cmd
                    
                    # Esegui l'importazione
                    self.log("Esecuzione diretta della funzione di importazione...")
                    cmd_import_workouts(args)
                    return 0, "Importazione completata con successo", ""
                    
                elif operation == "export":
                    for i, arg in enumerate(cmd):
                        if arg == "--export-file" and i+1 < len(cmd):
                            args.export_file = cmd[i+1]
                        elif arg == "--format" and i+1 < len(cmd):
                            args.format = cmd[i+1]
                        elif arg == "--name-filter" and i+1 < len(cmd):
                            args.name_filter = cmd[i+1]
                        elif arg == "--workout-ids" and i+1 < len(cmd):
                            args.workout_ids = cmd[i+1]
                    
                    args.clean = "--clean" in cmd
                    
                    # Esegui l'esportazione
                    self.log("Esecuzione diretta della funzione di esportazione...")
                    result = cmd_export_workouts(args)
                    return 0, "Esportazione completata con successo", ""
                    
                elif operation == "schedule":
                    for i, arg in enumerate(cmd):
                        if arg == "--training-plan" and i+1 < len(cmd):
                            args.training_plan = cmd[i+1]
                        elif arg == "--race-day" and i+1 < len(cmd):
                            args.race_day = cmd[i+1]
                        elif arg == "--workout-days" and i+1 < len(cmd):
                            args.workout_days = cmd[i+1]
                        elif arg == "--start-day" and i+1 < len(cmd):
                            args.start_day = cmd[i+1]
                    
                    args.dry_run = "--dry-run" in cmd
                    
                    # Esegui la pianificazione
                    self.log("Esecuzione diretta della funzione di pianificazione...")
                    cmd_schedule_workouts(args)
                    return 0, "Pianificazione completata con successo", ""
                    
                elif operation == "unschedule":
                    for i, arg in enumerate(cmd):
                        if arg == "--training-plan" and i+1 < len(cmd):
                            args.training_plan = cmd[i+1]
                        elif arg == "--start-date" and i+1 < len(cmd):
                            args.start_date = cmd[i+1]
                    
                    args.dry_run = "--dry-run" in cmd
                    
                    # Esegui la rimozione della pianificazione
                    self.log("Esecuzione diretta della funzione di unschedule...")
                    cmd_unschedule_workouts(args)
                    return 0, "Rimozione pianificazione completata con successo", ""
                    
                elif operation == "list":
                    for i, arg in enumerate(cmd):
                        if arg == "--start-date" and i+1 < len(cmd):
                            args.start_date = cmd[i+1]
                        elif arg == "--end-date" and i+1 < len(cmd):
                            args.end_date = cmd[i+1]
                        elif arg == "--date-range" and i+1 < len(cmd):
                            args.date_range = cmd[i+1]
                        elif arg == "--name-filter" and i+1 < len(cmd):
                            args.name_filter = cmd[i+1]
                    
                    # Esegui la funzione di lista
                    self.log("Esecuzione diretta della funzione di lista...")
                    result = get_scheduled(args)
                    return 0, str(result), ""
                    
                elif operation == "delete":
                    for i, arg in enumerate(cmd):
                        if arg == "--workout-ids" and i+1 < len(cmd):
                            args.workout_ids = cmd[i+1]
                        elif arg == "--name-filter" and i+1 < len(cmd):
                            args.name_filter = cmd[i+1]
                    
                    # Esegui la funzione di eliminazione
                    self.log("Esecuzione diretta della funzione di eliminazione...")
                    cmd_delete_workouts(args)
                    return 0, "Eliminazione completata con successo", ""
                    
                elif operation == "login":
                    for i, arg in enumerate(cmd):
                        if arg == "--email" and i+1 < len(cmd):
                            args.email = cmd[i+1]
                        elif arg == "--password" and i+1 < len(cmd):
                            args.password = cmd[i+1]
                    
                    # Esegui la funzione di login
                    self.log("Esecuzione diretta della funzione di login...")
                    cmd_login(args)
                    return 0, "Login completato con successo", ""
            
            # Se arriviamo qui, non è stato possibile identificare l'operazione
            self.log(f"AVVISO: Impossibile identificare l'operazione nel comando: {' '.join(cmd)}")
            return 1, "", "Impossibile identificare l'operazione nel comando"
        
        except Exception as e:
            self.log(f"Errore durante l'emulazione del comando: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            return 1, "", str(e)

    def analyze_training_plan(self):
        """Analizza il piano di allenamento selezionato e mostra informazioni all'utente"""
        try:
            training_plan_id = self.training_plan.get().strip()
            if not training_plan_id:
                self.training_plan_info.set("Nessun piano selezionato")
                self.plan_imported = False
                # Reset max_sessions quando non c'è piano selezionato
                self.max_sessions = 0
                return
                    
            # Aggiungi log di debug approfonditi
            self.log(f"Analisi del piano: '{training_plan_id}'")
            self.log(f"DEBUG: Formato dell'ID piano: '{training_plan_id}'")
            self.log(f"DEBUG: Lunghezza ID piano: {len(training_plan_id)} caratteri")
            self.log(f"DEBUG: Caratteri in esadecimale: {' '.join([hex(ord(c)) for c in training_plan_id])}")

            # Se il piano è AM18W115K, controlla specificamente
            if "AM18W115K" in training_plan_id:
                self.log(f"DEBUG: Il piano contiene 'AM18W115K', verificheremo match esatti")
            
            # Prima verifichiamo se esiste un file YAML corrispondente e lo analizziamo
            yaml_path = self.find_yaml_for_plan(training_plan_id)
            if yaml_path:
                self.log(f"Trovato file YAML per il piano: {yaml_path}")
                self.analyze_yaml_plan(yaml_path)
                return
                    
            # Inizializza i contatori
            total_workouts = 0
            sessions_per_week = {}
            
            # Cerca nella cache degli allenamenti
            if os.path.exists(WORKOUTS_CACHE_FILE):
                with open(WORKOUTS_CACHE_FILE, 'r') as f:
                    workouts = json.load(f)
                    
                    # Stampa info della cache
                    self.log(f"DEBUG: Trovati {len(workouts)} allenamenti nella cache")
                    workout_names = [w.get('workoutName', '') for w in workouts]
                    unique_prefixes = set()
                    for name in workout_names:
                        parts = name.split()
                        if parts:
                            unique_prefixes.add(parts[0])
                    self.log(f"DEBUG: Prefissi unici trovati: {unique_prefixes}")
                    
                    # Debug: stampa i primi allenamenti per vedere il formato
                    for i, workout in enumerate(workouts[:5]):
                        workout_name = workout.get('workoutName', '')
                        self.log(f"Allenamento di debug #{i}: '{workout_name}'")
                    
                    for workout in workouts:
                        workout_name = workout.get('workoutName', '')
                        
                        # Verifica con criteri più flessibili
                        is_match = False
                        
                        # 1. Controllo esatto (il piano è una sottostringa esatta)
                        if training_plan_id in workout_name:
                            is_match = True
                        
                        # 2. Controllo ignorando gli spazi alla fine
                        elif training_plan_id.rstrip() in workout_name:
                            is_match = True
                        
                        # 3. Controllo con pattern WxxSxx dopo il prefisso
                        pattern = re.escape(training_plan_id) + r'\s*W\d\dS\d\d'
                        if re.search(pattern, workout_name, re.IGNORECASE):
                            is_match = True
                            self.log(f"Corrispondenza per pattern regex: '{workout_name}' corrisponde a '{pattern}'")

                        # 4. Controllo più permissivo (rimuove caratteri problematici)
                        clean_plan_id = re.sub(r'[^a-zA-Z0-9]', '', training_plan_id)
                        clean_workout_name = re.sub(r'[^a-zA-Z0-9]', '', workout_name)
                        if clean_plan_id and clean_plan_id in clean_workout_name:
                            is_match = True
                            self.log(f"Corrispondenza con ID pulito: '{clean_workout_name}' contiene '{clean_plan_id}'")
                                
                        # Se c'è una corrispondenza, conta l'allenamento
                        if is_match:
                            total_workouts += 1
                            self.log(f"Trovato allenamento: '{workout_name}'")
                            
                            # Cerca il pattern WxxSxx per estrarre settimana e sessione
                            match = re.search(r'\s*(W\d\d)S(\d\d)\s*', workout_name)
                            if match:
                                week_id = match.group(1)
                                
                                # Incrementa il contatore per questa settimana
                                if week_id not in sessions_per_week:
                                    sessions_per_week[week_id] = 0
                                sessions_per_week[week_id] += 1
            
            # Se abbiamo trovato allenamenti, mostra le informazioni
            if total_workouts > 0:
                weeks = len(sessions_per_week)
                max_sessions = max(sessions_per_week.values()) if sessions_per_week else 0
                
                # Salva il numero massimo di sessioni come attributo della classe
                self.max_sessions = max_sessions
                
                info_text = f"Piano: {total_workouts} allenamenti, {weeks} settimane"
                if sessions_per_week:
                    info_text += f"\nAllenamenti per settimana: "
                    for week, count in sorted(sessions_per_week.items()):
                        info_text += f"{week}={count} "
                        
                self.training_plan_info.set(info_text)
                
                # Aggiorna il testo informativo sui giorni
                if max_sessions > 0:
                    self.day_info_label.config(text=f"Devi selezionare esattamente {max_sessions} giorni per settimana")
                else:
                    self.day_info_label.config(text="Seleziona i giorni preferiti per gli allenamenti")
                    
                # Preseleziona i giorni in base al numero di sessioni
                self.preselect_days(max_sessions)
                    
                # Imposta il flag di piano importato
                self.plan_imported = True
                self.log(f"Piano '{training_plan_id}' trovato con {total_workouts} allenamenti")
            else:
                self.training_plan_info.set(f"Nessun allenamento trovato per '{training_plan_id}'")
                self.day_info_label.config(text="Seleziona i giorni preferiti per gli allenamenti")
                # Imposta il flag di piano non importato
                self.plan_imported = False
                # Resetta max_sessions se non ci sono allenamenti
                self.max_sessions = 0
                self.log(f"Nessun allenamento trovato per il piano '{training_plan_id}'")
                    
        except Exception as e:
            self.log(f"Errore nell'analisi del piano: {str(e)}")
            self.training_plan_info.set("Errore nell'analisi del piano")
            self.plan_imported = False
            self.max_sessions = 0



    def load_workouts_from_cache(self):
        """Carica la lista degli allenamenti dalla cache locale"""
        try:
            if os.path.exists(WORKOUTS_CACHE_FILE):
                with open(WORKOUTS_CACHE_FILE, 'r') as f:
                    workouts = json.load(f)
                    
                    # Clear existing items
                    for item in self.workouts_tree.get_children():
                        self.workouts_tree.delete(item)
                    
                    # Add workouts to the tree view
                    for workout in workouts:
                        workout_id = workout.get('workoutId', 'N/A')
                        workout_name = workout.get('workoutName', 'Senza nome')
                        self.workouts_tree.insert("", "end", values=(workout_id, workout_name))
                    
                    self.log(f"Caricati {len(workouts)} allenamenti dalla cache")
            else:
                self.log("Nessuna cache di allenamenti trovata. Usa 'Aggiorna lista' per scaricare gli allenamenti.")
        except Exception as e:
            self.log(f"Errore nel caricamento della cache: {str(e)}")

    def save_workouts_to_cache(self, workouts):
        """Salva la lista degli allenamenti nella cache locale"""
        try:
            with open(WORKOUTS_CACHE_FILE, 'w') as f:
                json.dump(workouts, f, indent=2)
            self.log(f"Salvati {len(workouts)} allenamenti nella cache")
        except Exception as e:
            self.log(f"Errore nel salvataggio della cache: {str(e)}")

    def save_config(self):
        """Salva la configurazione attuale"""
        try:
            config = {
                'oauth_folder': self.oauth_folder.get()
            }
            with open(CONFIG_FILE, 'w') as f:
                json.dump(config, f)
        except Exception as e:
            logger.error(f"Errore nel salvataggio della configurazione: {str(e)}")

    def on_day_checkbox_clicked(self, day_index):
        """Gestisce direttamente il click sul checkbox, valutando il nuovo stato e limitando le selezioni"""
        # Se non c'è un limite massimo, permetti qualsiasi selezione
        if not hasattr(self, 'max_sessions') or self.max_sessions <= 0:
            return
        
        # Conta quanti giorni sono selezionati DOPO il click (il valore è già stato aggiornato)
        selected_count = sum(var.get() for var in self.day_selections)
        
        # Se abbiamo superato il limite e questo checkbox è selezionato, deselezionalo
        if selected_count > self.max_sessions and self.day_selections[day_index].get() == 1:
            # Deseleziona il checkbox
            self.day_selections[day_index].set(0)
            
            # Mostra il messaggio informativo
            messagebox.showinfo("Limite raggiunto", 
                             f"Puoi selezionare al massimo {self.max_sessions} giorni per questo piano di allenamento.")

    def create_settings_frame(self):
        """Frame per le impostazioni comuni (versione aggiornata senza OAuth)"""
        settings_frame = ttk.LabelFrame(self, text="Impostazioni comuni")
        settings_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Rimuovo le impostazioni della cartella OAuth poiché sono già nella tab Login
        
        # Per mantenere il layout, lasciamo uno spazio vuoto o un messaggio informativo
        info_label = ttk.Label(settings_frame, text="Usa la tab Login per gestire le impostazioni di autenticazione")
        info_label.pack(padx=10, pady=10)

    def browse_oauth_folder(self):
        folder = filedialog.askdirectory(initialdir=self.oauth_folder.get())
        if folder:
            self.oauth_folder.set(folder)
            self.save_config()  # Salva la configurazione quando viene cambiata la cartella

    def create_login_tab(self):
        """Crea la tab Login"""
        login_frame = ttk.Frame(self.notebook)
        self.notebook.add(login_frame, text="Login")
        
        # Chiama il metodo che crea il contenuto effettivo
        self.create_login_tab_content(login_frame)


    def create_login_tab_content(self, login_frame):
        """Crea il contenuto della tab Login (separato per permettere il refresh)"""
        # Verifica se esistono già file di autenticazione
        oauth_folder = self.oauth_folder.get()
        oauth_files_exist = False
        
        if os.path.exists(oauth_folder):
            # Controlla la presenza di file oauth
            oauth_files = [f for f in os.listdir(oauth_folder) if f.startswith('oauth') and f.endswith('.json')]
            if oauth_files:
                oauth_files_exist = True
        
        # Titolo della pagina
        ttk.Label(login_frame, text="Accesso a Garmin Connect", font=("", 14, "bold")).pack(pady=20)
        
        # Frame per lo stato di login
        status_frame = ttk.LabelFrame(login_frame, text="Stato di accesso")
        status_frame.pack(fill=tk.X, padx=20, pady=10)
        
        if oauth_files_exist:
            # Mostra informazioni di login esistente
            status_icon = "✅"  # Simbolo di spunta
            status_color = "green"
            status_message = "Hai già effettuato l'accesso a Garmin Connect."
            last_login = "Ultimo accesso: " + self.get_last_login_time(oauth_folder)
        else:
            # Mostra informazioni di login necessario
            status_icon = "❌"  # Simbolo X
            status_color = "red"
            status_message = "Non hai ancora effettuato l'accesso a Garmin Connect."
            last_login = "È necessario effettuare l'accesso per utilizzare le funzionalità di pianificazione."
        
        # Visualizza lo stato
        status_container = ttk.Frame(status_frame)
        status_container.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(status_container, text=status_icon, font=("", 24)).pack(side=tk.LEFT, padx=10)
        
        status_text = ttk.Frame(status_container)
        status_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        ttk.Label(status_text, text=status_message, font=("", 12, "bold"), foreground=status_color).pack(anchor=tk.W)
        ttk.Label(status_text, text=last_login).pack(anchor=tk.W, pady=5)
        
        # Descrizione
        description_frame = ttk.Frame(login_frame)
        description_frame.pack(fill=tk.X, padx=20, pady=10)
        
        description_text = (
            "Per utilizzare Garmin Planner è necessario effettuare l'accesso al tuo account Garmin Connect.\n\n"
            "Le credenziali verranno salvate in modo sicuro nella cartella OAuth e utilizzate per comunicare con Garmin Connect.\n\n"
            "Non è necessario effettuare nuovamente l'accesso ad ogni avvio dell'applicazione."
        )
        ttk.Label(description_frame, text=description_text, wraplength=600, justify="left").pack(pady=10)
        
        # Pulsanti
        button_frame = ttk.Frame(login_frame)
        button_frame.pack(pady=20)
        
        if oauth_files_exist:
            ttk.Button(button_frame, text="Aggiorna credenziali", command=self.perform_login).pack(side=tk.LEFT, padx=10)
            ttk.Button(button_frame, text="Verifica connessione", command=self.check_connection).pack(side=tk.LEFT, padx=10)
            ttk.Button(button_frame, text="Logout", command=self.perform_logout).pack(side=tk.LEFT, padx=10)
        else:
            # Usa un pulsante tk standard con colori espliciti
            login_button = tk.Button(button_frame, 
                                  text="Effettua login", 
                                  command=self.perform_login,
                                  bg="#0076c0",       # colore di sfondo blu
                                  fg="white",         # testo bianco
                                  font=("", 10, "bold"),  # font in grassetto
                                  padx=20,
                                  pady=10,
                                  relief=tk.RAISED,   # bordo in rilievo
                                  cursor="hand2")     # cambia il cursore al passaggio
            login_button.pack(padx=10, pady=5)
            
            # Gestione hover ed eventi del mouse
            def on_enter(e):
                login_button['bg'] = '#005486'  # blu più scuro al passaggio
                
            def on_leave(e):
                login_button['bg'] = '#0076c0'  # ritorna al blu normale
                
            def on_click(e):
                login_button['relief'] = tk.SUNKEN  # effetto premuto
                
            def on_release(e):
                login_button['relief'] = tk.RAISED  # torna al normale
                self.after(100, self.perform_login)  # esegue l'azione
            
            # Associa gli eventi
            login_button.bind("<Enter>", on_enter)
            login_button.bind("<Leave>", on_leave)
            login_button.bind("<ButtonPress-1>", on_click)
            login_button.bind("<ButtonRelease-1>", on_release)
        
        # Info sulla cartella OAuth
        info_frame = ttk.Frame(login_frame)
        info_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Label(info_frame, text=f"Cartella OAuth: {oauth_folder}", font=("", 8)).pack(side=tk.LEFT)
        ttk.Button(info_frame, text="Cambia", command=self.browse_oauth_folder, width=8).pack(side=tk.LEFT, padx=5)


    def get_last_login_time(self, oauth_folder):
        """Ottiene la data dell'ultimo login in base ai timestamp dei file OAuth"""
        try:
            if os.path.exists(oauth_folder):
                oauth_files = [os.path.join(oauth_folder, f) for f in os.listdir(oauth_folder) 
                              if f.startswith('oauth') and f.endswith('.json')]
                
                if oauth_files:
                    # Ottieni il timestamp più recente
                    timestamps = [os.path.getmtime(f) for f in oauth_files]
                    latest_timestamp = max(timestamps)
                    
                    # Converti il timestamp in data/ora leggibile
                    last_login_time = datetime.fromtimestamp(latest_timestamp)
                    return last_login_time.strftime("%d/%m/%Y %H:%M:%S")
                    
            return "Data non disponibile"
        except Exception as e:
            self.log(f"Errore nel recupero della data di ultimo accesso: {str(e)}")
            return "Data non disponibile"

    def check_connection(self):
        """Verifica la connessione a Garmin Connect usando le credenziali salvate"""
        self.log("Verifica della connessione a Garmin Connect...")
        
        # Disabilita temporaneamente i pulsanti
        for widget in self.winfo_children():
            if isinstance(widget, ttk.Button):
                widget.configure(state="disabled")
        
        # Esegui il controllo in un thread separato
        threading.Thread(target=self._do_check_connection).start()

    def _do_check_connection(self):
        """Implementazione della verifica della connessione"""
        try:
            # Importa le librerie necessarie
            import sys
            sys.path.append(SCRIPT_DIR)
            
            # Usa il GarminClient per verificare la connessione
            from planner.garmin_client import GarminClient
            
            client = GarminClient(self.oauth_folder.get())
            
            # Prova a ottenere la lista degli allenamenti come test
            response = client.list_workouts()
            
            if response:
                self.log("Connessione a Garmin Connect verificata con successo!")
                messagebox.showinfo("Connessione riuscita", 
                                  "La connessione a Garmin Connect è attiva e funzionante.\n\n"
                                  f"Trovati {len(response)} allenamenti nel tuo account.")
            else:
                self.log("Connessione a Garmin Connect non riuscita. Nessuna risposta ricevuta.")
                messagebox.showerror("Errore di connessione", 
                                   "La connessione a Garmin Connect non ha restituito dati.\n\n"
                                   "Prova ad aggiornare le credenziali.")
                
        except Exception as e:
            self.log(f"Errore nella verifica della connessione: {str(e)}")
            messagebox.showerror("Errore di connessione", 
                               f"Si è verificato un errore durante la verifica della connessione:\n\n{str(e)}")
        
        finally:
            # Riabilita i pulsanti nel thread principale
            self.after(0, self._re_enable_buttons)

    def _re_enable_buttons(self):
        """Riabilita i pulsanti dopo una operazione"""
        for widget in self.winfo_children():
            if isinstance(widget, ttk.Button):
                widget.configure(state="normal")


    def create_about_tab(self):
        """Crea la tab About con informazioni sull'applicazione e sulla licenza"""
        about_frame = ttk.Frame(self.notebook)
        self.notebook.add(about_frame, text="Info")
        
        # Header con logo (se disponibile)
        header_frame = ttk.Frame(about_frame)
        header_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # Titolo e versione
        ttk.Label(header_frame, text="Garmin Planner", font=("Helvetica", 20, "bold")).pack()
        ttk.Label(header_frame, text="Versione 1.1.0").pack()
        
        # Descrizione
        desc_frame = ttk.Frame(about_frame)
        desc_frame.pack(fill=tk.X, padx=20, pady=10)
        
        description = (
            "Garmin Planner è un'applicazione per la gestione e pianificazione\n"
            "degli allenamenti su Garmin Connect. Permette di importare, esportare\n"
            "e pianificare allenamenti in modo semplice ed efficace."
        )
        ttk.Label(desc_frame, text=description, justify=tk.CENTER).pack()
        
        # Informazioni licenza
        license_frame = ttk.LabelFrame(about_frame, text="Informazioni licenza")
        license_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Ottieni informazioni sulla licenza
        is_valid, message, features, expiry_date, username = self.license_manager.validate_license()
        
        if is_valid:
            license_status = "Licenza attiva"
            status_color = "green"
        else:
            license_status = "Licenza non attiva"
            status_color = "red"
        
        ttk.Label(license_frame, text=license_status, foreground=status_color, font=("Helvetica", 12, "bold")).pack(pady=5)
        
        license_details = ttk.Frame(license_frame)
        license_details.pack(fill=tk.X, padx=10, pady=5)
        
        if is_valid:
            # Username
            ttk.Label(license_details, text="Utente:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
            ttk.Label(license_details, text=username if username else "Non specificato").grid(row=0, column=1, sticky=tk.W, padx=5, pady=2)
            
            # Features
            ttk.Label(license_details, text="Funzionalità:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
            ttk.Label(license_details, text=", ".join(features)).grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
            
            # Expiry date
            ttk.Label(license_details, text="Scadenza:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
            expiry_text = expiry_date if expiry_date else "Licenza perpetua"
            ttk.Label(license_details, text=expiry_text).grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
            
            # Dettagli delle funzionalità disponibili in base al tipo di licenza
            features_details_frame = ttk.LabelFrame(license_frame, text="Funzionalità disponibili")
            features_details_frame.pack(fill=tk.X, padx=10, pady=10)
            
            # Dettaglio delle funzionalità
            feature_text = "Con la tua licenza attuale hai accesso a:\n\n"
            
            # Funzionalità Basic (sempre disponibili se la licenza è valida)
            feature_text += "✓ BASIC:\n"
            feature_text += "   • Importazione/esportazione di allenamenti\n\n"
            
            # Funzionalità Pro
            if "pro" in features:
                feature_text += "✓ PRO:\n"
                feature_text += "   • Pianificazione di base\n"
                feature_text += "   • Pianificazione Excel\n\n"
            else:
                feature_text += "✗ PRO: (non disponibile con la licenza attuale)\n"
                feature_text += "   • Pianificazione di base\n"
                feature_text += "   • Pianificazione Excel\n\n"
            
            # Funzionalità Premium
            if "premium" in features:
                feature_text += "✓ PREMIUM:\n"
                feature_text += "   • Conversione EXCEL a YAML\n"
                feature_text += "   • Editor di allenamenti\n"
            else:
                feature_text += "✗ PREMIUM: (non disponibile con la licenza attuale)\n"
                feature_text += "   • Conversione EXCEL a YAML\n"
                feature_text += "   • Editor di allenamenti\n"
            
            ttk.Label(features_details_frame, text=feature_text, justify=tk.LEFT, wraplength=500).pack(padx=10, pady=10, anchor=tk.W)
            
        else:
            # Per licenza non attiva, mostrare le informazioni dettagliate 
            # che prima erano mostrate nel popup
            info_frame = ttk.Frame(license_details)
            info_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            ttk.Label(info_frame, text=message, wraplength=500).pack(anchor=tk.W, pady=5)
            
            ttk.Label(info_frame, text="Per attivare Garmin Planner, posiziona un file license.dat valido "
                                  "nella cartella dell'applicazione.", wraplength=500).pack(anchor=tk.W, pady=5)
            
            # Aggiungiamo ulteriori informazioni sulla licenza
            info_text = (
                "Questo software è disponibile gratuitamente per uso personale ma richiede "
                "una licenza per l'utilizzo commerciale o in contesti professionali.\n\n"
                "La versione BASIC include:\n"
                "   • Importazione/esportazione di allenamenti\n"
                "La versione PRO include:\n"
                "   • Pianificazione di base e pianificazione Excel\n"
                "La versione PREMIUM include:\n"
                "   • Conversione EXCEL a YAML\n"
                "   • Editor di allenamenti\n\n"
                "Per informazioni sull'acquisto di una licenza, prochilo.francesco@gmail.com"
            )
            
            ttk.Label(info_frame, text=info_text, wraplength=500, justify=tk.LEFT).pack(anchor=tk.W, pady=10)
        
        # Credits e copyright
        footer_frame = ttk.Frame(about_frame)
        footer_frame.pack(fill=tk.X, padx=20, pady=20)
        
        current_year = datetime.now().year
        copyright_text = f"© {current_year} Garmin Planner Team. Tutti i diritti riservati."
        ttk.Label(footer_frame, text=copyright_text).pack()
        
        # Contatti
        contact_text = "Per supporto e informazioni: prochilo.francesco@gmail.com"
        ttk.Label(footer_frame, text=contact_text).pack()


    def create_import_tab(self):
        import_frame = ttk.Frame(self.notebook)
        self.notebook.add(import_frame, text="Importa")
        
        # Variables
        self.import_file = tk.StringVar()
        self.import_replace = tk.BooleanVar(value=False)
        
        # Widgets
        ttk.Label(import_frame, text="Importa allenamenti da file YAML", font=("", 12, "bold")).pack(pady=10)
        
        file_frame = ttk.Frame(import_frame)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(file_frame, text="File allenamenti:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Entry(file_frame, textvariable=self.import_file, width=50).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Sfoglia", command=self.browse_import_file).grid(row=0, column=2, padx=5, pady=5)
        
        options_frame = ttk.Frame(import_frame)
        options_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Checkbutton(options_frame, text="Sostituisci allenamenti esistenti", variable=self.import_replace).pack(side=tk.LEFT, padx=5, pady=5)
        
        ttk.Button(import_frame, text="Importa", command=self.perform_import).pack(pady=10)
        
        # Display available training plans
        tree_buttons_frame = ttk.Frame(import_frame)
        tree_buttons_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Label(tree_buttons_frame, text="Piani di allenamento disponibili:").pack(side=tk.LEFT, padx=5)
        ttk.Button(tree_buttons_frame, text="Aggiorna lista", command=self.load_training_plans).pack(side=tk.RIGHT, padx=5)
        
        # Crea il training_plans_tree
        self.training_plans_tree = ttk.Treeview(import_frame, columns=("plan", "weeks", "type"), show="headings")
        self.training_plans_tree.heading("plan", text="Piano")
        self.training_plans_tree.heading("weeks", text="Settimane")
        self.training_plans_tree.heading("type", text="Tipo")
        self.training_plans_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Populate tree with available training plans
        self.load_training_plans()
        
        # Double-click to select a plan
        self.training_plans_tree.bind("<Double-1>", self.select_training_plan)

    def create_export_tab(self):
        export_frame = ttk.Frame(self.notebook)
        self.notebook.add(export_frame, text="Esporta")
        
        # Variables
        self.export_file = tk.StringVar()
        # Rimossa variabile format: self.export_format = tk.StringVar(value="YAML")
        self.export_clean = tk.BooleanVar(value=True)
        
        # Widgets
        ttk.Label(export_frame, text="Esporta allenamenti", font=("", 12, "bold")).pack(pady=10)
        
        file_frame = ttk.Frame(export_frame)
        file_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(file_frame, text="File di destinazione:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Entry(file_frame, textvariable=self.export_file, width=50).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(file_frame, text="Sfoglia", command=self.browse_export_file).grid(row=0, column=2, padx=5, pady=5)
        
        options_frame = ttk.Frame(export_frame)
        options_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Rimosso il combobox per il formato
        ttk.Checkbutton(options_frame, text="Pulisci dati", variable=self.export_clean).pack(side=tk.LEFT, padx=5, pady=5)
        
        # Bottoni per l'esportazione
        export_buttons_frame = ttk.Frame(export_frame)
        export_buttons_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(export_buttons_frame, text="Esporta Selezionati", command=self.export_selected_workouts).pack(side=tk.LEFT, padx=5)
        ttk.Button(export_buttons_frame, text="Elimina Selezionati", command=self.delete_selected_workouts).pack(side=tk.LEFT, padx=5)
        
        # Lista degli allenamenti disponibili
        ttk.Label(export_frame, text="Allenamenti disponibili su Garmin Connect:").pack(pady=(10, 5))
        
        self.workouts_tree = ttk.Treeview(export_frame, columns=("id", "name"), show="headings")
        self.workouts_tree.heading("id", text="ID")
        self.workouts_tree.heading("name", text="Nome")
        self.workouts_tree.column("id", width=100)
        self.workouts_tree.column("name", width=500)
        self.workouts_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Carica gli allenamenti dalla cache
        self.load_workouts_from_cache()
        
        # Bottone per aggiornare la lista
        ttk.Button(export_frame, text="Aggiorna lista", command=self.refresh_workouts).pack(pady=5)



    def delete_selected_workouts(self):
        """Elimina gli allenamenti selezionati da Garmin Connect"""
        # Verifica se ci sono allenamenti selezionati
        selected_items = self.workouts_tree.selection()
        if not selected_items:
            messagebox.showwarning("Nessuna selezione", "Seleziona almeno un allenamento da eliminare")
            return
        
        # Ottieni gli ID degli allenamenti selezionati
        selected_ids = []
        selected_names = []
        for item in selected_items:
            values = self.workouts_tree.item(item, "values")
            workout_id = values[0]
            workout_name = values[1]
            selected_ids.append(workout_id)
            selected_names.append(workout_name)
        
        # Chiedi conferma all'utente
        if len(selected_ids) == 1:
            message = f"Sei sicuro di voler eliminare l'allenamento '{selected_names[0]}'?"
        else:
            message = f"Sei sicuro di voler eliminare {len(selected_ids)} allenamenti selezionati?"
        
        if not messagebox.askyesno("Conferma eliminazione", message):
            return
        
        self.log(f"Eliminazione di {len(selected_ids)} allenamenti selezionati...")
        
        # Avvia il processo di eliminazione in un thread separato
        threading.Thread(target=lambda: self._do_delete_workouts(selected_ids)).start()


    def _do_delete_workouts(self, workout_ids):
        """Esegue l'eliminazione diretta degli allenamenti tramite il client Garmin"""
        oauth_folder = self.oauth_folder.get()
        
        try:
            # Crea un client Garmin direttamente
            self.log(f"Creazione di un client Garmin Connect con OAuth folder: {oauth_folder}")
            client = GarminClient(oauth_folder)
            
            # Elimina gli allenamenti uno per uno
            deleted_count = 0
            for workout_id in workout_ids:
                self.log(f"Eliminazione diretta dell'allenamento {workout_id}...")
                try:
                    response = client.delete_workout(workout_id)
                    self.log(f"Risposta dall'API: {response}")
                    deleted_count += 1
                except Exception as e:
                    self.log(f"Errore nell'eliminazione dell'allenamento {workout_id}: {str(e)}")
            
            # Mostra un messaggio di conferma
            if deleted_count > 0:
                self.log(f"Eliminati {deleted_count} allenamenti con successo")
                messagebox.showinfo("Successo", f"Eliminati {deleted_count} allenamenti con successo")
                
                # Forza l'aggiornamento della lista degli allenamenti
                # Prima eliminiamo la cache per forzare un nuovo download
                if os.path.exists(WORKOUTS_CACHE_FILE):
                    os.remove(WORKOUTS_CACHE_FILE)
                    self.log("Cache degli allenamenti rimossa per forzare un nuovo download")
                
                # Quindi aggiorniamo la lista
                self.refresh_workouts()
            else:
                self.log("Nessun allenamento è stato eliminato")
                messagebox.showwarning("Attenzione", "Nessun allenamento è stato eliminato")
                
        except Exception as e:
            self.log(f"Errore durante l'eliminazione diretta: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore durante l'eliminazione: {str(e)}")


    def export_selected_workouts(self):
        """Export only selected workouts to a file"""
        # Check if any workouts are selected
        selected_items = self.workouts_tree.selection()
        if not selected_items:
            messagebox.showwarning("Nessuna selezione", "Seleziona almeno un allenamento da esportare")
            return
        
        # Get selected workout IDs
        selected_ids = []
        for item in selected_items:
            workout_id = self.workouts_tree.item(item, "values")[0]
            selected_ids.append(workout_id)
        
        # Richiedi il percorso del file se non è già specificato
        if not self.export_file.get():
            filename = filedialog.asksaveasfilename(
                title="Salva allenamenti selezionati",
                filetypes=[("YAML files", "*.yaml"), ("JSON files", "*.json"), ("All files", "*.*")],
                defaultextension=".yaml"
            )
            if not filename:
                return  # Utente ha annullato
            self.export_file.set(filename)
        
        # Esporta gli allenamenti selezionati
        self.log(f"Esportazione di {len(selected_ids)} allenamenti selezionati...")
        
        # Avvia il processo di esportazione in un thread separato
        threading.Thread(target=lambda: self._do_export_selected(selected_ids)).start()


    def _do_export_selected(self, selected_ids):
        try:
            # Crea un oggetto args simulato
            class Args:
                pass
                    
            args = Args()
            args.oauth_folder = self.oauth_folder.get()
            args.export_file = self.export_file.get()
            args.workout_ids = ",".join(selected_ids)
            args.clean = self.export_clean.get()
            args.name_filter = None
            
            # Determina il formato dall'estensione del file
            file_extension = os.path.splitext(self.export_file.get())[1].lower()
            if file_extension == '.json':
                args.format = "JSON"
            else:
                # Default to YAML for other extensions
                args.format = "YAML"
            
            # Esegui la funzione direttamente
            cmd_export_workouts(args)
            
            self.log(f"Esportazione completata con successo su {self.export_file.get()}")
            messagebox.showinfo("Successo", f"Esportazione completata con successo su {self.export_file.get()}")
        
        except Exception as e:
            self.log(f"Errore durante l'esportazione: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore durante l'esportazione: {str(e)}")

    def perform_unschedule(self):
        """Rimuove gli allenamenti pianificati per il piano selezionato"""
        # Verifica che sia selezionato un piano
        if not self.training_plan.get():
            messagebox.showerror("Errore", "Seleziona un piano di allenamento")
            return
        
        # Chiedi conferma prima di procedere
        if not messagebox.askyesno("Conferma", 
                                 f"Stai per rimuovere tutti gli allenamenti pianificati per il piano '{self.training_plan.get()}'.\n\n"
                                 f"Vuoi procedere?"):
            return
        
        # Esegui la rimozione della pianificazione
        self._do_unschedule()

    def _do_unschedule(self):
        """Esegue la rimozione degli allenamenti pianificati"""
        try:
            # Crea un oggetto args simulato
            class Args:
                pass
                    
            args = Args()
            args.oauth_folder = self.oauth_folder.get()
            args.training_plan = self.training_plan.get()
            args.start_date = None
            args.dry_run = self.schedule_dry_run.get()
            
            # Log della modalità simulazione
            if args.dry_run:
                self.log("Modalità simulazione attivata - nessuna modifica verrà apportata")
            
            # Esegui la funzione direttamente
            cmd_unschedule_workouts(args)
            
            self.log("Rimozione pianificazione completata con successo")
            messagebox.showinfo("Successo", "Rimozione pianificazione completata con successo")
            
            # Aggiorna il calendario
            self.refresh_calendar()
            
        except Exception as e:
            self.log(f"Errore durante la rimozione della pianificazione: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore durante la rimozione della pianificazione: {str(e)}")

    
    def on_excel_day_checkbox_clicked(self, day_index):
        """Gestisce direttamente il click sul checkbox nella tab Excel Tools"""
        # Se non c'è un limite massimo, permetti qualsiasi selezione
        if not hasattr(self, 'excel_max_sessions') or self.excel_max_sessions <= 0:
            return
        
        # Conta quanti giorni sono selezionati DOPO il click
        selected_count = sum(var.get() for var in self.day_selections)
        
        # Se abbiamo superato il limite e questo checkbox è selezionato, deselezionalo
        if selected_count > self.excel_max_sessions and self.day_selections[day_index].get() == 1:
            # Deseleziona il checkbox
            self.day_selections[day_index].set(0)
            
            # Mostra il messaggio informativo
            messagebox.showinfo("Limite raggiunto", 
                             f"Puoi selezionare al massimo {self.excel_max_sessions} giorni per questo piano di allenamento.")

    def create_custom_date_picker(self, parent, date_var):
        """Create a custom date picker with separate widgets for year, month, and day"""
        frame = ttk.Frame(parent)
        
        # Get current date from variable or use today
        try:
            current_date = datetime.strptime(date_var.get(), "%Y-%m-%d")
        except (ValueError, TypeError):
            current_date = datetime.now()
        
        # Year picker
        year_values = list(range(current_date.year, current_date.year + 10))
        year_var = tk.StringVar(value=str(current_date.year))
        ttk.Label(frame, text="Anno:").grid(row=0, column=0, padx=(0, 5))
        year_combo = ttk.Combobox(frame, textvariable=year_var, values=year_values, width=6, state="readonly")
        year_combo.grid(row=0, column=1, padx=(0, 10))
        
        # Month picker
        month_names = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", 
                      "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
        month_var = tk.StringVar(value=month_names[current_date.month-1])
        ttk.Label(frame, text="Mese:").grid(row=0, column=2, padx=(0, 5))
        month_combo = ttk.Combobox(frame, textvariable=month_var, values=month_names, width=10, state="readonly")
        month_combo.grid(row=0, column=3, padx=(0, 10))
        
        # Day picker - days will be adjusted based on month/year
        day_var = tk.StringVar(value=str(current_date.day))
        ttk.Label(frame, text="Giorno:").grid(row=0, column=4, padx=(0, 5))
        day_combo = ttk.Combobox(frame, textvariable=day_var, width=5, state="readonly")
        day_combo.grid(row=0, column=5)
        
        def update_days(*args):
            """Update the available days based on the selected month and year"""
            try:
                year = int(year_var.get())
                month = month_names.index(month_var.get()) + 1
                
                # Get the number of days in the month
                _, num_days = calendar.monthrange(year, month)
                
                # Update the day values
                day_values = [str(i) for i in range(1, num_days + 1)]
                day_combo['values'] = day_values
                
                # Adjust the day if it's out of range
                current_day = int(day_var.get()) if day_var.get() else 1
                if current_day > num_days:
                    day_var.set(str(num_days))
                elif not day_var.get():
                    day_var.set("1")
                    
            except (ValueError, TypeError):
                # Default to 31 days if there's an error
                day_combo['values'] = [str(i) for i in range(1, 32)]
        
        # Initial update of days
        update_days()
        
        # Function to update the main date variable
        def update_date(*args):
            try:
                year = int(year_var.get())
                month = month_names.index(month_var.get()) + 1
                day = int(day_var.get())
                
                # Validate the date
                date_obj = datetime(year, month, day)
                date_var.set(date_obj.strftime("%Y-%m-%d"))
            except (ValueError, TypeError):
                # Invalid date, don't update
                pass
        
        # Bind the update function to the variables
        year_var.trace_add("write", update_date)
        month_var.trace_add("write", lambda *args: [update_days(), update_date()])
        day_var.trace_add("write", update_date)
        
        return frame

    def analyze_training_plan(self):
        """Analizza il piano di allenamento selezionato e mostra informazioni all'utente"""
        try:
            training_plan_id = self.training_plan.get().strip()
            if not training_plan_id:
                self.training_plan_info.set("Nessun piano selezionato")
                self.plan_imported = False
                # Reset max_sessions quando non c'è piano selezionato
                self.max_sessions = 0
                return
                    
            # Aggiungi log di debug approfonditi
            self.log(f"Analisi del piano: '{training_plan_id}'")
            self.log(f"DEBUG: Formato dell'ID piano: '{training_plan_id}'")
            self.log(f"DEBUG: Lunghezza ID piano: {len(training_plan_id)} caratteri")
            self.log(f"DEBUG: Caratteri in esadecimale: {' '.join([hex(ord(c)) for c in training_plan_id])}")

            # Se il piano è AM18W115K, controlla specificamente
            if "AM18W115K" in training_plan_id:
                self.log(f"DEBUG: Il piano contiene 'AM18W115K', verificheremo match esatti")
            
            # Prima verifichiamo se esiste un file YAML corrispondente e lo analizziamo
            yaml_path = self.find_yaml_for_plan(training_plan_id)
            if yaml_path:
                self.log(f"Trovato file YAML per il piano: {yaml_path}")
                self.analyze_yaml_plan(yaml_path)
                return
                    
            # Inizializza i contatori
            total_workouts = 0
            sessions_per_week = {}
            
            # Cerca nella cache degli allenamenti
            if os.path.exists(WORKOUTS_CACHE_FILE):
                with open(WORKOUTS_CACHE_FILE, 'r') as f:
                    workouts = json.load(f)
                    
                    # Stampa info della cache
                    self.log(f"DEBUG: Trovati {len(workouts)} allenamenti nella cache")
                    workout_names = [w.get('workoutName', '') for w in workouts]
                    unique_prefixes = set()
                    for name in workout_names:
                        parts = name.split()
                        if parts:
                            unique_prefixes.add(parts[0])
                    self.log(f"DEBUG: Prefissi unici trovati: {unique_prefixes}")
                    
                    # Debug: stampa i primi allenamenti per vedere il formato
                    for i, workout in enumerate(workouts[:5]):
                        workout_name = workout.get('workoutName', '')
                        self.log(f"Allenamento di debug #{i}: '{workout_name}'")
                    
                    for workout in workouts:
                        workout_name = workout.get('workoutName', '')
                        
                        # Verifica con criteri più flessibili
                        is_match = False
                        
                        # 1. Controllo esatto (il piano è una sottostringa esatta)
                        if training_plan_id in workout_name:
                            is_match = True
                        
                        # 2. Controllo ignorando gli spazi alla fine
                        elif training_plan_id.rstrip() in workout_name:
                            is_match = True
                        
                        # 3. Controllo con pattern WxxSxx dopo il prefisso
                        pattern = re.escape(training_plan_id) + r'\s*W\d\dS\d\d'
                        if re.search(pattern, workout_name, re.IGNORECASE):
                            is_match = True
                            self.log(f"Corrispondenza per pattern regex: '{workout_name}' corrisponde a '{pattern}'")

                        # 4. Controllo più permissivo (rimuove caratteri problematici)
                        clean_plan_id = re.sub(r'[^a-zA-Z0-9]', '', training_plan_id)
                        clean_workout_name = re.sub(r'[^a-zA-Z0-9]', '', workout_name)
                        if clean_plan_id and clean_plan_id in clean_workout_name:
                            is_match = True
                            self.log(f"Corrispondenza con ID pulito: '{clean_workout_name}' contiene '{clean_plan_id}'")
                                
                        # Se c'è una corrispondenza, conta l'allenamento
                        if is_match:
                            total_workouts += 1
                            self.log(f"Trovato allenamento: '{workout_name}'")
                            
                            # Cerca il pattern WxxSxx per estrarre settimana e sessione
                            match = re.search(r'\s*(W\d\d)S(\d\d)\s*', workout_name)
                            if match:
                                week_id = match.group(1)
                                
                                # Incrementa il contatore per questa settimana
                                if week_id not in sessions_per_week:
                                    sessions_per_week[week_id] = 0
                                sessions_per_week[week_id] += 1
            
            # Se abbiamo trovato allenamenti, mostra le informazioni
            if total_workouts > 0:
                weeks = len(sessions_per_week)
                max_sessions = max(sessions_per_week.values()) if sessions_per_week else 0
                
                # Salva il numero massimo di sessioni come attributo della classe
                self.max_sessions = max_sessions
                
                info_text = f"Piano: {total_workouts} allenamenti, {weeks} settimane"
                if sessions_per_week:
                    info_text += f"\nAllenamenti per settimana: "
                    for week, count in sorted(sessions_per_week.items()):
                        info_text += f"{week}={count} "
                        
                self.training_plan_info.set(info_text)
                
                # Non aggiorniamo più la label informativa
                # self.day_info_label.config(text="...")
                    
                # Preseleziona i giorni in base al numero di sessioni
                self.preselect_days(max_sessions)
                    
                # Imposta il flag di piano importato
                self.plan_imported = True
                self.log(f"Piano '{training_plan_id}' trovato con {total_workouts} allenamenti")
            else:
                self.training_plan_info.set(f"Nessun allenamento trovato per '{training_plan_id}'")
                # Non aggiorniamo più la label informativa
                # self.day_info_label.config(text="...")
                
                # Imposta il flag di piano non importato
                self.plan_imported = False
                # Resetta max_sessions se non ci sono allenamenti
                self.max_sessions = 0
                self.log(f"Nessun allenamento trovato per il piano '{training_plan_id}'")
                    
        except Exception as e:
            self.log(f"Errore nell'analisi del piano: {str(e)}")
            self.training_plan_info.set("Errore nell'analisi del piano")
            self.plan_imported = False
            self.max_sessions = 0


    def create_log_tab(self):
        """Crea la tab Log (aggiornata con l'opzione per il livello di log)"""
        log_frame = ttk.Frame(self.notebook)
        self.notebook.add(log_frame, text="Log")
        
        # Aggiungiamo le impostazioni del livello di log qui
        log_settings_frame = ttk.Frame(log_frame)
        log_settings_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(log_settings_frame, text="Livello di log:").pack(side=tk.LEFT, padx=5)
        log_level_combo = ttk.Combobox(log_settings_frame, textvariable=self.log_level, 
                                      values=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
        log_level_combo.pack(side=tk.LEFT, padx=5)
        log_level_combo.current(1)  # Default to INFO
        
        # Log text widget
        self.log_text = tk.Text(log_frame, wrap=tk.WORD, height=20)
        scroll = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Add handler to log to the text widget
        self.log_handler = TextHandler(self.log_text)
        logging.getLogger().addHandler(self.log_handler)
        
        # Clear button
        button_frame = ttk.Frame(log_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(button_frame, text="Pulisci log", command=self.clear_log).pack(side=tk.RIGHT)


    def create_excel_tools_tab(self):
        """Crea la tab Pianificazione con flusso guidato a step"""
        excel_frame = ttk.Frame(self.notebook)
        self.notebook.add(excel_frame, text="Pianificazione")
        
        # Variabili
        self.excel_input_file = tk.StringVar()
        self.yaml_output_file = tk.StringVar()
        self.athlete_name_var = tk.StringVar()
        
        # Default alla data corrente + 30 giorni per la data della gara
        default_race_date = (datetime.now() + timedelta(days=30)).strftime("%Y-%m-%d")
        self.race_day = tk.StringVar(value=default_race_date)
        
        # Variabili per i giorni della settimana
        self.day_selections = [tk.IntVar() for _ in range(7)]
        
        # Inizializza la variabile per il numero di sessioni massime
        self.excel_max_sessions = 0
        
        # Variabile per il monitoraggio dello step corrente
        self.current_step = tk.IntVar(value=1)
        
        # Creiamo un contenitore per gli step
        steps_container = ttk.Frame(excel_frame)
        steps_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Creiamo i frame per ogni step
        self.step_frames = []
        for i in range(1, 5):  # 4 step totali
            step_frame = ttk.LabelFrame(steps_container, text=f"Step {i}")
            step_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            self.step_frames.append(step_frame)
            
            # Nascondiamo tutti gli step tranne il primo
            if i > 1:
                step_frame.pack_forget()
        
        # Aggiungi contenuti specifici per ogni step
        self.create_step1_content(self.step_frames[0])  # Pianificazione Allenamenti
        self.create_step2_content(self.step_frames[1])  # Conversione
        self.create_step3_content(self.step_frames[2])  # Informazioni Piano
        self.create_step4_content(self.step_frames[3])  # Pianificazione Calendario
        
        # Frame per i pulsanti di navigazione
        nav_frame = ttk.Frame(excel_frame)
        nav_frame.pack(fill=tk.X, padx=10, pady=10)
        
        self.prev_button = ttk.Button(nav_frame, text="Precedente", command=self.go_to_prev_step)
        self.prev_button.pack(side=tk.LEFT, padx=5)
        self.prev_button.config(state=tk.DISABLED)  # Disabilitato all'inizio
        
        self.next_button = ttk.Button(nav_frame, text="Successivo", command=self.go_to_next_step)
        self.next_button.pack(side=tk.RIGHT, padx=5)


    def create_step1_content(self, parent_frame):
        """Step 1: Pianificazione Allenamenti (creazione file Excel)"""
        # Titolo
        ttk.Label(parent_frame, text="Step 1: Pianificazione Allenamenti", font=("", 12, "bold")).pack(pady=10)
        
        # Descrizione
        desc_text = "In questo step, inserisci i dati dell'atleta, la data della gara e seleziona i giorni preferiti per gli allenamenti."
        ttk.Label(parent_frame, text=desc_text, wraplength=600).pack(pady=5)
        
        # Frame principale con elementi di input
        input_frame = ttk.Frame(parent_frame)
        input_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Nome Atleta (con asterisco rosso per indicare campo obbligatorio)
        athlete_frame = ttk.Frame(input_frame)
        athlete_frame.pack(fill=tk.X, pady=5)
        
        athlete_label_frame = ttk.Frame(athlete_frame)
        athlete_label_frame.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(athlete_label_frame, text="Nome dell'atleta:").pack(side=tk.LEFT)
        ttk.Label(athlete_label_frame, text="*", foreground="red", font=("", 12, "bold")).pack(side=tk.LEFT)
        
        ttk.Entry(athlete_frame, textvariable=self.athlete_name_var, width=30).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # Data della gara
        race_date_frame = ttk.Frame(input_frame)
        race_date_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(race_date_frame, text="Giorno della gara:").pack(side=tk.LEFT, padx=5)
        race_date_picker = self.create_custom_date_picker(race_date_frame, self.race_day)
        race_date_picker.pack(side=tk.LEFT, padx=5)
        
        # Giorni selezionati
        days_frame = ttk.LabelFrame(input_frame, text="Giorni preferiti per gli allenamenti")
        days_frame.pack(fill=tk.X, pady=10)
        
        days_container = ttk.Frame(days_frame)
        days_container.pack(fill=tk.X, padx=10, pady=5)
        
        day_names = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
        for i, day_name in enumerate(day_names):
            var = self.day_selections[i]
            cb = ttk.Checkbutton(days_container, text=day_name, variable=var,
                              command=lambda i=i: self.on_excel_day_checkbox_clicked(i))
            cb.grid(row=0, column=i, padx=5)
        
        # Output file
        output_frame = ttk.Frame(input_frame)
        output_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(output_frame, text="File Excel di output:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Entry(output_frame, textvariable=self.excel_input_file, width=50).grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        ttk.Button(output_frame, text="Sfoglia", command=self.browse_excel_output_path).grid(row=0, column=2, padx=5, pady=5)
        
        output_frame.columnconfigure(1, weight=1)
        
        # Bottone per creare il file Excel
        ttk.Button(input_frame, text="Crea Piano Allenamenti", 
                   command=self.create_training_plan_excel).pack(pady=10)

    def create_step2_content(self, parent_frame):
        """Step 2: Conversione da Excel a YAML"""
        # Titolo
        ttk.Label(parent_frame, text="Step 2: Conversione Excel → YAML", font=("", 12, "bold")).pack(pady=10)
        
        # Descrizione
        desc_text = "In questo step, converti il file Excel in un file YAML compatibile con Garmin Connect."
        ttk.Label(parent_frame, text=desc_text, wraplength=600).pack(pady=5)
        
        # Frame principale con elementi di input
        input_frame = ttk.Frame(parent_frame)
        input_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # File Excel di input (recuperato dallo step 1)
        excel_frame = ttk.Frame(input_frame)
        excel_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(excel_frame, text="File Excel:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Entry(excel_frame, textvariable=self.excel_input_file, width=50, state="readonly").grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        excel_frame.columnconfigure(1, weight=1)
        
        # File YAML di output
        yaml_frame = ttk.Frame(input_frame)
        yaml_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(yaml_frame, text="File YAML:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        ttk.Entry(yaml_frame, textvariable=self.yaml_output_file, width=50).grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        ttk.Button(yaml_frame, text="Sfoglia", command=self.browse_yaml_output).grid(row=0, column=2, padx=5, pady=5)
        
        yaml_frame.columnconfigure(1, weight=1)
        
        # Bottone per convertire
        ttk.Button(input_frame, text="Converti Excel → YAML", 
                   command=self.convert_excel_to_yaml).pack(pady=10)

    def create_step3_content(self, parent_frame):
        """Step 3: Visualizzazione informazioni sul piano"""
        # Titolo
        ttk.Label(parent_frame, text="Step 3: Informazioni Piano", font=("", 12, "bold")).pack(pady=10)
        
        # Descrizione
        desc_text = "In questo step, puoi verificare le informazioni sul piano creato."
        ttk.Label(parent_frame, text=desc_text, wraplength=600).pack(pady=5)
        
        # Frame principale con elementi informativi
        info_frame = ttk.Frame(parent_frame)
        info_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Informazioni sul piano
        self.plan_info_frame = ttk.LabelFrame(info_frame, text="Piano di allenamento")
        self.plan_info_frame.pack(fill=tk.X, pady=5)
        
        # Piano ID
        plan_id_frame = ttk.Frame(self.plan_info_frame)
        plan_id_frame.pack(fill=tk.X, pady=5, padx=10)
        
        ttk.Label(plan_id_frame, text="ID Piano:").pack(side=tk.LEFT, padx=5)
        
        # Assicurati che training_plan sia già inizializzata nel costruttore della classe
        ttk.Entry(plan_id_frame, textvariable=self.training_plan, width=30, state="readonly").pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # File YAML
        yaml_frame = ttk.Frame(self.plan_info_frame)
        yaml_frame.pack(fill=tk.X, pady=5, padx=10)
        
        ttk.Label(yaml_frame, text="File YAML:").pack(side=tk.LEFT, padx=5)
        
        # Mostriamo il percorso del file YAML
        self.step3_yaml_path = tk.StringVar()
        ttk.Entry(yaml_frame, textvariable=self.step3_yaml_path, width=50, state="readonly").pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # Statistiche del piano
        stats_frame = ttk.LabelFrame(info_frame, text="Statistiche")
        stats_frame.pack(fill=tk.X, pady=10)
        
        # Usiamo una variabile per le statistiche
        self.plan_stats_text = tk.StringVar(value="Nessuna informazione disponibile")
        ttk.Label(stats_frame, textvariable=self.plan_stats_text, justify=tk.LEFT, wraplength=600).pack(padx=10, pady=10, anchor=tk.W)
        
        # Bottone per analizzare il piano
        ttk.Button(info_frame, text="Analizza Piano", 
                   command=self.analyze_step3_plan).pack(pady=10)

    def create_step4_content(self, parent_frame):
        """Step 4: Pianificazione Calendar"""
        # Titolo
        ttk.Label(parent_frame, text="Step 4: Pianificazione Calendar", font=("", 12, "bold")).pack(pady=10)
        
        # Descrizione
        desc_text = "In questo step, puoi simulare e pianificare gli allenamenti nel calendario di Garmin Connect."
        ttk.Label(parent_frame, text=desc_text, wraplength=600).pack(pady=5)
        
        # Frame principale con elementi di azione
        action_frame = ttk.Frame(parent_frame)
        action_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Opzioni di pianificazione
        options_frame = ttk.Frame(action_frame)
        options_frame.pack(fill=tk.X, pady=5)
            
        # Bottoni per azioni
        buttons_frame = ttk.Frame(action_frame)
        buttons_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(buttons_frame, text="Simula Pianificazione", 
                   command=self.simulate_schedule_step4).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(buttons_frame, text="Pianifica Allenamenti", 
                   command=self.perform_schedule_step4).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(buttons_frame, text="Rimuovi Pianificazione", 
                   command=self.perform_unschedule_step4).pack(side=tk.LEFT, padx=5)
        
        # Vista calendario
        calendar_frame = ttk.LabelFrame(action_frame, text="Calendario Allenamenti")
        calendar_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Tree view per visualizzare il calendario
        self.step4_calendar_tree = ttk.Treeview(calendar_frame, columns=("date", "workout"), show="headings")
        self.step4_calendar_tree.heading("date", text="Data")
        self.step4_calendar_tree.heading("workout", text="Allenamento")
        self.step4_calendar_tree.column("date", width=100)
        self.step4_calendar_tree.column("workout", width=500)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(calendar_frame, orient="vertical", command=self.step4_calendar_tree.yview)
        self.step4_calendar_tree.configure(yscrollcommand=scrollbar.set)
        
        # Layout
        self.step4_calendar_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
        
        # Bottone per aggiornare il calendario
        ttk.Button(action_frame, text="Aggiorna Calendario", 
                   command=self.refresh_step4_calendar).pack(pady=5)

    def go_to_next_step(self):
        """Passa allo step successivo"""
        current = self.current_step.get()
        if current < len(self.step_frames):
            # Nascondi step corrente
            self.step_frames[current-1].pack_forget()
            
            # Passa al prossimo step
            self.current_step.set(current + 1)
            
            # Mostra il nuovo step
            self.step_frames[current].pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Aggiorna stato pulsanti
            self.prev_button.config(state=tk.NORMAL)
            if current + 1 == len(self.step_frames):
                self.next_button.config(state=tk.DISABLED)
            
            # Azioni specifiche per ogni step
            if current + 1 == 2:  # Passando allo step 2
                # Aggiorna il file YAML di output in base al file Excel
                if self.excel_input_file.get():
                    yaml_path = os.path.splitext(self.excel_input_file.get())[0] + ".yaml"
                    self.yaml_output_file.set(yaml_path)
            
            elif current + 1 == 3:  # Passando allo step 3
                # Aggiorna automaticamente le informazioni sul piano
                self.analyze_step3_plan()
            
            elif current + 1 == 4:  # Passando allo step 4
                # Aggiorna calendario
                self.refresh_step4_calendar()

    def go_to_prev_step(self):
        """Torna allo step precedente"""
        current = self.current_step.get()
        if current > 1:
            # Nascondi step corrente
            self.step_frames[current-1].pack_forget()
            
            # Torna allo step precedente
            self.current_step.set(current - 1)
            
            # Mostra il nuovo step
            self.step_frames[current-2].pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            # Aggiorna stato pulsanti
            self.next_button.config(state=tk.NORMAL)
            if current - 1 == 1:
                self.prev_button.config(state=tk.DISABLED)

    def browse_excel_output_path(self):
        """Seleziona il percorso per il file Excel di output"""
        filename = filedialog.asksaveasfilename(
            title="Salva file Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            defaultextension=".xlsx"
        )
        if filename:
            self.excel_input_file.set(filename)
            self.log(f"File Excel di output impostato: {filename}")
            
            # Aggiorna anche il file YAML di output
            yaml_path = os.path.splitext(filename)[0] + ".yaml"
            self.yaml_output_file.set(yaml_path)


    def create_training_plan_excel(self):
        """Crea il file Excel del piano di allenamento con tutti i dettagli"""
        try:
            # Verifica campi obbligatori
            athlete_name = self.athlete_name_var.get().strip()
            if not athlete_name:
                messagebox.showerror("Errore", "Il nome dell'atleta è obbligatorio!")
                return
            
            # Verifica data della gara
            race_day_str = self.race_day.get()
            try:
                race_date = datetime.strptime(race_day_str, "%Y-%m-%d").date()
                today = datetime.today().date()
                if race_date < today:
                    messagebox.showerror("Errore", "La data della gara deve essere nel futuro!")
                    return
            except ValueError:
                messagebox.showerror("Errore", "Formato data non valido. Usa YYYY-MM-DD")
                return
            
            # Verifica giorni selezionati
            selected_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
            if not selected_days:
                messagebox.showerror("Errore", "Seleziona almeno un giorno della settimana!")
                return
            
            # Chiedi file di output se non specificato
            if not self.excel_input_file.get():
                filename = filedialog.asksaveasfilename(
                    title="Salva file Excel",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    defaultextension=".xlsx"
                )
                if not filename:
                    return  # Utente ha annullato
                self.excel_input_file.set(filename)
            
            # Crea un file Excel personalizzato con il numero corretto di sessioni
            excel_file = self.create_custom_excel_plan(self.excel_input_file.get(), len(selected_days))
            self.log(f"File Excel creato: {excel_file}")
            
            # Aggiorna il file Excel con i dettagli dell'atleta, la data della gara e i giorni selezionati
            self._schedule_excel_workouts_from_race_day(excel_file, race_date, selected_days, athlete_name)
            
            # Mostra messaggio di successo
            messagebox.showinfo("Successo", 
                              f"File Excel creato con successo!\n\n"
                              f"File: {excel_file}\n\n"
                              f"Ora puoi procedere con la conversione YAML.")
            
            # Passa automaticamente allo step successivo
            self.go_to_next_step()
            
        except Exception as e:
            self.log(f"Errore nella creazione del file Excel: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def create_custom_excel_plan(self, output_file, sessions_per_week):
        """
        Crea un file Excel personalizzato con il numero corretto di sessioni per settimana.
        
        Args:
            output_file: Percorso del file Excel di output
            sessions_per_week: Numero di sessioni per settimana (giorni selezionati)
            
        Returns:
            Percorso del file Excel creato
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            import random
            import string
            
            self.log(f"Creazione piano Excel con {sessions_per_week} sessioni per settimana")
            
            wb = openpyxl.Workbook()
            
            # Define a thin border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
            prefix = f"MYRUN_{random_suffix}_"
            
            # Config sheet
            config_sheet = wb.active
            config_sheet.title = 'Config'
            
            # Config sheet headers
            config_sheet['A1'] = 'Parameter'
            config_sheet['B1'] = 'Value'
            config_sheet['C1'] = 'Slower'
            config_sheet['D1'] = 'HR Up'
            config_sheet['E1'] = 'HR Down'
            
            # Config sheet values
            config_sheet['A2'] = 'name_prefix'
            config_sheet['B2'] = prefix
            
            config_sheet['A3'] = 'margins'
            config_sheet['B3'] = '0:03'  # faster
            config_sheet['C3'] = '0:03'  # slower
            config_sheet['D3'] = 5       # hr_up
            config_sheet['E3'] = 5       # hr_down
            
            # Aggiungi la race_day nel foglio Config
            config_sheet['A4'] = 'race_day'
            config_sheet['B4'] = datetime.now().strftime("%Y-%m-%d")
            
            # Aggiungi i giorni preferiti nel foglio Config
            selected_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
            config_sheet['A5'] = 'preferred_days'
            config_sheet['B5'] = str(selected_days)
            
            # Format header
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            for col in ['A', 'B', 'C', 'D', 'E']:
                config_sheet[f'{col}1'].font = Font(bold=True)
                config_sheet[f'{col}1'].fill = header_fill
            
            # Paces sheet (Z1-Z5 zones)
            paces_sheet = wb.create_sheet(title='Paces')
            
            paces_sheet['A1'] = 'Name'
            paces_sheet['B1'] = 'Value'
            
            paces_sheet['A2'] = 'Z1'
            paces_sheet['B2'] = '6:30'
            
            paces_sheet['A3'] = 'Z2'
            paces_sheet['B3'] = '6:20'
            
            paces_sheet['A4'] = 'Z3'
            paces_sheet['B4'] = '6:00'
            
            paces_sheet['A5'] = 'Z4'
            paces_sheet['B5'] = '5:20'
            
            paces_sheet['A6'] = 'Z5'
            paces_sheet['B6'] = '4:50'
            
            # Format header
            for col in ['A', 'B']:
                paces_sheet[f'{col}1'].font = Font(bold=True)
                paces_sheet[f'{col}1'].fill = header_fill
            
            # HeartRates sheet (Z1-Z5 zones)
            hr_sheet = wb.create_sheet(title='HeartRates')
            
            hr_sheet['A1'] = 'Name'
            hr_sheet['B1'] = 'Value'
            
            # Example of using max_hr with percentages
            hr_sheet['A2'] = 'max_hr'
            hr_sheet['B2'] = 198  # Use an integer instead of a string
            
            hr_sheet['A3'] = 'Z1'
            hr_sheet['B3'] = '62-76% max_hr'
            
            hr_sheet['A4'] = 'Z2'
            hr_sheet['B4'] = '76-85% max_hr'
            
            hr_sheet['A5'] = 'Z3'
            hr_sheet['B5'] = '85-91% max_hr'
            
            hr_sheet['A6'] = 'Z4'
            hr_sheet['B6'] = '91-95% max_hr'
            
            hr_sheet['A7'] = 'Z5'
            hr_sheet['B7'] = '95-100% max_hr'
            
            # Format header
            for col in ['A', 'B']:
                hr_sheet[f'{col}1'].font = Font(bold=True)
                hr_sheet[f'{col}1'].fill = header_fill
            
            # Single Workouts sheet for all workouts
            workouts_sheet = wb.create_sheet(title='Workouts')
            
            # Add a row for the athlete's name
            # Create a merged cell for the athlete's name
            workouts_sheet.merge_cells('A1:E1')
            athlete_cell = workouts_sheet['A1']
            athlete_cell.value = "Atleta: "  # Prepared to be filled in
            athlete_cell.alignment = Alignment(horizontal='center', vertical='center')
            athlete_cell.font = Font(size=12, bold=True)
            # Add border to the athlete cell
            athlete_cell.border = thin_border

            # Headers in row 2
            workouts_sheet['A2'] = 'Week'
            workouts_sheet['B2'] = 'Date'
            workouts_sheet['C2'] = 'Session'
            workouts_sheet['D2'] = 'Description'
            workouts_sheet['E2'] = 'Steps'

            # Format header
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = workouts_sheet[f'{col}2']
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.border = thin_border  # Add border to all header cells
            
            # Definisci tipi di allenamento basati sul numero di sessioni per settimana
            if sessions_per_week == 1:
                # Piano minimo: solo una sessione lunga a settimana
                workout_types = [
                    "Long slow run"
                ]
            elif sessions_per_week == 2:
                # Piano base: una sessione lunga e una di intervalli
                workout_types = [
                    "Interval training",
                    "Long slow run"
                ]
            elif sessions_per_week == 3:
                # Piano intermedio: più varietà
                workout_types = [
                    "Easy run",
                    "Interval training",
                    "Long slow run"
                ]
            elif sessions_per_week == 4:
                # Piano avanzato: aggiunta di un allenamento di recupero
                workout_types = [
                    "Recovery run",
                    "Tempo run",
                    "Interval training",
                    "Long slow run"
                ]
            elif sessions_per_week == 5:
                # Piano molto avanzato
                workout_types = [
                    "Easy run",
                    "Recovery run", 
                    "Tempo run",
                    "Interval training",
                    "Long slow run"
                ]
            elif sessions_per_week >= 6:
                # Piano professionale
                workout_types = [
                    "Easy run",
                    "Recovery run", 
                    "Tempo run",
                    "Interval training",
                    "Hill repeats",
                    "Long slow run"
                ]
                # Aggiungi sessioni aggiuntive se necessario
                while len(workout_types) < sessions_per_week:
                    workout_types.append("Extra session")
            
            # Definisci dettagli degli allenamenti
            workout_details = {
                "Easy run": {
                    "steps": "warmup: 10min @ Z1\ninterval: 30min @ Z2\ncooldown: 5min @ Z1"
                },
                "Recovery run": {
                    "steps": "interval: 30min @ Z1"
                },
                "Tempo run": {
                    "steps": "warmup: 15min @ Z1\ninterval: 20min @ Z4\ncooldown: 10min @ Z1"
                },
                "Interval training": {
                    "steps": "warmup: 15min @ Z1\nrepeat 5:\n  interval: 400m @ Z5\n  recovery: 2min @ Z1\ncooldown: 10min @ Z1"
                },
                "Hill repeats": {
                    "steps": "warmup: 15min @ Z1\nrepeat 6:\n  interval: 1min @ Z5\n  recovery: 2min @ Z1\ncooldown: 10min @ Z1"
                },
                "Long slow run": {
                    "steps": "warmup: 10min @ Z1\ninterval: 45min @ Z2\ncooldown: 5min @ Z1"
                },
                "Extra session": {
                    "steps": "warmup: 10min @ Z1\ninterval: 20min @ Z2\ncooldown: 5min @ Z1"
                }
            }
            
            # Definisci durate progressive per ciascuna settimana del piano
            # Per semplicità, generiamo un piano di 8 settimane
            weeks = 3
            
            # Define alternating colors for weeks
            week_colors = [
                "FFF2CC",  # Light yellow
                "DAEEF3",  # Light blue
                "E2EFDA",  # Light green
                "FCE4D6",  # Light orange
                "EAD1DC",  # Light pink
                "D9D9D9",  # Light gray
            ]
            
            # Aggiungi allenamenti al foglio
            row_index = 3  # Start from row 3 (after header and athlete row)
            
            for week in range(1, weeks + 1):
                # Determina il colore per questa settimana
                color_index = (week - 1) % len(week_colors)
                row_fill = PatternFill(start_color=week_colors[color_index], 
                                      end_color=week_colors[color_index], 
                                      fill_type="solid")
                
                # Aggiungi ciascuna sessione per questa settimana
                for session in range(1, sessions_per_week + 1):
                    # Usa l'indice corretto in base al numero di sessioni
                    workout_type = workout_types[(session - 1) % len(workout_types)]
                    
                    # Se è l'ultima settimana e la prima sessione, rendiamola la gara
                    if week == weeks and session == 1:
                        workout_type = "Race day"
                        workout_details["Race day"] = {
                            "steps": "warmup: 10min @ Z2\ninterval: 3000m in 13:48\ncooldown: 10min @ Z1"
                        }
                    
                    # Crea la riga per questa sessione
                    workouts_sheet[f'A{row_index}'] = week
                    workouts_sheet[f'B{row_index}'] = None  # La data verrà impostata successivamente
                    workouts_sheet[f'C{row_index}'] = session
                    workouts_sheet[f'D{row_index}'] = workout_type
                    workouts_sheet[f'E{row_index}'] = workout_details[workout_type]["steps"]
                    
                    # Applica lo stile a tutte le celle della riga
                    for col in ['A', 'B', 'C', 'D', 'E']:
                        cell = workouts_sheet[f'{col}{row_index}']
                        cell.fill = row_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(wrapText=True, vertical='top')
                    
                    # Calcola altezza appropriata per il contenuto
                    steps_text = workout_details[workout_type]["steps"]
                    num_lines = 1 + steps_text.count('\n') + steps_text.count(';')
                    
                    # Considera indentazione per i ripetuti
                    if 'repeat' in steps_text and '\n' in steps_text:
                        # Conta le righe indentate dopo un repeat
                        lines_after_repeat = steps_text.split('repeat')[1].count('\n')
                        if lines_after_repeat > 0:
                            num_lines += lines_after_repeat - 1  # -1 perché la riga con repeat è già contata
                    
                    # Imposta altezza minima più altezza per ogni riga di testo (circa 15 punti per riga)
                    row_height = max(20, 15 * num_lines)  # Altezza minima aumentata
                    workouts_sheet.row_dimensions[row_index].height = row_height
                    
                    # Passa alla prossima riga
                    row_index += 1
            
            # Set column widths
            workouts_sheet.column_dimensions['A'].width = 10  # Week
            workouts_sheet.column_dimensions['B'].width = 15  # Date
            workouts_sheet.column_dimensions['C'].width = 10  # Session
            workouts_sheet.column_dimensions['D'].width = 25  # Description
            workouts_sheet.column_dimensions['E'].width = 60  # Steps
            
            # Auto-adjust column widths in Config, Paces, and HR sheets
            self.auto_adjust_column_widths(config_sheet)
            self.auto_adjust_column_widths(paces_sheet)
            self.auto_adjust_column_widths(hr_sheet)
            
            # Save the file
            wb.save(output_file)
            self.log(f"File Excel creato con {sessions_per_week} sessioni per settimana")
            return output_file
            
        except Exception as e:
            self.log(f"Errore nella creazione del file Excel: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def auto_adjust_column_widths(self, worksheet):
        """Adatta automaticamente le larghezze delle colonne in base al contenuto"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            adjusted_width = max(max_length + 2, 8)  # Aggiungi spazio extra
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 60)  # Limita a 60 per evitare colonne troppo larghe


    def analyze_step3_plan(self):
        """Analizza il piano YAML per lo step 3"""
        try:
            yaml_file = self.yaml_output_file.get()
            if not yaml_file or not os.path.exists(yaml_file):
                messagebox.showerror("Errore", "File YAML non trovato. Completa prima lo step 2.")
                return
            
            self.log(f"Analisi del piano YAML: {yaml_file}")
            
            # Leggi il file YAML
            with open(yaml_file, 'r') as f:
                plan_data = yaml.safe_load(f)
            
            # Estrai l'ID del piano
            plan_id = ""
            if 'config' in plan_data and 'name_prefix' in plan_data['config']:
                plan_id = plan_data['config']['name_prefix'].strip()
            
            # Estrai informazioni
            total_workouts = 0
            sessions_per_week = {}
            workouts_with_dates = 0
            
            # Rimuovi la configurazione per contare solo gli allenamenti
            plan_data_copy = plan_data.copy()
            if 'config' in plan_data_copy:
                plan_data_copy.pop('config')
            
            # Conta allenamenti e sessioni per settimana
            for workout_name, steps in plan_data_copy.items():
                total_workouts += 1
                
                # Cerca la data nell'allenamento
                if steps and isinstance(steps, list) and len(steps) > 0:
                    if isinstance(steps[0], dict) and 'date' in steps[0]:
                        workouts_with_dates += 1
                
                # Estrai settimana e sessione
                match = re.search(r'(W\d\d)S(\d\d)', workout_name)
                if match:
                    week_id = match.group(1)
                    if week_id not in sessions_per_week:
                        sessions_per_week[week_id] = 0
                    sessions_per_week[week_id] += 1
            
            # Calcola settimane e sessioni massime
            num_weeks = len(sessions_per_week)
            max_sessions = max(sessions_per_week.values()) if sessions_per_week else 0
            
            # Componi testo informativo
            stats_text = f"Piano: {total_workouts} allenamenti, {num_weeks} settimane\n"
            if sessions_per_week:
                stats_text += f"Allenamenti per settimana: "
                for week, count in sorted(sessions_per_week.items()):
                    stats_text += f"{week}={count} "
                stats_text += f"\n"
            
            stats_text += f"Allenamenti con date: {workouts_with_dates}/{total_workouts}\n"
            
            if 'config' in plan_data and 'race_day' in plan_data['config']:
                stats_text += f"Data gara: {plan_data['config']['race_day']}\n"
            
            if 'config' in plan_data and 'athlete_name' in plan_data['config']:
                stats_text += f"Atleta: {plan_data['config']['athlete_name']}"
            
            # Aggiorna interfaccia
            self.training_plan  .set(plan_id)
            self.step3_yaml_path.set(yaml_file)
            self.plan_stats_text.set(stats_text)
            
            # Salva il percorso YAML e il plan ID per lo step 4
            self.current_yaml_path = yaml_file
            self.training_plan.set(plan_id)  # Questo serve per lo step 4
            
            self.log("Analisi del piano completata")
            
        except Exception as e:
            self.log(f"Errore nell'analisi del piano: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def simulate_schedule_step4(self):
        """Simula la pianificazione per lo step 4"""
        try:
            # Verifica che siano presenti le informazioni necessarie
            if not self.training_plan.get():
                messagebox.showerror("Errore", "ID Piano non impostato. Completa prima lo step 3.")
                return
                        
            # Usa la funzione di pianificazione esistente
            selected_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
            
            if not selected_days:
                messagebox.showerror("Errore", "Nessun giorno selezionato per la pianificazione.")
                return
            
            self.log("Avvio simulazione pianificazione...")
            
            # Il flag dry_run viene gestito internamente
            self._simulate_schedule(selected_days)
            
        except Exception as e:
            self.log(f"Errore nella simulazione: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def perform_schedule_step4(self):
        """Esegue la pianificazione effettiva per lo step 4"""
        try:
            # Verifica che siano presenti le informazioni necessarie
            if not self.training_plan.get():
                messagebox.showerror("Errore", "ID Piano non impostato. Completa prima lo step 3.")
                return
                        
            # Chiedi conferma
            if not messagebox.askyesno("Conferma", 
                                     "Stai per pianificare gli allenamenti in Garmin Connect.\n\n"
                                     "Vuoi procedere?"):
                return
            
            # Usa la funzione di pianificazione esistente
            selected_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
            
            if not selected_days:
                messagebox.showerror("Errore", "Nessun giorno selezionato per la pianificazione.")
                return
            
            # Verifica login a Garmin Connect
            if not self.verify_garmin_login():
                return
            
            self.log("Avvio pianificazione allenamenti...")
            
            # Invoca la funzione di pianificazione
            self.perform_schedule()
            
            # Aggiorna calendario
            self.refresh_step4_calendar()
            
        except Exception as e:
            self.log(f"Errore nella pianificazione: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def perform_unschedule_step4(self):
        """Rimuove gli allenamenti pianificati per lo step 4"""
        try:
            # Verifica che siano presenti le informazioni necessarie
            if not self.training_plan.get():
                messagebox.showerror("Errore", "ID Piano non impostato. Completa prima lo step 3.")
                return
            
            # Chiedi conferma
            if not messagebox.askyesno("Conferma", 
                                     f"Stai per rimuovere tutti gli allenamenti pianificati per '{self.training_plan.get()}'.\n\n"
                                     f"Vuoi procedere?"):
                return
            
            # Verifica login a Garmin Connect
            if not self.verify_garmin_login():
                return
            
            # Usa la funzione esistente
            self.perform_unschedule()
            
            # Aggiorna calendario
            self.refresh_step4_calendar()
            
        except Exception as e:
            self.log(f"Errore nella rimozione: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def refresh_step4_calendar(self):
        """Aggiorna la visualizzazione del calendario nello step 4"""
        try:
            # Usa la stessa logica di refresh_calendar, ma aggiorna il tree view dello step 4
            self.log("Aggiornamento calendario...")
            
            # Crea un oggetto args simulato
            class Args:
                pass
                    
            args = Args()
            args.oauth_folder = self.oauth_folder.get()
            args.start_date = None
            args.end_date = None
            args.date_range = None
            
            # Aggiungi il filtro per nome se specificato
            if self.training_plan.get():
                args.name_filter = self.training_plan.get()
            else:
                args.name_filter = None
            
            # Ottieni gli allenamenti pianificati
            from planner.manage import get_scheduled
            calendar_data = get_scheduled(args)
            
            # Pulisci la tabella
            for item in self.step4_calendar_tree.get_children():
                self.step4_calendar_tree.delete(item)
            
            # Aggiungi alla tabella
            for item in calendar_data:
                date = item.get('date', 'N/A')
                workout_name = item.get('title', 'Senza nome')
                self.step4_calendar_tree.insert("", "end", values=(date, workout_name))
            
            self.log(f"Trovati {len(calendar_data)} allenamenti pianificati")
            
        except Exception as e:
            self.log(f"Errore nell'aggiornamento del calendario: {str(e)}")


    def verify_garmin_login(self):
        """Verifica che l'utente sia loggato a Garmin Connect"""
        # Verifica se esistono già file di autenticazione
        oauth_folder = self.oauth_folder.get()
        oauth_files_exist = False
        
        if os.path.exists(oauth_folder):
            # Controlla la presenza di file oauth
            oauth_files = [f for f in os.listdir(oauth_folder) if f.startswith('oauth') and f.endswith('.json')]
            if oauth_files:
                oauth_files_exist = True
        
        if not oauth_files_exist:
            messagebox.showerror("Login Richiesto", 
                              "Per pianificare gli allenamenti, è necessario effettuare il login a Garmin Connect.\n\n"
                              "Vai alla tab Login e completa l'accesso.")
            return False
        
        return True

    def schedule_excel_workouts(self):
        """Pianifica le date degli allenamenti nel file Excel"""
        if not LicenseManager.get_instance().check_feature_access("pro"):
            return

        try:
            # Verifica che sia stato selezionato un file Excel
            excel_file = self.excel_input_file.get()
            if not excel_file:
                messagebox.showerror("Errore", "Seleziona prima un file Excel.")
                return
                
            if not os.path.exists(excel_file):
                messagebox.showerror("Errore", f"Il file Excel non esiste: {excel_file}")
                return
                
            # Verifica che sia stata selezionata una data di gara
            race_day_str = self.race_day.get()
            try:
                race_date = datetime.strptime(race_day_str, "%Y-%m-%d").date()
            except ValueError:
                messagebox.showerror("Errore", f"Data non valida: {race_day_str}")
                return
            
            # Verifica che sia stato inserito il nome dell'atleta (nuovo)
            athlete_name = self.athlete_name_var.get().strip()
            if not athlete_name:
                messagebox.showerror("Errore", "Il nome dell'atleta è obbligatorio.\nInserisci il nome dell'atleta prima di procedere.")
                return
                
            # Verifica che siano stati selezionati dei giorni della settimana
            selected_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
            if not selected_days:
                messagebox.showwarning("Nessun giorno selezionato", "Seleziona almeno un giorno della settimana.")
                return
            
            # Verifica che il numero di giorni selezionati sia corretto (nuovo)
            if hasattr(self, 'excel_max_sessions') and self.excel_max_sessions > 0:
                if len(selected_days) != self.excel_max_sessions:
                    messagebox.showwarning("Numero errato di giorni", 
                                        f"Questo piano richiede esattamente {self.excel_max_sessions} giorni di allenamento a settimana.\n"
                                        f"Hai selezionato {len(selected_days)} giorni.\n\n"
                                        f"Seleziona esattamente {self.excel_max_sessions} giorni per continuare.")
                    return
                
            # Esegui la pianificazione a ritroso partendo dalla data della gara
            self._schedule_excel_workouts_from_race_day(excel_file, race_date, selected_days, athlete_name)
            
        except Exception as e:
            self.log(f"Errore: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")


    def preselect_excel_days(self, sessions_per_week):
        """Preseleziona i giorni della settimana in base al numero di sessioni nella tab Excel Tools"""
        # Prima deseleziona tutti i giorni
        for var in self.day_selections:
            var.set(0)
        
        # Poi seleziona i giorni appropriati
        if sessions_per_week == 1:
            self.day_selections[2].set(1)  # Mercoledì
        elif sessions_per_week == 2:
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[4].set(1)  # Venerdì
        elif sessions_per_week == 3:
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[3].set(1)  # Giovedì
            self.day_selections[5].set(1)  # Sabato
        elif sessions_per_week == 4:
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[3].set(1)  # Giovedì
            self.day_selections[5].set(1)  # Sabato
            self.day_selections[6].set(1)  # Domenica
        elif sessions_per_week >= 5:
            self.day_selections[0].set(1)  # Lunedì
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[3].set(1)  # Giovedì
            self.day_selections[4].set(1)  # Venerdì
            self.day_selections[6].set(1)  # Domenica


    def analyze_excel_file(self, excel_file):
        """Analizza il file Excel per determinare il numero massimo di sessioni settimanali"""
        try:
            self.log(f"Analisi del file Excel: {excel_file}")
            
            # Carica il file Excel
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            
            # Verifica che esista il foglio Workouts
            if 'Workouts' not in wb.sheetnames:
                self.log("Il foglio 'Workouts' non esiste nel file Excel.")
                self.excel_max_sessions = 0
                return
                
            ws = wb['Workouts']
            
            # Verifica intestazioni
            if ws.cell(row=2, column=1).value != "Week" or ws.cell(row=2, column=3).value != "Session":
                self.log("Il file Excel non ha la struttura attesa. Intestazioni non trovate.")
                self.excel_max_sessions = 0
                return
            
            # Raccogli informazioni sulle sessioni per settimana
            week_col = 1  # Colonna "Week"
            session_col = 3  # Colonna "Session"
            sessions_per_week = {}
            
            for row in range(3, ws.max_row + 1):  # Inizia dalla terza riga (dopo intestazioni)
                week_cell = ws.cell(row=row, column=week_col)
                session_cell = ws.cell(row=row, column=session_col)
                
                if week_cell.value is None or session_cell.value is None:
                    continue
                    
                # Converti in numeri se necessario
                week = week_cell.value
                if isinstance(week, str):
                    try:
                        week = int(float(week.strip()))
                    except ValueError:
                        continue
                        
                # Incrementa il contatore per questa settimana
                if week not in sessions_per_week:
                    sessions_per_week[week] = 0
                sessions_per_week[week] += 1
            
            # Trova il numero massimo di sessioni per settimana
            max_sessions = max(sessions_per_week.values()) if sessions_per_week else 0
            
            self.log(f"Analisi Excel completata: max sessioni per settimana = {max_sessions}")
            
            # Salva il valore e aggiorna l'interfaccia
            self.excel_max_sessions = max_sessions
            
            if max_sessions > 0:
                # Non aggiorniamo più la label informativa
                # self.excel_day_info_label.config(text="...")
                
                # Preseleziona i giorni in base al numero di sessioni
                self.preselect_excel_days(max_sessions)
                
        except Exception as e:
            self.log(f"Errore nell'analisi del file Excel: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            self.excel_max_sessions = 0


    def _schedule_excel_workouts_from_race_day(self, excel_file, race_date, selected_days, athlete_name=""):
        """Pianifica gli allenamenti nel file Excel procedendo a ritroso dalla data della gara"""
        try:
            # Carica il file Excel
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            ws = wb['Workouts']
            
            self.log(f"File Excel caricato: {excel_file}")
            
            # Aggiorna il nome dell'atleta
            if athlete_name:
                athlete_cell = ws.cell(row=1, column=1)
                athlete_cell.value = f"Atleta: {athlete_name}"
                self.log(f"Nome atleta impostato: {athlete_name}")
            
            # Verifica intestazioni e ottieni indici colonne
            if ws.cell(row=2, column=1).value != "Week":
                messagebox.showerror("Errore", "Il file Excel non ha la struttura attesa. La seconda riga dovrebbe contenere 'Week' nella prima colonna.")
                return
            
            col_indices = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=2, column=col).value
                if header:
                    col_indices[header] = col
            
            # Verifica colonne necessarie
            required_columns = ["Week", "Session", "Description", "Steps"]
            missing_columns = [col for col in required_columns if col not in col_indices]
            if missing_columns:
                messagebox.showerror("Errore", f"Colonne mancanti nel file Excel: {', '.join(missing_columns)}")
                return
            
            # Aggiungi colonna Date se non esiste
            if "Date" not in col_indices:
                week_col = col_indices["Week"]
                ws.insert_cols(week_col + 1)
                
                # Aggiorna indici colonne
                for key in list(col_indices.keys()):
                    if col_indices[key] > week_col:
                        col_indices[key] += 1
                
                date_col = week_col + 1
                ws.cell(row=2, column=date_col, value="Date")
                
                # Formatta intestazione
                header_cell = ws.cell(row=2, column=date_col)
                header_cell.font = openpyxl.styles.Font(bold=True)
                header_cell.fill = openpyxl.styles.PatternFill(
                    fill_type="solid",
                    start_color="DDEBF7",
                    end_color="DDEBF7"
                )
                header_cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    bottom=openpyxl.styles.Side(style="thin")
                )
                
                col_indices["Date"] = date_col
                self.log("Aggiunta colonna 'Date' al file Excel")
            
            # Ottieni gli indici di colonna
            week_col = col_indices["Week"]
            session_col = col_indices["Session"]
            date_col = col_indices["Date"]
            
            # Pulisci tutte le celle delle date esistenti prima di procedere con la nuova pianificazione
            self.log("Pulizia delle date esistenti...")
            for row in range(3, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=date_col)
                if date_cell.value is not None:
                    date_cell.value = None
            self.log("Pulizia delle date completata")
            
            # Verifica che ci siano giorni selezionati
            if not selected_days:
                messagebox.showerror("Errore", "Seleziona almeno un giorno della settimana.")
                return
            
            # Converti gli indici dei giorni in nomi per il log
            day_names = {
                0: "Lunedì", 1: "Martedì", 2: "Mercoledì", 
                3: "Giovedì", 4: "Venerdì", 5: "Sabato", 6: "Domenica"
            }
            
            selected_day_names = [day_names[day] for day in sorted(selected_days)]
            self.log(f"Giorni selezionati: {', '.join(selected_day_names)}")
            self.log(f"Data della gara: {race_date.strftime('%d/%m/%Y')} ({day_names[race_date.weekday()]})")
            
            # Raccogli ed ordina gli allenamenti per settimana e sessione
            workouts_by_week = {}
            
            for row in range(3, ws.max_row + 1):
                week_cell = ws.cell(row=row, column=week_col)
                session_cell = ws.cell(row=row, column=session_col)
                
                if week_cell.value is None or session_cell.value is None:
                    continue
                
                # Converti in numeri
                week = week_cell.value
                session = session_cell.value
                
                if isinstance(week, str):
                    try:
                        week = int(float(week.strip()))
                    except ValueError:
                        continue
                
                if isinstance(session, str):
                    try:
                        session = int(float(session.strip()))
                    except ValueError:
                        continue
                
                # Ottieni il colore per la formattazione
                color = None
                if week_cell.fill and week_cell.fill.start_color and week_cell.fill.start_color.index:
                    color = week_cell.fill.start_color.index
                
                # Aggiungi alle informazioni
                if week not in workouts_by_week:
                    workouts_by_week[week] = {}
                
                workouts_by_week[week][session] = {
                    'row': row,
                    'color': color
                }
            
            if not workouts_by_week:
                messagebox.showerror("Errore", "Nessun allenamento trovato nel file Excel.")
                return
            
            # Ottieni il numero di settimane totali e ordina le settimane in ordine decrescente
            weeks = sorted(workouts_by_week.keys(), reverse=True)
            self.log(f"Trovate {len(weeks)} settimane di allenamenti")
            
            # Data corrente e lunedì della settimana della gara
            today = datetime.today().date()
            
            # Trova il lunedì della settimana della gara
            days_to_monday = race_date.weekday()  # 0=lunedì, 1=martedì, ecc.
            race_week_monday = race_date - timedelta(days=days_to_monday)
            
            self.log(f"Lunedì della settimana della gara: {race_week_monday.strftime('%d/%m/%Y')}")
            
            # Ordina selected_days
            selected_days = sorted(selected_days)
            
            # Pianifica gli allenamenti per ogni settimana
            assigned_dates = set()
            
            # Pianifica a ritroso dalle settimane più alte (vicine alla gara) a quelle più basse
            for week in weeks:  # Le settimane sono già ordinate in ordine decrescente
                # Calcola il lunedì della settimana attuale
                week_index = max(weeks) - week
                week_monday = race_week_monday - timedelta(weeks=week_index)
                self.log(f"Pianificazione settimana W{week:02d}: inizia {week_monday.strftime('%d/%m/%Y')} (Lunedì)")
                
                # Ordina sessioni in ordine crescente (S01, S02, S03...)
                sessions = sorted(workouts_by_week[week].keys())
                
                # Assegna date alle sessioni
                for session_idx, session in enumerate(sessions):
                    workout_info = workouts_by_week[week][session]
                    
                    # Calcola la data per questa sessione
                    if session_idx < len(selected_days):
                        day_idx = selected_days[session_idx]
                    else:
                        # Se ci sono più sessioni che giorni selezionati, cicla
                        day_idx = selected_days[session_idx % len(selected_days)]
                    
                    workout_date = week_monday + timedelta(days=day_idx)
                    date_str = workout_date.strftime("%Y-%m-%d")
                    
                    # Verifica se questa data coincide con il giorno della gara
                    if workout_date == race_date:
                        self.log(f"Allenamento W{week:02d}S{session:02d} coinciderebbe con il giorno della gara ({date_str}). Saltato.")
                        continue
                    
                    # Verifica che non sia nel passato
                    if workout_date < today:
                        self.log(f"Allenamento W{week:02d}S{session:02d} cadrebbe nel passato ({date_str}). Saltato.")
                        continue
                    
                    # Verifica che non sia dopo la gara
                    if workout_date > race_date:
                        self.log(f"Allenamento W{week:02d}S{session:02d} cadrebbe dopo la gara ({date_str}). Saltato.")
                        continue
                    
                    # Verifica se questa data è già assegnata ad un altro allenamento
                    if date_str in assigned_dates:
                        # Cerca una data alternativa tra i giorni selezionati
                        self.log(f"Data {date_str} già assegnata. Cercando alternativa.")
                        found_alternative = False
                        
                        # Prova le altre date nella stessa settimana
                        for alt_day in [d for d in selected_days if d != day_idx]:
                            alt_date = week_monday + timedelta(days=alt_day)
                            alt_date_str = alt_date.strftime("%Y-%m-%d")
                            
                            # Verifica che la data alternativa sia valida
                            if (alt_date != race_date and
                                alt_date <= race_date and
                                alt_date >= today and
                                alt_date_str not in assigned_dates):
                                workout_date = alt_date
                                date_str = alt_date_str
                                found_alternative = True
                                self.log(f"Trovata data alternativa: {date_str}")
                                break
                        
                        if not found_alternative:
                            self.log(f"Nessuna data alternativa disponibile per W{week:02d}S{session:02d}. Saltato.")
                            continue
                    
                    # Aggiorna la cella con la data
                    row = workout_info['row']
                    date_cell = ws.cell(row=row, column=date_col)
                    date_cell.value = workout_date
                    date_cell.number_format = "YYYY-MM-DD"  # Formato ISO standard
                    
                    # Formattazione
                    date_cell.alignment = openpyxl.styles.Alignment(
                        horizontal="center", 
                        vertical="center"
                    )
                    
                    # Applica lo stile
                    if workout_info['color']:
                        try:
                            date_cell.fill = openpyxl.styles.PatternFill(
                                fill_type="solid",
                                start_color=workout_info['color'],
                                end_color=workout_info['color']
                            )
                        except:
                            pass  # Ignora errori di formattazione
                    
                    date_cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style="thin"),
                        right=openpyxl.styles.Side(style="thin"),
                        top=openpyxl.styles.Side(style="thin"),
                        bottom=openpyxl.styles.Side(style="thin")
                    )
                    
                    # Registra la data come assegnata
                    assigned_dates.add(date_str)
                    self.log(f"Allenamento W{week:02d}S{session:02d} pianificato per {workout_date.strftime('%d/%m/%Y')} ({day_names[workout_date.weekday()]})")
            
            # Adatta la larghezza della colonna Date
            date_col_letter = openpyxl.utils.get_column_letter(date_col)
            ws.column_dimensions[date_col_letter].width = 15
            
            # Aggiungiamo un allenamento speciale per la gara o impostiamo direttamente la data della gara
            # nel foglio Config, che creiamo se non esiste
            if 'Config' not in wb.sheetnames:
                config_sheet = wb.create_sheet('Config')
                config_sheet['A1'] = 'Parameter'
                config_sheet['B1'] = 'Value'
                config_sheet['A2'] = 'race_day'
                config_sheet['B2'] = race_date.strftime("%Y-%m-%d")
                self.log(f"Creato foglio Config con data della gara: {race_date.strftime('%Y-%m-%d')}")
            else:
                config_sheet = wb['Config']
                
                # Cerca se esiste già race_day
                race_day_found = False
                for row in range(1, config_sheet.max_row + 1):
                    if config_sheet.cell(row=row, column=1).value == 'race_day':
                        config_sheet.cell(row=row, column=2).value = race_date.strftime("%Y-%m-%d")
                        race_day_found = True
                        self.log(f"Aggiornata data della gara nel foglio Config: {race_date.strftime('%Y-%m-%d')}")
                        break
                
                # Se non esiste, la aggiungiamo
                if not race_day_found:
                    next_row = config_sheet.max_row + 1
                    config_sheet.cell(row=next_row, column=1, value='race_day')
                    config_sheet.cell(row=next_row, column=2, value=race_date.strftime("%Y-%m-%d"))
                    self.log(f"Aggiunta data della gara nel foglio Config: {race_date.strftime('%Y-%m-%d')}")
            
            # Salva il file
            wb.save(excel_file)
            
            # Mostra un messaggio di conferma
            messagebox.showinfo("Successo", 
                             f"Pianificazione completata con successo!\n\n"
                             f"Gli allenamenti sono stati pianificati a ritroso a partire dalla data della gara ({race_date.strftime('%d/%m/%Y')}).\n\n"
                             f"Totale allenamenti pianificati: {len(assigned_dates)}")
            
            self.log(f"Pianificazione completata con successo: {len(assigned_dates)} allenamenti pianificati")
            
        except Exception as e:
            self.log(f"Errore durante la pianificazione: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Errore", f"Si è verificato un errore durante la pianificazione:\n{str(e)}")


    def browse_excel_input(self):
        """Apre il dialogo per selezionare il file Excel di input"""
        filename = filedialog.askopenfilename(
            title="Seleziona file Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.excel_input_file.set(filename)
            self.log(f"File Excel selezionato: {filename}")
            
            # Analizza il file Excel per determinare il numero massimo di sessioni
            self.analyze_excel_file(filename)
            
            # Se il file YAML non è stato specificato, proponi lo stesso nome con estensione YAML
            if not self.yaml_output_file.get():
                yaml_path = os.path.splitext(filename)[0] + ".yaml"
                self.yaml_output_file.set(yaml_path)
    
    def browse_yaml_output(self):
        """Apre il dialogo per selezionare il file YAML di output"""
        filename = filedialog.asksaveasfilename(
            title="Salva file YAML",
            filetypes=[("YAML files", "*.yaml"), ("All files", "*.*")],
            defaultextension=".yaml"
        )
        if filename:
            self.yaml_output_file.set(filename)
            self.log(f"File YAML di output impostato: {filename}")

    def create_excel_sample(self):
        """Crea un file Excel di esempio"""
        try:
            # Importa il modulo excel_to_yaml_converter
            from planner.excel_to_yaml_converter import create_sample_excel
            
            filename = filedialog.asksaveasfilename(
                title="Salva file Excel di esempio",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                defaultextension=".xlsx"
            )
            if not filename:
                return
                
            self.log(f"Creazione file Excel di esempio: {filename}")
            
            # Esegui la creazione in un thread separato
            def _create_sample_thread():
                try:
                    created_file = create_sample_excel(filename)
                    
                    # Aggiorna l'interfaccia nel thread principale
                    self.after(0, lambda: self._sample_created(created_file))
                except Exception as e:
                    self.after(0, lambda: self.log(f"Errore nella creazione del file di esempio: {str(e)}"))
                    messagebox.showerror("Errore", f"Si è verificato un errore durante la creazione del file:\n{str(e)}")
            
            threading.Thread(target=_create_sample_thread).start()
            
        except Exception as e:
            self.log(f"Errore: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def _sample_created(self, file_path):
        """Chiamato quando il file di esempio è stato creato"""
        self.log(f"File Excel di esempio creato con successo!")
        self.excel_input_file.set(file_path)
        
        # Analizza il file Excel appena creato
        self.analyze_excel_file(file_path)
        
        # Proponi un file YAML con lo stesso nome
        yaml_path = os.path.splitext(file_path)[0] + ".yaml"
        self.yaml_output_file.set(yaml_path)
        
        messagebox.showinfo("Successo", 
                           f"File Excel di esempio creato con successo:\n{file_path}")

    def convert_excel_to_yaml(self):
        """Converte il file Excel in formato YAML"""
        if not LicenseManager.get_instance().check_feature_access("premium"):
            return
        try:
            # Verifica che il file Excel esista
            excel_file = self.excel_input_file.get()
            if not excel_file:
                messagebox.showerror("Errore", "Seleziona un file Excel da convertire")
                return
                
            if not os.path.exists(excel_file):
                messagebox.showerror("Errore", f"Il file Excel non esiste: {excel_file}")
                return
                
            # Determina il file YAML di output
            yaml_file = self.yaml_output_file.get()
            if not yaml_file:
                yaml_file = os.path.splitext(excel_file)[0] + ".yaml"
                self.yaml_output_file.set(yaml_file)
                
            self.log(f"Conversione di {excel_file} in {yaml_file}...")
            
            # Importa il modulo excel_to_yaml_converter
            from planner.excel_to_yaml_converter import excel_to_yaml
            
            # Esegui la conversione in un thread separato
            def _convert_thread():
                try:
                    plan = excel_to_yaml(excel_file, yaml_file)
                    
                    # Aggiorna l'interfaccia nel thread principale
                    workout_count = len(plan) - 1 if "config" in plan else len(plan)  # -1 per escludere config
                    self.after(0, lambda: self._conversion_success(yaml_file, workout_count))
                except Exception as e:
                    self.after(0, lambda: self.log(f"Errore nella conversione: {str(e)}"))
                    messagebox.showerror("Errore", f"Si è verificato un errore durante la conversione:\n{str(e)}")
            
            threading.Thread(target=_convert_thread).start()
            
        except Exception as e:
            self.log(f"Errore: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def _conversion_success(self, yaml_file, workout_count):
        """Chiamato quando la conversione è stata completata con successo"""
        self.log(f"Conversione completata con successo!")
        self.log(f"File YAML salvato: {yaml_file}")
        self.log(f"Piano di allenamento con {workout_count} allenamenti")
        
        messagebox.showinfo("Successo", 
                          f"Conversione completata con successo!\n\n"
                          f"File YAML salvato: {yaml_file}\n"
                          f"Allenamenti creati: {workout_count}")

    def schedule_workouts_in_excel(self):
        """Apre il dialogo per pianificare gli allenamenti nel file Excel"""
        try:
            # Verifica che sia stato selezionato un file Excel
            excel_file = self.excel_input_file.get()
            if not excel_file:
                messagebox.showerror("Errore", "Seleziona prima un file Excel.")
                return
                
            if not os.path.exists(excel_file):
                messagebox.showerror("Errore", f"Il file Excel non esiste: {excel_file}")
                return
                
            # Importa il modulo necessario
            from planner.excel_to_yaml_converter import excel_to_yaml
            
            # Crea una finestra di dialogo personalizzata per la pianificazione
            self._open_schedule_dialog(excel_file)
            
        except Exception as e:
            self.log(f"Errore: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")

    def _open_schedule_dialog(self, excel_file):
        """Apre un dialogo personalizzato per la pianificazione degli allenamenti"""
        try:
            # Importa la classe di dialogo
            from planner.excel_to_yaml_gui import ScheduleDialog
            
            # Crea il dialogo
            dialog = ScheduleDialog(self, excel_file)
            
            # Se la pianificazione è stata completata con successo
            if dialog.result:
                self.log("Pianificazione degli allenamenti in Excel completata con successo.")
                messagebox.showinfo("Successo", "Le date sono state aggiunte al file Excel.")
                
        except Exception as e:
            self.log(f"Errore durante l'apertura del dialogo di pianificazione: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n{str(e)}")



    def clear_log(self):
        """Pulisce il contenuto della finestra di log"""
        self.log_text.config(state=tk.NORMAL)  # Abilita la modifica del text widget
        self.log_text.delete(1.0, tk.END)      # Cancella tutto il contenuto
        self.log_text.config(state=tk.DISABLED)  # Riporta il widget in stato di sola lettura
        self.log("Log pulito")                 # Aggiunge un messaggio di conferma

    def log(self, message):
        """Add a message to the log tab"""
        logger.info(message)
        self.status_var.set(message)
    
    # Tab functionality methods
    def perform_login(self):
        """Gestisce il processo di login a Garmin Connect"""
        self.log("Avvio procedura di login...")
        
        # Verifica se c'è già una finestra di login aperta (evita duplicati)
        for widget in self.winfo_children():
            if isinstance(widget, tk.Toplevel) and widget.title() == "Login Garmin Connect":
                self.log("Finestra di login già aperta")
                widget.lift()  # Porta in primo piano la finestra esistente
                return
        
        # Assicurati che la cartella OAuth esista
        oauth_folder = self.oauth_folder.get()
        if not os.path.exists(oauth_folder):
            try:
                os.makedirs(oauth_folder, exist_ok=True)
                self.log(f"Creata cartella OAuth: {oauth_folder}")
            except Exception as e:
                self.log(f"Errore nella creazione della cartella OAuth: {str(e)}")
                return
                
        # Creiamo una finestra di dialogo personalizzata
        login_dialog = tk.Toplevel(self)
        login_dialog.title("Login Garmin Connect")
        login_dialog.geometry("300x150")
        login_dialog.resizable(False, False)
        login_dialog.transient(self)  # Rende la finestra modale
        login_dialog.grab_set()       # Impedisce di interagire con la finestra principale
        
        # Variabili per memorizzare email e password
        email_var = tk.StringVar()
        password_var = tk.StringVar()
        
        # Frame per organizzare i widget
        frame = ttk.Frame(login_dialog, padding=10)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Widget per l'email
        ttk.Label(frame, text="Email:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=email_var, width=30).grid(row=0, column=1, pady=5)
        
        # Widget per la password
        ttk.Label(frame, text="Password:").grid(row=1, column=0, sticky=tk.W, pady=5)
        password_entry = ttk.Entry(frame, textvariable=password_var, show="*", width=30)
        password_entry.grid(row=1, column=1, pady=5)

        def do_login():
            """Callback quando l'utente preme Login"""
            email = email_var.get()
            password = password_var.get()
            
            if not email or not password:
                messagebox.showerror("Errore", "Email e password sono obbligatorie", parent=login_dialog)
                return
            
            # Cambia l'interfaccia per mostrare che stiamo elaborando
            for widget in frame.winfo_children():
                if isinstance(widget, ttk.Entry) or isinstance(widget, ttk.Button):
                    widget.configure(state="disabled")
            
            # Mostra un messaggio di attesa
            wait_label = ttk.Label(frame, text="Login in corso...", font=("", 10, "italic"))
            wait_label.grid(row=3, column=0, columnspan=2, pady=5)
            login_dialog.update()
            
            # Esegui il login in un thread separato
            threading.Thread(target=lambda: self._do_login_process(email, password, login_dialog)).start()

        def cancel():
            """Callback quando l'utente preme Annulla"""
            login_dialog.destroy()
            self.log("Login annullato")

        # Pulsanti
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Login", command=do_login).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Annulla", command=cancel).pack(side=tk.LEFT, padx=5)
        
        # Centra la finestra sullo schermo
        login_dialog.update_idletasks()
        width = login_dialog.winfo_width()
        height = login_dialog.winfo_height()
        x = (login_dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (login_dialog.winfo_screenheight() // 2) - (height // 2)
        login_dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # Imposta il focus sull'input dell'email
        login_dialog.focus_set()

        # Pulsanti
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=10)
        
        ttk.Button(button_frame, text="Login", command=do_login).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Annulla", command=cancel).pack(side=tk.LEFT, padx=5)
        
        # Centra la finestra sullo schermo
        login_dialog.update_idletasks()
        width = login_dialog.winfo_width()
        height = login_dialog.winfo_height()
        x = (login_dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (login_dialog.winfo_screenheight() // 2) - (height // 2)
        login_dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        # Imposta il focus sull'input dell'email
        login_dialog.focus_set()
        
        # Attendiamo che la finestra sia chiusa prima di continuare
        self.wait_window(login_dialog)

    def _do_login_process(self, email, password, login_dialog=None):
        """Esegue effettivamente il processo di login con le credenziali fornite"""
        # Assicurati che la cartella OAuth esista
        oauth_folder = self.oauth_folder.get()
        success = False
        
        try:
            # Non dichiarare os all'interno della funzione perché è già importato a livello globale
            if not os.path.exists(oauth_folder):
                os.makedirs(oauth_folder, exist_ok=True)
                self.log(f"Creata cartella OAuth: {oauth_folder}")
            
            # Aggiungiamo la directory corrente al percorso Python
            sys.path.append(SCRIPT_DIR)
            
            # Impostiamo l'ambiente per disabilitare la verifica SSL
            os.environ['PYTHONHTTPSVERIFY'] = '0'
            
            # Importiamo direttamente i moduli necessari per il login
            import garth
            
            # Esegui il login direttamente
            self.log(f"Tentativo di login con email: {email[:3]}***")
            try:
                garth.login(email, password)
                garth.save(oauth_folder)
                self.log("Login completato con successo")
                success = True
                
                # Chiudi il dialogo di login se esiste
                if login_dialog and login_dialog.winfo_exists():
                    login_dialog.destroy()
                    
                # Aggiorna l'interfaccia dopo il login riuscito
                self.after(0, self.refresh_login_tab)
                self.after(100, lambda: messagebox.showinfo("Successo", "Login completato con successo"))
                
            except Exception as e:
                self.log(f"Errore durante il login: {str(e)}")
                
                # Se il dialogo esiste ancora, aggiorna l'interfaccia per mostrare l'errore
                if login_dialog and login_dialog.winfo_exists():
                    def update_ui():
                        # Rimuovi messaggio di attesa se esiste
                        for widget in login_dialog.winfo_children():
                            if isinstance(widget, ttk.Label) and "Login in corso" in str(widget.cget("text")):
                                widget.destroy()
                        
                        # Riattiva i campi
                        for widget in login_dialog.winfo_children():
                            if isinstance(widget, ttk.Entry):
                                widget.configure(state="normal")
                        
                        # Riattiva i pulsanti
                        for widget in login_dialog.winfo_children():
                            if isinstance(widget, ttk.Button):
                                widget.configure(state="normal")
                        
                        # Mostra errore
                        messagebox.showerror("Errore", f"Errore durante il login: {str(e)}", parent=login_dialog)
                    
                    self.after(0, update_ui)
                else:
                    # Mostra errore nella finestra principale
                    self.after(0, lambda: messagebox.showerror("Errore", f"Errore durante il login: {str(e)}"))
                
        except Exception as e:
            self.log(f"Errore durante l'inizializzazione del login: {str(e)}")
            
            # Se il dialogo esiste ancora, chiudilo
            if login_dialog and login_dialog.winfo_exists():
                login_dialog.destroy()
                
            # Mostra errore nella finestra principale
            self.after(0, lambda: messagebox.showerror("Errore", f"Errore durante l'inizializzazione del login: {str(e)}"))

    def perform_logout(self):
        """Esegue il logout cancellando i file OAuth"""
        # Chiedi conferma all'utente
        if not messagebox.askyesno("Conferma logout", 
                                 "Sei sicuro di voler effettuare il logout?\n\n"
                                 "Verranno eliminati tutti i file di autenticazione e dovrai effettuare nuovamente il login."):
            return
        
        oauth_folder = self.oauth_folder.get()
        if os.path.exists(oauth_folder):
            try:
                # Elimina tutti i file OAuth
                oauth_files = [f for f in os.listdir(oauth_folder) if f.startswith('oauth') and f.endswith('.json')]
                
                if not oauth_files:
                    messagebox.showinfo("Informazione", "Nessun file di autenticazione trovato.")
                    return
                    
                for file in oauth_files:
                    file_path = os.path.join(oauth_folder, file)
                    os.remove(file_path)
                    self.log(f"File eliminato: {file_path}")
                
                # Aggiorna la tab Login
                self.refresh_login_tab()
                
                messagebox.showinfo("Logout completato", "Logout effettuato con successo.")
                
            except Exception as e:
                self.log(f"Errore durante il logout: {str(e)}")
                messagebox.showerror("Errore", f"Si è verificato un errore durante il logout: {str(e)}")
        else:
            messagebox.showinfo("Informazione", "Cartella OAuth non trovata.")

    def refresh_login_tab(self):
        """Aggiorna l'interfaccia della tab Login ricostruendola"""
        # Salva l'indice della tab attualmente selezionata
        current_tab = self.notebook.index("current")
        
        # Rimuovi la tab esistente
        login_tab_index = 0  # Assumiamo che Login sia la prima tab
        self.notebook.forget(login_tab_index)
        
        # Ricrea la tab Login
        login_frame = ttk.Frame(self.notebook)
        self.notebook.insert(login_tab_index, login_frame, text="Login")
        
        # Ridisegna il contenuto della tab
        self.create_login_tab_content(login_frame)
        
        # Ripristina la tab che era selezionata
        self.notebook.select(current_tab)
        
        self.log("Tab Login aggiornata con il nuovo stato")


    def browse_import_file(self):
        filename = filedialog.askopenfilename(
            title="Seleziona file allenamenti",
            filetypes=[("YAML files", "*.yaml *.yml"), ("All files", "*.*")]
        )
        if filename:
            self.import_file.set(filename)
    
    def browse_export_file(self):
        filename = filedialog.asksaveasfilename(
            title="Salva file allenamenti",
            filetypes=[("YAML files", "*.yaml"), ("JSON files", "*.json"), ("All files", "*.*")],
            defaultextension=".yaml"
        )
        if filename:
            self.export_file.set(filename)
            self.log(f"File di esportazione impostato: {filename}")
            self.log(f"Il formato sarà determinato dall'estensione del file (.yaml per YAML, .json per JSON)")
    
    def load_training_plans(self):
        """Load available training plans from the training_plans directory"""
        try:
            # Clear existing items
            for item in self.training_plans_tree.get_children():
                self.training_plans_tree.delete(item)
            
            # Look for the training_plans directory
            plans_dir = os.path.join(SCRIPT_DIR, "training_plans")
            if not os.path.exists(plans_dir):
                self.log("Directory training_plans non trovata")
                return
            
            # Scan for YAML files in the directory structure
            for root, _, files in os.walk(plans_dir):
                for file in files:
                    if file.endswith(('.yaml', '.yml')):
                        rel_path = os.path.relpath(os.path.join(root, file), plans_dir)
                        parts = rel_path.split(os.sep)
                        
                        # Supporta sia la struttura gerarchica che quella semplice
                        if len(parts) >= 3:
                            # Struttura gerarchica originale: tipo/settimane/variante/file.yaml
                            plan_type = parts[0]  # e.g., half_marathon
                            weeks = parts[1]      # e.g., 8_weeks
                            variant = parts[2]    # e.g., paris
                            
                            # Extract target time if available in the filename
                            target_time = ""
                            match = re.search(r'(\d+[KM])@([\dh]+)', file)
                            if match:
                                distance, time = match.groups()
                                target_time = f" ({distance} in {time})"
                            
                            plan_name = f"{plan_type} - {variant}{target_time}"
                            
                            # Add to treeview
                            self.training_plans_tree.insert("", "end", 
                                                         values=(plan_name, weeks.replace("_", " "), 
                                                                 os.path.join(root, file)))
                        elif len(parts) == 2:
                            # Struttura semplice: cartella/file.yaml
                            folder = parts[0]     # e.g., frank
                            filename = parts[1]   # e.g., my_plan.yaml
                            
                            # Extract name without extension
                            plan_name = os.path.splitext(filename)[0]
                            
                            # Add to treeview with folder name as the plan type
                            self.training_plans_tree.insert("", "end", 
                                                         values=(f"{plan_name} ({folder})", 
                                                                 "N/A", 
                                                                 os.path.join(root, file)))
                        elif len(parts) == 1:
                            # File direttamente nella cartella training_plans
                            filename = parts[0]
                            plan_name = os.path.splitext(filename)[0]
                            
                            # Add to treeview
                            self.training_plans_tree.insert("", "end", 
                                                         values=(plan_name, 
                                                                 "N/A", 
                                                                 os.path.join(root, file)))
            
            if not self.training_plans_tree.get_children():
                self.log("Nessun piano di allenamento trovato")
            else:
                self.log(f"Trovati {len(self.training_plans_tree.get_children())} piani di allenamento")
        
        except Exception as e:
            self.log(f"Errore nel caricamento dei piani di allenamento: {str(e)}")
    
    def select_training_plan(self, event):
        """Gestisce il doppio click su un piano di allenamento nella tree view"""
        try:
            # Ottieni l'elemento selezionato
            item = self.training_plans_tree.selection()[0]
            plan_path = self.training_plans_tree.item(item, "values")[2]
            
            # Imposta il file di importazione
            if hasattr(self, 'import_file'):
                self.import_file.set(plan_path)
            
            # Leggi il file YAML
            with open(plan_path, 'r') as f:
                plan_data = yaml.safe_load(f)
                
                # Estrai l'ID del piano (name_prefix)
                plan_id = ""
                if 'config' in plan_data and 'name_prefix' in plan_data['config']:
                    plan_id = plan_data['config']['name_prefix'].strip()
                    self.log(f"Piano selezionato: {plan_id}")
                
                # Ottieni l'oggetto principale dell'applicazione (root)
                # Questo è necessario perché potremmo essere in un contesto dove self 
                # non è l'oggetto principale dell'applicazione
                root = self
                while hasattr(root, 'master') and root.master is not None:
                    root = root.master
                
                # Verifica se root o self hanno training_plan
                if hasattr(root, 'training_plan'):
                    root.training_plan.set(plan_id)
                    self.log("Impostato ID piano nell'oggetto root")
                elif hasattr(self, 'training_plan'):
                    self.training_plan.set(plan_id)
                    self.log("Impostato ID piano nell'oggetto self")
                else:
                    self.log("Attenzione: Attributo training_plan non trovato")
                
                # Esegui l'analisi del piano usando il metodo appropriato
                if hasattr(root, 'analyze_yaml_plan'):
                    root.analyze_yaml_plan(plan_path)
                    self.log("Analisi piano eseguita sull'oggetto root")
                elif hasattr(self, 'analyze_yaml_plan'):
                    self.analyze_yaml_plan(plan_path)
                    self.log("Analisi piano eseguita sull'oggetto self")
                else:
                    self.log("Attenzione: Metodo analyze_yaml_plan non trovato")
                
        except Exception as e:
            # Log dettagliato dell'errore
            import traceback
            error_traceback = traceback.format_exc()
            self.log(f"Errore durante la selezione del piano: {str(e)}")
            self.log(f"Traceback:\n{error_traceback}")
            

    def update_date_selectors(self, date):
        """Aggiorna i selettori di data Anno/Mese/Giorno in base alla data fornita"""
        try:
            # Cerca i selettori di data nella tab Pianifica
            schedule_tab = None
            for tab_id in self.notebook.tabs():
                if self.notebook.tab(tab_id, "text") == "Pianifica":
                    schedule_tab = self.notebook.nametowidget(tab_id)
                    break
            
            if not schedule_tab:
                self.log("Tab Pianifica non trovata")
                return
            
            # Cerca i frame contenenti i selettori di data
            date_selectors_found = False
            month_names = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", 
                          "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
            
            for frame in schedule_tab.winfo_children():
                if isinstance(frame, ttk.Frame):
                    # Cerca ricorsivamente nei sotto-frame
                    for child in frame.winfo_children():
                        if isinstance(child, ttk.Frame):
                            # Cerca i combobox
                            year_combo = None
                            month_combo = None
                            day_combo = None
                            
                            for combo in child.winfo_children():
                                if isinstance(combo, ttk.Combobox):
                                    values = combo.cget("values")
                                    if values and len(values) > 0:
                                        # Identifica il tipo di combo in base ai valori
                                        if str(date.year) in values:
                                            year_combo = combo
                                        elif len(values) == 12 and month_names[0] in values:
                                            month_combo = combo
                                        elif len(values) <= 31:
                                            day_combo = combo
                            
                            # Se abbiamo trovato tutti i combobox, aggiorniamo i valori
                            if year_combo and month_combo and day_combo:
                                year_combo.set(str(date.year))
                                month_combo.set(month_names[date.month - 1])
                                day_combo.set(str(date.day))
                                date_selectors_found = True
                                self.log("Selettori di data aggiornati con successo")
                                break
            
            if not date_selectors_found:
                self.log("Non è stato possibile trovare i selettori di data nella tab Pianifica")
                
        except Exception as e:
            self.log(f"Errore nell'aggiornamento dei selettori di data: {str(e)}")


    def analyze_yaml_plan(self, yaml_path):
        """Analizza il piano direttamente dal file YAML senza necessità di importarlo"""
        try:
            self.log(f"Analisi piano da file YAML: {yaml_path}")
            
            with open(yaml_path, 'r') as f:
                plan_data = yaml.safe_load(f)
                
                # Estrai la data della gara se presente nella configurazione
                config = plan_data.get('config', {})
                
                # Estrai l'ID del piano (prefisso)
                if 'name_prefix' in config:
                    plan_id = config['name_prefix'].strip()
                    self.log(f"ID del piano estratto dal YAML: {plan_id}")
                    
                    # Imposta l'ID piano sia nella vecchia variabile che nella nuova 
                    # a seconda di quali esistono nell'interfaccia attuale
                    if hasattr(self, 'training_plan'):
                        self.training_plan.set(plan_id)
                    if hasattr(self, 'training_plan '):
                        self.training_plan  .set(plan_id)
                else:
                    self.log("ID del piano (name_prefix) non trovato nella configurazione")
                    
                # Estrai la data della gara
                if 'race_day' in config:
                    race_day_str = config['race_day']
                    self.log(f"Trovata data della gara nel YAML: {race_day_str}")
                    
                    # Imposta la data della gara nelle variabili appropriate
                    if hasattr(self, 'race_day'):
                        self.race_day.set(race_day_str)
                    if hasattr(self, 'race_day'):
                        self.race_day.set(race_day_str)
                    
                    # Aggiorna i selettori di data Anno/Mese/Giorno se il metodo esiste
                    try:
                        race_date = datetime.strptime(race_day_str, "%Y-%m-%d")
                        if hasattr(self, 'update_date_selectors'):
                            self.update_date_selectors(race_date)
                    except Exception as e:
                        self.log(f"Errore nell'aggiornamento dei selettori di data: {str(e)}")
                    
                    # Memorizza il valore originale della data gara per confronti futuri
                    self.original_race_day = race_day_str
                else:
                    self.log("Data della gara non trovata nella configurazione")
                    self.original_race_day = None
                
                # Estrai i giorni preferiti e aggiornali nell'interfaccia
                if 'preferred_days' in config:
                    preferred_days = config['preferred_days']
                    
                    # Se preferred_days è una stringa rappresentante una lista, convertila
                    if isinstance(preferred_days, str):
                        try:
                            # Rimuovi parentesi quadre, split per virgole e converti in int
                            preferred_days = preferred_days.strip('[]').split(',')
                            preferred_days = [int(d.strip()) for d in preferred_days if d.strip()]
                            self.log(f"Giorni preferiti convertiti da stringa: {preferred_days}")
                        except Exception as e:
                            self.log(f"Errore nella conversione dei giorni preferiti: {str(e)}")
                            preferred_days = []
                    
                    self.log(f"Trovati giorni preferiti nel YAML: {preferred_days}")
                    
                    # Memorizza i giorni originali per confronti futuri
                    self.original_preferred_days = preferred_days
                    
                    # Aggiorna i checkbox nell'interfaccia appropriata
                    selections = []
                    if hasattr(self, 'day_selections'):
                        selections = self.day_selections
                    elif hasattr(self, 'day_selections'):
                        selections = self.day_selections
                    
                    if selections:
                        # Resetta tutti i checkbox
                        for i in range(len(selections)):
                            selections[i].set(0)
                        
                        # Seleziona i giorni preferiti
                        for day_index in preferred_days:
                            if 0 <= day_index < len(selections):  # Verifica che l'indice sia valido
                                selections[day_index].set(1)
                else:
                    self.log("Giorni preferiti non trovati nella configurazione")
                    # Se non ci sono giorni preferiti, usa preselect_days se è disponibile
                    if hasattr(self, 'preselect_days'):
                        self.preselect_days(self.max_sessions if hasattr(self, 'max_sessions') else 0)
                    self.original_preferred_days = None
                
                # Rimuovi la sezione config per continuare l'analisi degli allenamenti
                plan_data_copy = plan_data.copy()
                if 'config' in plan_data_copy:
                    plan_data_copy.pop('config')
                
                # Conta gli allenamenti (escludendo la configurazione)
                workout_count = len(plan_data_copy)
                
                # Analizza le settimane e sessioni
                weeks = {}
                workouts_with_dates = {}
                
                for workout_name, steps in plan_data_copy.items():
                    # Cerca il pattern WxxSxx nel nome
                    match = re.search(r'\s*(W\d\d)S(\d\d)\s*', workout_name)
                    if match:
                        week_id = match.group(1)
                        session_id = match.group(2)
                        
                        # Conteggia le sessioni per settimana
                        if week_id not in weeks:
                            weeks[week_id] = 0
                        weeks[week_id] += 1
                        
                        # Estrai la data se presente
                        if steps and isinstance(steps, list) and len(steps) > 0:
                            first_step = steps[0]
                            if isinstance(first_step, dict) and 'date' in first_step:
                                date_str = first_step['date']
                                workouts_with_dates[workout_name] = date_str
                                self.log(f"Trovata data {date_str} per allenamento {workout_name}")
                
                # Salva le date originali per confronti futuri
                self.original_workout_dates = workouts_with_dates
                
                # Costruisci il testo informativo
                num_weeks = len(weeks)
                max_sessions = max(weeks.values()) if weeks else 0
                
                # Salva il numero massimo di sessioni come attributo della classe
                self.max_sessions = max_sessions
                if hasattr(self, 'excel_max_sessions'):
                    self.excel_max_sessions = max_sessions
                
                info_text = f"Piano da file YAML: {workout_count} allenamenti, {num_weeks} settimane"
                if weeks:
                    info_text += f"\nAllenamenti per settimana: "
                    for week, count in sorted(weeks.items()):
                        info_text += f"{week}={count} "
                
                # Aggiungi informazioni sulle date se presenti
                if workouts_with_dates:
                    num_dates = len(workouts_with_dates)
                    info_text += f"\nDate pianificate: {num_dates}/{workout_count} allenamenti"
                
                # Aggiorna l'interfaccia
                if hasattr(self, 'training_plan_info'):
                    self.training_plan_info.set(info_text)
                if hasattr(self, 'plan_stats_text'):
                    self.plan_stats_text.set(info_text)
                
                # Aggiorna il percorso del file YAML per lo step 3 se esiste
                if hasattr(self, 'step3_yaml_path'):
                    self.step3_yaml_path.set(yaml_path)
                
                # Imposta il flag di piano caricato
                self.plan_imported = True
                
                # Salva il percorso del YAML come attributo per riferimento futuro
                self.current_yaml_path = yaml_path
                
                self.log(f"Analisi completata: {workout_count} allenamenti in {num_weeks} settimane")
                
                return True
                
        except Exception as e:
            self.log(f"Errore nell'analisi del piano YAML: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            
            # Reimposta le variabili di stato
            if hasattr(self, 'training_plan_info'):
                self.training_plan_info.set(f"Errore nell'analisi del piano YAML: {str(e)}")
            if hasattr(self, 'plan_stats_text'):
                self.plan_stats_text.set(f"Errore nell'analisi del piano YAML: {str(e)}")
            
            self.plan_imported = False
            self.max_sessions = 0
            if hasattr(self, 'excel_max_sessions'):
                self.excel_max_sessions = 0
                
            return False

    def preselect_days(self, sessions_per_week):
        """Preseleziona i giorni della settimana in base al numero di sessioni"""
        # Prima deseleziona tutti i giorni
        for var in self.day_selections:
            var.set(0)
        
        # Poi seleziona i giorni appropriati
        if sessions_per_week == 1:
            self.day_selections[2].set(1)  # Mercoledì
        elif sessions_per_week == 2:
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[4].set(1)  # Venerdì
        elif sessions_per_week == 3:
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[3].set(1)  # Giovedì
            self.day_selections[5].set(1)  # Sabato
        elif sessions_per_week == 4:
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[3].set(1)  # Giovedì
            self.day_selections[5].set(1)  # Sabato
            self.day_selections[6].set(1)  # Domenica
        elif sessions_per_week >= 5:
            self.day_selections[0].set(1)  # Lunedì
            self.day_selections[1].set(1)  # Martedì
            self.day_selections[3].set(1)  # Giovedì
            self.day_selections[4].set(1)  # Venerdì
            self.day_selections[6].set(1)  # Domenica

    def perform_import(self):
        """Import workouts from a YAML file"""
        if not self.import_file.get():
            messagebox.showerror("Errore", "Seleziona un file di allenamento")
            return
        
        self.log(f"Importazione degli allenamenti da {self.import_file.get()}...")
        
        # Run the import command in a separate thread
        threading.Thread(target=self._do_import).start()
    
    def _do_import(self):
        try:
            self.log("Importazione degli allenamenti in corso...")
            
            # Crea un oggetto args simulato
            class Args:
                pass
                    
            args = Args()
            args.oauth_folder = self.oauth_folder.get()
            args.workouts_file = self.import_file.get()
            args.name_filter = None
            args.replace = self.import_replace.get()
            args.dry_run = False
            args.treadmill = False
            
            # Esegui la funzione direttamente
            cmd_import_workouts(args)
            
            self.log("Importazione completata con successo")
            messagebox.showinfo("Successo", "Importazione completata con successo")
            
            # Aggiorna la lista degli allenamenti
            self.refresh_workouts()
            
        except Exception as e:
            self.log(f"Errore durante l'importazione: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore durante l'importazione: {str(e)}")
    
    def refresh_workouts(self):
        """Refresh the list of workouts from Garmin Connect"""
        self.log("Aggiornamento lista allenamenti...")
        
        # Assicurati che la cartella OAuth esista
        oauth_folder = self.oauth_folder.get()
        if not os.path.exists(oauth_folder):
            try:
                os.makedirs(oauth_folder, exist_ok=True)
                self.log(f"Creata cartella OAuth: {oauth_folder}")
            except Exception as e:
                self.log(f"Errore nella creazione della cartella OAuth: {str(e)}")
                return
        
        # Clear existing items
        for item in self.workouts_tree.get_children():
            self.workouts_tree.delete(item)
        
        # Esegui l'aggiornamento in un thread separato
        threading.Thread(target=self._refresh_workouts).start()
    
    def _refresh_workouts(self):
        try:
            # Crea un client Garmin direttamente
            client = GarminClient(self.oauth_folder.get())
            
            # Ottieni la lista degli allenamenti
            workouts = client.list_workouts()
            
            # Salva gli allenamenti nella cache
            self.save_workouts_to_cache(workouts)
            
            # Aggiorna la treeview nel thread principale
            def update_ui():
                # Cancella gli elementi esistenti
                for item in self.workouts_tree.get_children():
                    self.workouts_tree.delete(item)
                
                # Aggiungi i nuovi allenamenti
                for workout in workouts:
                    workout_id = workout.get('workoutId', 'N/A')
                    workout_name = workout.get('workoutName', 'Senza nome')
                    self.workouts_tree.insert("", "end", values=(workout_id, workout_name))
                
                self.log(f"Trovati {len(workouts)} allenamenti su Garmin Connect")
            
            # Esegui l'aggiornamento dell'UI nel thread principale
            self.after(0, update_ui)
            
        except Exception as e:
            self.log(f"Errore durante l'aggiornamento: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
    
    def refresh_calendar(self):
        """Refresh the calendar view"""
        self.log("Aggiornamento calendario...")
        
        # Assicurati che la cartella OAuth esista
        oauth_folder = self.oauth_folder.get()
        if not os.path.exists(oauth_folder):
            try:
                os.makedirs(oauth_folder, exist_ok=True)
                self.log(f"Creata cartella OAuth: {oauth_folder}")
            except Exception as e:
                self.log(f"Errore nella creazione della cartella OAuth: {str(e)}")
                return
        
        # Clear existing items
        for item in self.calendar_tree.get_children():
            self.calendar_tree.delete(item)
        
        # Esegui l'aggiornamento in un thread separato
        threading.Thread(target=self._refresh_calendar).start()
    
    def _refresh_calendar(self):
        try:
            # Crea un oggetto args simulato
            class Args:
                pass
                    
            args = Args()
            args.oauth_folder = self.oauth_folder.get()
            args.start_date = None
            args.end_date = None
            args.date_range = None
            
            # Aggiungi il filtro per nome se specificato
            if self.training_plan.get():
                args.name_filter = self.training_plan.get()
            else:
                args.name_filter = None
            
            # Ottieni gli allenamenti pianificati
            calendar_data = get_scheduled(args)
            
            # Aggiorna l'UI nel thread principale
            def update_ui():
                # Pulisci la tabella corrente
                for item in self.calendar_tree.get_children():
                    self.calendar_tree.delete(item)
                
                # Aggiungi alla tabella
                for item in calendar_data:
                    date = item.get('date', 'N/A')
                    workout_name = item.get('title', 'Senza nome')
                    self.calendar_tree.insert("", "end", values=(date, workout_name))
                
                self.log(f"Trovati {len(calendar_data)} allenamenti pianificati")
            
            # Esegui l'aggiornamento dell'UI nel thread principale
            self.after(0, update_ui)
            
        except Exception as e:
            self.log(f"Errore durante l'aggiornamento del calendario: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
    

    def perform_schedule(self):
        """Handle Schedule button click"""
        if not LicenseManager.get_instance().check_feature_access("pro"):
            return

        if not self.training_plan.get():
            messagebox.showerror("Errore", "Inserisci l'ID del piano di allenamento")
            return
        
        if not self.race_day.get():
            messagebox.showerror("Errore", "Inserisci la data della gara")
            return
        
        # Validate race day format
        try:
            race_date = datetime.strptime(self.race_day.get(), "%Y-%m-%d").date()
            today = datetime.today().date()
            
            if race_date < today:
                messagebox.showerror("Errore", "La data della gara deve essere nel futuro")
                return
        except ValueError:
            messagebox.showerror("Errore", "Formato data non valido. Usa YYYY-MM-DD")
            return
        
        # Ottieni i giorni selezionati
        selected_days = []
        for i, var in enumerate(self.day_selections):
            if var.get():
                selected_days.append(i)  # Nota: usiamo gli indici numerici, non le stringhe
        
        if not selected_days:
            messagebox.showwarning("Nessun giorno selezionato", 
                                  "Non hai selezionato nessun giorno per gli allenamenti.\n"
                                  "Per favore, seleziona almeno un giorno della settimana.")
            return
        
        # Verifica che il numero di giorni selezionati sia corretto
        if hasattr(self, 'max_sessions') and self.max_sessions > 0:
            if len(selected_days) != self.max_sessions:
                messagebox.showwarning("Numero errato di giorni", 
                                      f"Questo piano richiede esattamente {self.max_sessions} giorni di allenamento a settimana.\n"
                                      f"Hai selezionati {len(selected_days)} giorni.\n\n"
                                      f"Seleziona esattamente {self.max_sessions} giorni per continuare.")
                return
        
        # Verifica se siamo in modalità simulazione
        is_dry_run = False
        training_plan_id = self.training_plan.get().strip()
        
        # Verifica se abbiamo un file YAML caricato
        yaml_path = None
        original_race_day = None
        
        if hasattr(self, 'current_yaml_path') and self.current_yaml_path:
            yaml_path = self.current_yaml_path
            
            # Carica il file YAML per verificare la data della gara
            try:
                with open(yaml_path, 'r') as f:
                    plan_data = yaml.safe_load(f)
                    config = plan_data.get('config', {})
                    if 'race_day' in config:
                        original_race_day = config['race_day']
            except Exception as e:
                self.log(f"Errore nel caricamento del file YAML per verifica: {str(e)}")
        
        # Controlla se ci sono state modifiche alle date o ai giorni
        date_modified = False
        days_modified = False
        
        if hasattr(self, 'original_race_day') and self.original_race_day:
            if self.original_race_day != self.race_day.get():
                self.log(f"Data gara modificata: {self.original_race_day} -> {self.race_day.get()}")
                date_modified = True
        
        if hasattr(self, 'original_preferred_days') and self.original_preferred_days:
            current_days = selected_days
            if set(self.original_preferred_days) != set(current_days):
                self.log(f"Giorni preferiti modificati: {self.original_preferred_days} -> {current_days}")
                days_modified = True
        
        # Se siamo in modalità dry run, non dobbiamo importare nulla, solo simulare
        if is_dry_run:
            self.log("Modalità simulazione attivata, nessuna modifica effettiva verrà apportata")
            
            # Se non ci sono modifiche e abbiamo date originali nel YAML, usa quelle
            if (not date_modified and not days_modified and 
                hasattr(self, 'original_workout_dates') and self.original_workout_dates):
                self._display_original_dates()
                return
            
            # Altrimenti, esegui la simulazione normale
            self._simulate_schedule(selected_days)
            return
        
        # NON siamo in modalità simulazione
        # Se non ci sono modifiche e abbiamo date nel YAML, possiamo usare quelle per pianificare
        use_original_dates = (not date_modified and not days_modified and 
                              hasattr(self, 'original_workout_dates') and self.original_workout_dates)
        
        # Verifica se gli allenamenti sono già stati importati
        self.log("Verifica se il piano è già importato...")
        self.plan_imported = False  # Reset dello stato di importazione
        
        # Controlla nella cache se esistono allenamenti per questo piano
        self.check_workouts_in_cache(training_plan_id)
        
        # Se non è stato importato, dobbiamo importarlo prima di pianificare
        if not self.plan_imported:
            self.log("Piano non importato. Ricerca del file YAML...")
            
            # Trova il file YAML corrispondente
            if not yaml_path:
                yaml_path = self.find_yaml_for_plan(training_plan_id)
            
            if not yaml_path:
                # Nessun file YAML trovato, chiedi all'utente di specificarne uno
                yaml_path = filedialog.askopenfilename(
                    title=f"Seleziona file YAML per il piano '{training_plan_id}'",
                    filetypes=[("YAML files", "*.yaml *.yml"), ("All files", "*.*")]
                )
                
                if not yaml_path:
                    self.log("Operazione annullata. Nessun file selezionato.")
                    return
            
            # Se ci sono modifiche e stiamo usando un file YAML, dobbiamo aggiornarlo
            if (date_modified or days_modified) and yaml_path:
                self.log("Sono state apportate modifiche, aggiornamento del file YAML...")
                if self._update_yaml_dates(yaml_path, self.race_day.get(), selected_days):
                    # Ricarica le date dal YAML aggiornato
                    self.log("File YAML aggiornato, ricarico le date...")
                    self.analyze_yaml_plan(yaml_path)
                    # Imposta flag per usare le date aggiornate
                    use_original_dates = True
            
            # Conferma importazione all'utente
            if not messagebox.askyesno("Conferma importazione", 
                                      f"Prima di pianificare, gli allenamenti devono essere importati in Garmin Connect.\n\n"
                                      f"Importare gli allenamenti dal file:\n{yaml_path}?"):
                self.log("Operazione annullata dall'utente.")
                return
            
            # Imposta il file di importazione e opzioni
            self.import_file.set(yaml_path)
            self.import_replace.set(True)  # Sostituisci eventuali allenamenti esistenti
            
            self.log(f"Importazione degli allenamenti da {yaml_path}...")
            
            # Esegui l'importazione e verifica che abbia avuto successo
            import_success = self._do_import_for_schedule()
            
            if not import_success:
                self.log("Impossibile procedere con la pianificazione a causa di errori nell'importazione.")
                messagebox.showerror("Errore", "La pianificazione è stata annullata a causa di problemi durante l'importazione.")
                return
            else:
                # Imposta il flag di piano importato
                self.plan_imported = True
                self.log("Importazione completata con successo. Procedo con la pianificazione.")
        else:
            self.log("Piano già importato in Garmin Connect. Procedo con la pianificazione.")
            
            # Se ci sono modifiche e stiamo usando un file YAML, aggiorniamolo
            if (date_modified or days_modified) and yaml_path:
                self.log("Sono state apportate modifiche, aggiornamento del file YAML...")
                if self._update_yaml_dates(yaml_path, self.race_day.get(), selected_days):
                    # Ricarica le date dal YAML aggiornato
                    self.log("File YAML aggiornato, ricarico le date...")
                    self.analyze_yaml_plan(yaml_path)
                    # Imposta flag per usare le date aggiornate
                    use_original_dates = True
        
        # Ora che siamo sicuri che il piano sia importato, procedi con la pianificazione
        self.log(f"Pianificazione allenamenti per il piano {training_plan_id}...")
        self.log(f"Data di gara: {self.race_day.get()}")
        self.log(f"Giorni selezionati: {[self.day_names[d] for d in selected_days]}")
        
        # Se stiamo usando le date originali dal YAML, pianificazione diretta
        if use_original_dates:
            self.log("Utilizzo le date originali dal file YAML per la pianificazione")
            self._schedule_with_yaml_dates()
        else:
            # Altrimenti esegui la pianificazione standard
            self._do_schedule(selected_days, False)


    def _schedule_with_yaml_dates(self):
        """Pianifica gli allenamenti usando le date specificate nel file YAML"""
        try:
            # Verifica se abbiamo le date originali dai workout
            if not hasattr(self, 'original_workout_dates') or not self.original_workout_dates:
                self.log("Nessuna data originale trovata, uso la pianificazione standard")
                self._do_schedule(selected_days, False)
                return
            
            self.log("Pianificazione con le date dal file YAML in corso...")
            
            # Assicurati che il cliente Garmin sia creato
            client = GarminClient(self.oauth_folder.get())
            
            # Ottieni la lista degli allenamenti
            all_workouts = client.list_workouts()
            workouts_by_name = {}
            
            # Crea un mapping nome -> id
            for workout in all_workouts:
                name = workout.get('workoutName', '')
                workout_id = workout.get('workoutId', '')
                if name and workout_id:
                    workouts_by_name[name] = workout_id
            
            # Pianifica ogni allenamento con la sua data
            scheduled_count = 0
            for workout_name, date_str in self.original_workout_dates.items():
                # Verifica se l'allenamento esiste
                if workout_name in workouts_by_name:
                    workout_id = workouts_by_name[workout_name]
                    self.log(f"Pianificazione dell'allenamento '{workout_name}' per la data {date_str}")
                    
                    try:
                        response = client.schedule_workout(workout_id, date_str)
                        scheduled_count += 1
                        self.log(f"Allenamento pianificato con successo: {workout_name} ({date_str})")
                    except Exception as e:
                        self.log(f"Errore nella pianificazione dell'allenamento {workout_name}: {str(e)}")
                else:
                    self.log(f"Allenamento non trovato in Garmin Connect: {workout_name}")
            
            # Aggiorna l'interfaccia
            self.refresh_calendar()
            
            # Mostra un messaggio all'utente
            messagebox.showinfo("Pianificazione completata", 
                              f"Pianificazione completata con successo!\n\n"
                              f"Sono stati pianificati {scheduled_count} allenamenti.")
            
        except Exception as e:
            self.log(f"Errore nella pianificazione con le date dal YAML: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore nella pianificazione: {str(e)}")


    def check_workouts_in_cache(self, training_plan_id):
        """Verifica se gli allenamenti del piano sono presenti nella cache di Garmin Connect."""
        # Aggiorna la cache se necessario
        if not os.path.exists(WORKOUTS_CACHE_FILE):
            self.log("La cache degli allenamenti non esiste. Aggiornamento in corso...")
            self.refresh_workouts()
            return
        
        try:
            with open(WORKOUTS_CACHE_FILE, 'r') as f:
                cache = json.load(f)
                
            # Cerca allenamenti che corrispondono al piano
            matching_workouts = []
            for workout in cache:
                if training_plan_id in workout.get('workoutName', ''):
                    matching_workouts.append(workout)
            
            if matching_workouts:
                self.log(f"Trovati {len(matching_workouts)} allenamenti per il piano {training_plan_id} in Garmin Connect")
                self.plan_imported = True
            else:
                self.log(f"Non sono stati trovati allenamenti per il piano {training_plan_id} in Garmin Connect")
                
                # Chiedi all'utente se vuole importare gli allenamenti
                yaml_path = self.find_yaml_for_plan(training_plan_id)
                
                if yaml_path and messagebox.askyesno("Importazione necessaria", 
                                              f"Non sono stati trovati allenamenti per il piano {training_plan_id} in Garmin Connect.\n\n"
                                              f"Vuoi importare gli allenamenti dal file {yaml_path} prima di pianificarli?"):
                    self.log(f"Importazione degli allenamenti da {yaml_path}...")
                    self.import_file.set(yaml_path)
                    self.import_replace.set(True)  # Sostituisci eventuali allenamenti esistenti
                    
                    # Eseguiamo l'importazione e attendiamo il completamento
                    self._do_import_for_schedule()
                    self.plan_imported = True
        except Exception as e:
            self.log(f"Errore durante il controllo della cache: {str(e)}")

    def _do_import_and_schedule(self, selected_days):
        """Importa gli allenamenti e poi pianifica"""
        try:
            self.log("Importazione e pianificazione degli allenamenti in corso...")
            
            # Assicurati che la cartella OAuth esista
            oauth_folder = self.oauth_folder.get()
            if not os.path.exists(oauth_folder):
                try:
                    os.makedirs(oauth_folder, exist_ok=True)
                    self.log(f"Creata cartella OAuth: {oauth_folder}")
                except Exception as e:
                    self.log(f"Errore nella creazione della cartella OAuth: {str(e)}")
                    messagebox.showerror("Errore", f"Errore nella creazione della cartella OAuth: {str(e)}")
                    return
            
            # Esegui l'importazione usando la funzione diretta
            class ImportArgs:
                pass
                    
            import_args = ImportArgs()
            import_args.oauth_folder = oauth_folder
            import_args.workouts_file = self.import_file.get()
            import_args.name_filter = None
            import_args.replace = self.import_replace.get()
            import_args.dry_run = False
            import_args.treadmill = False
            
            self.log(f"Importazione degli allenamenti da {self.import_file.get()}")
            cmd_import_workouts(import_args)
            
            self.log("Importazione completata con successo")
            
            # Aggiorna la lista degli allenamenti in background
            self.log("Aggiornamento cache degli allenamenti...")
            refresh_thread = threading.Thread(target=self._refresh_workouts_silent)
            refresh_thread.start()
            refresh_thread.join()  # Attendi il completamento
            
            # Imposta il flag di piano importato
            self.plan_imported = True
            
            # Ora procedi con la pianificazione
            self.log("Procedo con la pianificazione degli allenamenti...")
            self._do_schedule(selected_days, False)  # False = non è dry run
            
        except Exception as e:
            self.log(f"Errore durante l'importazione e pianificazione: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore durante l'importazione e pianificazione: {str(e)}")


    def plan_yaml_workouts(self, yaml_path, start_date, race_date, selected_days):
        """Pianifica gli allenamenti dal file YAML procedendo a ritroso dalla data della gara"""
        try:
            today = datetime.today().date()
            
            # Carica il file YAML
            with open(yaml_path, 'r') as f:
                plan_data = yaml.safe_load(f)
                
                # Rimuovi la configurazione
                config = plan_data.pop('config', {})
                
                # Estrai e organizza gli allenamenti per settimana e sessione
                workouts_by_week = {}
                for name, _ in plan_data.items():
                    match = re.search(r'(W\d\d)S(\d\d)', name)
                    if match:
                        week_id = match.group(1)
                        week_num = int(week_id.replace('W', ''))
                        session_id = int(match.group(2).replace('S', ''))
                        
                        if week_id not in workouts_by_week:
                            workouts_by_week[week_id] = {}
                        
                        workouts_by_week[week_id][session_id] = name
                
                # Verifica presenza di allenamenti
                if not workouts_by_week:
                    self.log("Nessun allenamento trovato nel file YAML.")
                    messagebox.showinfo("Simulazione", "Nessun allenamento trovato nel file YAML.")
                    return
                
                # Verifica che ci siano giorni selezionati
                if not selected_days:
                    messagebox.showwarning("Nessun giorno selezionato", "Seleziona almeno un giorno della settimana.")
                    return
                
                # Ordina giorni selezionati
                selected_days = sorted([int(d) for d in selected_days])
                self.log(f"Giorni selezionati: {[calendar.day_name[d] for d in selected_days]}")
                self.log(f"Data della gara: {race_date.strftime('%d/%m/%Y')} ({calendar.day_name[race_date.weekday()]})")
                
                # Converti la data della gara in stringa una volta sola per i confronti
                race_date_str = race_date.strftime('%Y-%m-%d')
                
                # Verifica se il giorno della gara è tra i giorni selezionati
                if race_date.weekday() in selected_days:
                    self.log(f"Il giorno della gara ({calendar.day_name[race_date.weekday()]}) è uno dei giorni selezionati")
                
                # Calcola il lunedì della settimana della gara
                days_to_monday = race_date.weekday()  # 0=lunedì, 1=martedì, ecc.
                race_week_monday = race_date - timedelta(days=days_to_monday)
                self.log(f"Lunedì della settimana della gara: {race_week_monday.strftime('%d/%m/%Y')}")
                
                # Pianificazione
                scheduled = []
                used_dates = set()
                
                # Ordina le settimane in ordine decrescente (dalla settimana della gara a ritroso)
                weeks = sorted(workouts_by_week.keys(), reverse=True)
                
                # Trova la settimana massima
                max_week = max([int(w.replace('W', '')) for w in weeks])
                
                # Pianifica a ritroso dalla settimana più alta (solitamente quella della gara)
                for week_id in weeks:
                    # Estrai il numero della settimana
                    week_num = int(week_id.replace('W', ''))
                    
                    # Calcola offset dalla settimana della gara
                    week_offset = max_week - week_num
                    
                    # Calcola il lunedì di questa settimana
                    week_monday = race_week_monday - timedelta(weeks=week_offset)
                    self.log(f"Simulazione settimana {week_id}: inizia il {week_monday.strftime('%d/%m/%Y')} (Lunedì)")
                    
                    # Ordina le sessioni
                    sessions = sorted(workouts_by_week[week_id].keys())
                    
                    # Assegna date alle sessioni
                    for session_idx, session_id in enumerate(sessions):
                        workout_name = workouts_by_week[week_id][session_id]
                        
                        # Determina il giorno della settimana
                        if session_idx < len(selected_days):
                            day_idx = selected_days[session_idx]
                        else:
                            # Se ci sono più sessioni che giorni selezionati, cicla
                            day_idx = selected_days[session_idx % len(selected_days)]
                        
                        # Calcola la data
                        workout_date = week_monday + timedelta(days=day_idx)
                        date_str = workout_date.strftime("%Y-%m-%d")
                        
                        # DEBUG: Mostra cosa stiamo valutando
                        self.log(f"DEBUG: Valutando allenamento {workout_name} per data {date_str}")
                        
                        # ===== CONTROLLO CRITICO: verifica se la data coincide con il giorno della gara =====
                        if date_str == race_date_str:
                            self.log(f"Allenamento {workout_name} coinciderebbe con il giorno della gara ({date_str}). SALTATO.")
                            continue  # Salta questo allenamento e passa al successivo
                        
                        # Salta date nel passato
                        if workout_date < today:
                            self.log(f"Allenamento {workout_name} cadrebbe nel passato ({date_str}). Saltato.")
                            continue
                            
                        # Salta date dopo la gara
                        if workout_date > race_date:
                            self.log(f"Allenamento {workout_name} cadrebbe dopo la gara ({date_str}). Saltato.")
                            continue
                        
                        # Verifica se la data è già utilizzata
                        if date_str in used_dates:
                            # Cerca una data alternativa tra i giorni selezionati
                            self.log(f"Data {date_str} già utilizzata. Cercando un'alternativa...")
                            found_alternative = False
                            
                            for alt_day in [d for d in selected_days if d != day_idx]:
                                alt_date = week_monday + timedelta(days=alt_day)
                                alt_date_str = alt_date.strftime("%Y-%m-%d")
                                
                                # Verifica che non sia la data della gara
                                if alt_date_str == race_date_str:
                                    self.log(f"Alternativa {alt_date_str} coincide con la gara. Saltata.")
                                    continue
                                    
                                if (alt_date <= race_date and 
                                    alt_date >= today and 
                                    alt_date_str not in used_dates):
                                    workout_date = alt_date
                                    date_str = alt_date_str
                                    found_alternative = True
                                    self.log(f"Trovata data alternativa: {workout_date.strftime('%d/%m/%Y')}")
                                    break
                            
                            if not found_alternative:
                                self.log(f"Nessuna data alternativa disponibile per {workout_name}. Saltato.")
                                continue
                        
                        # Aggiungi alla pianificazione
                        used_dates.add(date_str)
                        scheduled.append({
                            'date': date_str,
                            'workout': workout_name + " (SIMULAZIONE)"
                        })
                        
                        day_name = calendar.day_name[workout_date.weekday()]
                        self.log(f"{workout_name} pianificato per {day_name} {date_str}")
                
                # Ordina la pianificazione per data
                scheduled.sort(key=lambda x: x['date'])
                
                # DEBUG: Mostra il contenuto finale di scheduled prima di visualizzarlo
                self.log(f"DEBUG: Contenuto finale della pianificazione ({len(scheduled)} allenamenti):")
                for item in scheduled:
                    self.log(f"DEBUG: Pianificato: {item['date']} - {item['workout']}")
                
                # Visualizza nella tabella del calendario
                self.display_yaml_schedule(scheduled)
                
                # Mostra un messaggio all'utente
                messagebox.showinfo("Simulazione completata", 
                                  f"Simulazione pianificazione completata con successo.\n"
                                  f"Sono stati pianificati {len(scheduled)} allenamenti.\n"
                                  f"Questi allenamenti sono visualizzati nella tabella con l'etichetta (SIMULAZIONE).")
        
        except Exception as e:
            self.log(f"Errore nella pianificazione: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Si è verificato un errore durante la pianificazione: {str(e)}")


    def display_yaml_schedule(self, scheduled):
        """Visualizza la pianificazione nella tabella del calendario"""
        try:
            # Pulisci la tabella esistente
            for item in self.calendar_tree.get_children():
                self.calendar_tree.delete(item)
            
            # Aggiungi gli allenamenti pianificati con debug
            for item in scheduled:
                # Debug per vedere esattamente quali allenamenti stiamo visualizzando
                self.log(f"DEBUG: Visualizzo da YAML: {item['date']} - {item['workout']}")
                
                date = item['date']
                workout = item['workout']
                self.calendar_tree.insert("", "end", values=(date, workout))
        
        except Exception as e:
            self.log(f"Errore nella visualizzazione della pianificazione: {str(e)}")
            import traceback
            traceback.print_exc()




    def find_yaml_for_plan(self, plan_id):
        """Cerca il file YAML corrispondente al piano specificato"""
        try:
            self.log(f"Ricerca del file YAML per il piano: '{plan_id}'")
            
            # Normalizza il plan_id per la ricerca
            plan_id_normalized = plan_id.lower().strip()
            
            # Controlla se il file YAML è già stato selezionato nella tab Importa
            import_file = self.import_file.get()
            if import_file and os.path.exists(import_file) and import_file.endswith(('.yaml', '.yml')):
                self.log(f"Controllando il file selezionato: {import_file}")
                # Verifica se il file contiene il piano
                try:
                    with open(import_file, 'r', encoding='utf-8') as f:
                        data = yaml.safe_load(f)
                        if 'config' in data and 'name_prefix' in data['config']:
                            name_prefix = data['config']['name_prefix'].strip().lower()
                            
                            # Verifica se corrisponde al piano cercato
                            self.log(f"Confronto: '{name_prefix}' con '{plan_id_normalized}'")
                            if plan_id_normalized in name_prefix or name_prefix in plan_id_normalized:
                                self.log(f"Corrispondenza trovata! Usando il file YAML già selezionato: {import_file}")
                                
                                # Se c'è una race_day, impostala
                                if 'race_day' in data['config']:
                                    race_day_str = data['config']['race_day']
                                    self.log(f"Trovata data della gara nel YAML: {race_day_str}")
                                    self.race_day.set(race_day_str)
                                
                                # Salva il percorso per riferimento futuro
                                self.current_yaml_path = import_file
                                
                                return import_file
                            else:
                                self.log(f"Prefisso {name_prefix} non corrisponde a {plan_id_normalized}")
                except Exception as e:
                    self.log(f"Errore nella lettura del file {import_file}: {str(e)}")
            
            # Cerca nelle directory standard
            search_dirs = [
                os.path.join(SCRIPT_DIR, "training_plans"),  # Directory principale training_plans
                SCRIPT_DIR,  # Directory principale dello script
                os.getcwd()  # Directory di lavoro corrente
            ]
            
            self.log(f"Cercando nelle directory: {search_dirs}")
            
            for search_dir in search_dirs:
                if os.path.exists(search_dir):
                    self.log(f"Cercando in: {search_dir}")
                    # Cerca ricorsivamente nei file YAML
                    for root, _, files in os.walk(search_dir):
                        for file in files:
                            if file.endswith(('.yaml', '.yml')):
                                file_path = os.path.join(root, file)
                                self.log(f"Trovato file YAML: {file_path}, verifico se contiene il piano")
                                
                                # Controlla se il file contiene il piano
                                try:
                                    with open(file_path, 'r', encoding='utf-8') as f:
                                        data = yaml.safe_load(f)
                                        if 'config' in data and 'name_prefix' in data['config']:
                                            name_prefix = data['config']['name_prefix'].strip().lower()
                                            
                                            # Verifica se corrisponde al piano cercato
                                            self.log(f"Confronto: '{name_prefix}' con '{plan_id_normalized}'")
                                            if plan_id_normalized in name_prefix or name_prefix in plan_id_normalized:
                                                self.log(f"Corrispondenza trovata! File YAML per il piano '{plan_id}': {file_path}")
                                                
                                                # Se c'è una race_day, impostala
                                                if 'race_day' in data['config']:
                                                    race_day_str = data['config']['race_day']
                                                    self.log(f"Trovata data della gara nel YAML: {race_day_str}")
                                                    self.race_day.set(race_day_str)
                                                
                                                # Salva il percorso per riferimento futuro
                                                self.current_yaml_path = file_path
                                                
                                                return file_path
                                            else:
                                                self.log(f"Prefisso {name_prefix} non corrisponde a {plan_id_normalized}")
                                except Exception as e:
                                    self.log(f"Errore nella lettura del file {file_path}: {str(e)}")
            
            # Se non trova automaticamente, chiede all'utente di selezionare un file
            if messagebox.askyesno("File YAML non trovato", 
                                  f"Non è stato trovato un file YAML per il piano '{plan_id}'.\n\n"
                                  f"Vuoi selezionare manualmente un file YAML?"):
                file_path = filedialog.askopenfilename(
                    title="Seleziona file YAML",
                    filetypes=[("YAML files", "*.yaml *.yml"), ("All files", "*.*")]
                )
                if file_path:
                    self.log(f"File YAML selezionato manualmente: {file_path}")
                    
                    # Verifica se contiene race_day
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = yaml.safe_load(f)
                            if 'config' in data and 'race_day' in data['config']:
                                race_day_str = data['config']['race_day']
                                self.log(f"Trovata data della gara nel YAML: {race_day_str}")
                                self.race_day.set(race_day_str)
                    except Exception as e:
                        self.log(f"Errore nella lettura del file {file_path}: {str(e)}")
                    
                    # Salva il percorso per riferimento futuro
                    self.current_yaml_path = file_path
                    
                    return file_path
            
            self.log(f"Nessun file YAML trovato per il piano '{plan_id}'")
            return None
            
        except Exception as e:
            self.log(f"Errore nella ricerca del file YAML: {str(e)}")
            return None


    def _display_original_dates(self):
        """Visualizza le date originali dal file YAML nella tabella del calendario"""
        try:
            # Pulisci la tabella esistente
            for item in self.calendar_tree.get_children():
                self.calendar_tree.delete(item)
            
            # Verifica che abbiamo le date originali
            if not hasattr(self, 'original_workout_dates') or not self.original_workout_dates:
                self.log("Nessuna data originale trovata nel file YAML")
                return
            
            # Prepara una lista di allenamenti con date per l'ordinamento
            workout_list = []
            for workout_name, date_str in self.original_workout_dates.items():
                workout_list.append({
                    'date': date_str,
                    'title': workout_name + " (DA YAML)",
                })
            
            # Ordina per data
            workout_list.sort(key=lambda x: x['date'])
            
            # Visualizza nella tabella
            for workout in workout_list:
                self.calendar_tree.insert("", "end", values=(workout['date'], workout['title']))
            
            # Log e messaggio all'utente
            self.log(f"Visualizzate {len(workout_list)} date originali dal file YAML")
            messagebox.showinfo("Date originali", 
                              f"Sono state visualizzate {len(workout_list)} date originali dal file YAML.\n"
                              f"Queste date verranno utilizzate per la pianificazione effettiva se non modifichi\n"
                              f"la data della gara o i giorni preferiti.")
        
        except Exception as e:
            self.log(f"Errore nella visualizzazione delle date originali: {str(e)}")
            import traceback
            traceback.print_exc()


    def _update_yaml_dates(self, yaml_path, race_day, selected_days):
        """Aggiorna il file YAML con la nuova data di gara e ricalcola le date degli allenamenti"""
        try:
            # Carica il file YAML
            with open(yaml_path, 'r') as f:
                plan_data = yaml.safe_load(f)
            
            # Aggiorna la data della gara
            if 'config' not in plan_data:
                plan_data['config'] = {}
            
            plan_data['config']['race_day'] = race_day
            plan_data['config']['preferred_days'] = selected_days
            
            # Ora ricalcola tutte le date degli allenamenti
            race_date = datetime.strptime(race_day, "%Y-%m-%d").date()
            today = datetime.today().date()
            
            # Trova il lunedì della settimana della gara
            days_to_monday = race_date.weekday()
            race_week_monday = race_date - timedelta(days=days_to_monday)
            
            # Organizziamo gli allenamenti per settimana e sessione
            workouts_by_week = {}
            for workout_name, steps in list(plan_data.items()):
                if workout_name == 'config':
                    continue
                    
                match = re.search(r'(W\d\d)S(\d\d)', workout_name)
                if match:
                    week_id = match.group(1)
                    session_id = int(match.group(2))
                    week_num = int(week_id[1:])
                    
                    if week_id not in workouts_by_week:
                        workouts_by_week[week_id] = {}
                    
                    workouts_by_week[week_id][session_id] = {
                        'name': workout_name,
                        'steps': steps
                    }
            
            # Calcola la settimana massima
            max_week = 0
            for week_id in workouts_by_week:
                week_num = int(week_id[1:])
                max_week = max(max_week, week_num)
            
            # Pianifica le date
            used_dates = set()
            assigned_dates = {}
            
            # Pianifica a ritroso dalla settimana più alta
            weeks = sorted(workouts_by_week.keys(), reverse=True)  # Ordina in ordine decrescente
            
            for week_id in weeks:
                week_num = int(week_id[1:])
                
                # Calcola l'offset rispetto alla settimana della gara
                week_offset = max_week - week_num
                
                # Calcola il lunedì di questa settimana
                week_monday = race_week_monday - timedelta(weeks=week_offset)
                
                # Ordina le sessioni
                sessions = sorted(workouts_by_week[week_id].keys())
                
                # Assegna le date
                for session_idx, session_id in enumerate(sessions):
                    workout_info = workouts_by_week[week_id][session_id]
                    
                    # Determina il giorno della settimana
                    if session_idx < len(selected_days):
                        day_idx = selected_days[session_idx]
                    else:
                        # Se ci sono più sessioni che giorni selezionati, cicla
                        day_idx = selected_days[session_idx % len(selected_days)]
                    
                    # Calcola la data
                    workout_date = week_monday + timedelta(days=day_idx)
                    date_str = workout_date.strftime("%Y-%m-%d")
                    
                    # Verifica se coincide con il giorno della gara
                    if workout_date.date() == race_date:
                        self.log(f"Allenamento {workout_info['name']} coinciderebbe con il giorno della gara. Saltato.")
                        continue
                    
                    # Verifica se è nel passato
                    if workout_date.date() < today:
                        self.log(f"Allenamento {workout_info['name']} cadrebbe nel passato. Saltato.")
                        continue
                    
                    # Verifica se è dopo la gara
                    if workout_date.date() > race_date:
                        self.log(f"Allenamento {workout_info['name']} cadrebbe dopo la gara. Saltato.")
                        continue
                    
                    # Verifica se la data è già utilizzata
                    if date_str in used_dates:
                        # Cerca una data alternativa
                        found_alternative = False
                        
                        for alt_day in [d for d in selected_days if d != day_idx]:
                            alt_date = week_monday + timedelta(days=alt_day)
                            alt_date_str = alt_date.strftime("%Y-%m-%d")
                            
                            # Verifica che l'alternativa sia valida
                            if (alt_date.date() != race_date and
                                alt_date.date() <= race_date and
                                alt_date.date() >= today and
                                alt_date_str not in used_dates):
                                workout_date = alt_date
                                date_str = alt_date_str
                                found_alternative = True
                                self.log(f"Trovata data alternativa per {workout_info['name']}: {date_str}")
                                break
                        
                        if not found_alternative:
                            self.log(f"Nessuna data alternativa disponibile per {workout_info['name']}. Saltato.")
                            continue
                    
                    # Aggiungi alla pianificazione
                    used_dates.add(date_str)
                    assigned_dates[workout_info['name']] = date_str
                    
                    # Aggiorna i passi dell'allenamento (aggiungi o aggiorna la data)
                    steps = workout_info['steps']
                    date_added = False
                    
                    if steps and isinstance(steps, list):
                        # Verifica se c'è già un elemento date
                        for i, step in enumerate(steps):
                            if isinstance(step, dict) and 'date' in step:
                                # Aggiorna la data esistente
                                steps[i]['date'] = date_str
                                date_added = True
                                break
                        
                        # Se non c'è una data, aggiungi come primo elemento
                        if not date_added:
                            steps.insert(0, {'date': date_str})
            
            # Salva il file YAML aggiornato
            with open(yaml_path, 'w', encoding='utf-8') as f:
                yaml.dump(plan_data, f, default_flow_style=False, sort_keys=False, Dumper=NoAliasDumper)
            
            self.log(f"File YAML aggiornato con successo: {len(assigned_dates)} date pianificate")
            
            # Aggiorna le date originali
            self.original_workout_dates = assigned_dates
            self.original_race_day = race_day
            self.original_preferred_days = selected_days
            
            return True
            
        except Exception as e:
            self.log(f"Errore nell'aggiornamento del file YAML: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            return False


    def _simulate_schedule(self, selected_days):
        """Simula la pianificazione degli allenamenti senza apportare modifiche"""
        try:
            # Recupera i parametri necessari
            training_plan_id = self.training_plan.get().strip()
            race_day_str = self.race_day.get()  # Usa direttamente race_day
            race_day = datetime.strptime(race_day_str, "%Y-%m-%d").date()
            today = datetime.today().date()
            
            # Log per debug
            self.log(f"DEBUG: Data della gara (stringa): {race_day_str}")
            self.log(f"DEBUG: Data della gara (oggetto): {race_day}")
            self.log(f"DEBUG: Formato della data: {race_day.strftime('%Y-%m-%d')}")
            
            # Verifica che la data della gara sia valida
            if race_day < today:
                self.log("Errore: La data della gara è nel passato.")
                messagebox.showerror("Errore", "La data della gara deve essere nel futuro.")
                return
            
            # Verifica se abbiamo date originali nel file YAML
            has_original_dates = False
            use_original_dates = False
            
            # Controlla se è stata modificata la data della gara o i giorni selezionati
            date_modified = False
            days_modified = False
            
            if hasattr(self, 'original_race_day') and self.original_race_day:
                if self.original_race_day != race_day_str:
                    self.log(f"Data gara modificata: {self.original_race_day} -> {race_day_str}")
                    date_modified = True
            
            if hasattr(self, 'original_preferred_days') and self.original_preferred_days:
                current_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
                if set(self.original_preferred_days) != set(current_days):
                    self.log(f"Giorni preferiti modificati: {self.original_preferred_days} -> {current_days}")
                    days_modified = True
            
            # Se abbiamo date originali nei workout e non sono state fatte modifiche, possiamo usarle
            if hasattr(self, 'original_workout_dates') and self.original_workout_dates:
                has_original_dates = True
                
                # Usa le date originali solo se non sono state fatte modifiche rilevanti
                if not date_modified and not days_modified:
                    use_original_dates = True
                    self.log("Utilizzo le date originali dal file YAML per la simulazione")
            
            # Se dobbiamo usare le date originali dal file YAML
            if use_original_dates:
                self._display_original_dates()
                return
            
            # Inizializza struttura per gli allenamenti
            workouts = []
            workout_infos = {}
            
            # Verifica se abbiamo un file YAML caricato
            yaml_plan_loaded = False
            yaml_path = None
            
            if hasattr(self, 'current_yaml_path') and self.current_yaml_path and os.path.exists(self.current_yaml_path):
                yaml_path = self.current_yaml_path
                yaml_plan_loaded = True
                self.log(f"File YAML trovato: {yaml_path}")
            else:
                # Cerca il file YAML corrispondente al piano
                yaml_path = self.find_yaml_for_plan(training_plan_id)
                if yaml_path:
                    yaml_plan_loaded = True
                    self.log(f"File YAML trovato per il piano {training_plan_id}: {yaml_path}")
            
            # Se abbiamo un file YAML, lo usiamo come fonte primaria
            if yaml_plan_loaded:
                try:
                    with open(yaml_path, 'r') as f:
                        plan_data = yaml.safe_load(f)
                        
                        # Rimuovi la configurazione
                        config = plan_data.pop('config', {}) if 'config' in plan_data else {}
                        
                        # Crea una struttura di allenamenti dal YAML
                        for workout_name, steps in plan_data.items():
                            match = re.search(r'(W\d\d)S(\d\d)', workout_name)
                            if match:
                                week_id = match.group(1)
                                session_id = int(match.group(2).replace('S', ''))
                                week_num = int(week_id.replace('W', ''))
                                
                                if week_id not in workout_infos:
                                    workout_infos[week_id] = {}
                                
                                workout_infos[week_id][session_id] = {
                                    'id': f"yaml-{workout_name.replace(' ', '-')}",  # ID fittizio
                                    'name': workout_name,
                                    'week_num': week_num
                                }
                                
                                # Estrai la data se presente
                                if steps and isinstance(steps, list) and len(steps) > 0:
                                    first_step = steps[0]
                                    if isinstance(first_step, dict) and 'date' in first_step:
                                        date_str = first_step['date']
                                        workout_infos[week_id][session_id]['existing_date'] = date_str
                                        self.log(f"Trovata data esistente {date_str} per {workout_name}")
                    
                    self.log(f"Utilizzando {len(workout_infos)} settimane dal file YAML")
                except Exception as e:
                    self.log(f"Errore nella lettura del file YAML: {str(e)}")
            
            # Se ancora non abbiamo allenamenti, cerchiamo nella cache
            if not workout_infos:
                self.log("Nessun allenamento trovato nel YAML, tentativo di utilizzo della cache...")
                if os.path.exists(WORKOUTS_CACHE_FILE):
                    with open(WORKOUTS_CACHE_FILE, 'r') as f:
                        all_workouts = json.load(f)
                        
                        # Filtra per il piano di allenamento
                        plan_workouts = []
                        for workout in all_workouts:
                            workout_name = workout.get('workoutName', '')
                            if training_plan_id in workout_name:
                                plan_workouts.append(workout)
                        
                        # Estrai informazioni settimana/sessione
                        for workout in plan_workouts:
                            workout_name = workout.get('workoutName', '')
                            workout_id = workout.get('workoutId', '')
                            
                            match = re.search(r'\s(W\d\d)S(\d\d)\s', workout_name)
                            if match:
                                week_id = match.group(1)
                                session_id = int(match.group(2))
                                week_num = int(week_id[1:])
                                
                                if week_id not in workout_infos:
                                    workout_infos[week_id] = {}
                                workout_infos[week_id][session_id] = {
                                    'id': workout_id,
                                    'name': workout_name,
                                    'week_num': week_num
                                }
            
            # Se non ci sono ancora allenamenti, non possiamo simulare
            if not workout_infos:
                self.log("Nessun allenamento trovato per la simulazione")
                messagebox.showinfo("Simulazione", 
                                  f"Nessun allenamento trovato per il piano '{training_plan_id}'.\n\n"
                                  f"Importa prima gli allenamenti per poter simulare la pianificazione.")
                return
            
            # Verifica che ci siano giorni selezionati e convertili in numeri
            if hasattr(self, 'day_selections'):
                selected_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
            else:
                selected_days = [i for i, var in enumerate(self.day_selections) if var.get() == 1]
                
            if not selected_days:
                messagebox.showwarning("Nessun giorno selezionato", "Seleziona almeno un giorno della settimana.")
                return
                
            # Ordina i giorni selezionati
            selected_days = sorted(selected_days)
            day_names = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
            self.log(f"Giorni selezionati: {[day_names[d] for d in selected_days]}")
            self.log(f"Data della gara: {race_day.strftime('%d/%m/%Y')} ({day_names[race_day.weekday()]})")
            
            # Trova il lunedì della settimana della gara
            days_to_monday = race_day.weekday()  # 0=lunedì, 1=martedì, ecc.
            race_week_monday = race_day - timedelta(days=days_to_monday)
            self.log(f"Lunedì della settimana della gara: {race_week_monday.strftime('%d/%m/%Y')}")
            
            # Ottieni le settimane in ordine decrescente (dall'ultima alla prima)
            weeks = sorted(workout_infos.keys(), reverse=True)
            
            # Trova la settimana massima per calcolare l'offset corretto
            max_week = max([int(w[1:]) for w in weeks])
            
            # Ora simula la pianificazione
            simulated_schedule = []
            dates_used = set()
            
            # IMPORTANTE: Memorizza la data della gara come stringa una volta sola per il confronto
            race_date_str = race_day.strftime('%Y-%m-%d')
            self.log(f"DEBUG: Data della gara per confronto: {race_date_str}")
            
            for week_id in weeks:
                # Estrai il numero della settimana (da W01, W02, ecc.)
                week_num = int(week_id[1:])
                
                # Calcola l'offset rispetto alla settimana della gara
                week_offset = max_week - week_num
                
                # Calcola il lunedì di questa settimana
                week_monday = race_week_monday - timedelta(weeks=week_offset)
                self.log(f"Simulazione settimana {week_id}: inizia il {week_monday.strftime('%d/%m/%Y')} (Lunedì)")
                
                # Ordina le sessioni (S01, S02, ecc.)
                sessions = sorted(workout_infos[week_id].keys())
                
                # Assegna allenamenti ai giorni selezionati
                for session_idx, session_id in enumerate(sessions):
                    workout_info = workout_infos[week_id][session_id]
                    
                    # Controlla se questo allenamento ha già una data assegnata
                    if 'existing_date' in workout_info:
                        existing_date = workout_info['existing_date']
                        self.log(f"Allenamento {workout_info['name']} ha già data {existing_date}")
                        
                        # Usa la data esistente
                        simulated_schedule.append({
                            'date': existing_date,
                            'title': workout_info['name'],
                            'id': workout_info['id'],
                            'day': "Data esistente"
                        })
                        dates_used.add(existing_date)
                        continue
                    
                    # Determina il giorno della settimana
                    if session_idx < len(selected_days):
                        day_idx = selected_days[session_idx]
                    else:
                        # Se ci sono più sessioni che giorni selezionati, cicla
                        day_idx = selected_days[session_idx % len(selected_days)]
                    
                    # Calcola la data per questa sessione
                    workout_date = week_monday + timedelta(days=day_idx)
                    date_str = workout_date.strftime('%Y-%m-%d')
                    
                    self.log(f"DEBUG: Valutando {week_id}S{session_id} per la data {date_str}")
                    
                    # ===== VERIFICA IMPORTANTE: controlla se coincide con la data della gara =====
                    if date_str == race_date_str:
                        self.log(f"DEBUG: MATCH! {date_str} == {race_date_str}")
                        self.log(f"Allenamento {week_id}S{session_id:02d} coinciderebbe con il giorno della gara ({date_str}). Saltato.")
                        continue
                    
                    # Salta date nel passato
                    if workout_date < today:
                        self.log(f"Allenamento {week_id}S{session_id:02d} cadrebbe nel passato ({date_str}). Saltato.")
                        continue
                        
                    # Salta date dopo la gara
                    if workout_date > race_day:
                        self.log(f"Allenamento {week_id}S{session_id:02d} cadrebbe dopo la gara ({date_str}). Saltato.")
                        continue
                    
                    # Verifica se la data è già utilizzata
                    if date_str in dates_used:
                        # Cerca una data alternativa tra i giorni selezionati
                        self.log(f"Data {date_str} già utilizzata. Cercando un'alternativa...")
                        found_alternative = False
                        
                        for alt_day in [d for d in selected_days if d != day_idx]:
                            alt_date = week_monday + timedelta(days=alt_day)
                            alt_date_str = alt_date.strftime('%Y-%m-%d')
                            
                            self.log(f"DEBUG: Considerando alternativa {alt_date_str}")
                            
                            # Confronta stringhe per verificare che non sia la data della gara
                            if alt_date_str == race_date_str:
                                self.log(f"DEBUG: Alternativa {alt_date_str} coincide con la gara. Saltata.")
                                continue
                                
                            if (alt_date <= race_day and 
                                alt_date >= today and 
                                alt_date_str not in dates_used):
                                workout_date = alt_date
                                date_str = alt_date_str
                                found_alternative = True
                                self.log(f"Trovata data alternativa: {workout_date.strftime('%d/%m/%Y')}")
                                break
                        
                        if not found_alternative:
                            self.log(f"Nessuna data alternativa disponibile per {week_id}S{session_id:02d}. Saltato.")
                            continue
                    
                    # Aggiungi alla simulazione
                    dates_used.add(date_str)
                    simulated_schedule.append({
                        'date': date_str,
                        'title': workout_info['name'],
                        'id': workout_info['id'],
                        'day': day_names[workout_date.weekday()]
                    })
                    
                    self.log(f"SIMULAZIONE: Allenamento {week_id}S{session_id:02d} pianificato per {day_names[workout_date.weekday()]} {date_str}")
            
            # Ordina per data
            simulated_schedule.sort(key=lambda x: x['date'])
            
            # Visualizza gli allenamenti simulati
            self._display_simulated_workouts(simulated_schedule)
            
        except Exception as e:
            self.log(f"Errore durante la simulazione: {str(e)}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Errore", f"Errore durante la simulazione: {str(e)}")

            
    def _do_schedule(self, selected_days, is_dry_run=False):
        """Esegue il comando di pianificazione degli allenamenti"""
        # Assicurati che la cartella OAuth esista
        oauth_folder = self.oauth_folder.get()
        if not os.path.exists(oauth_folder):
            try:
                os.makedirs(oauth_folder, exist_ok=True)
                self.log(f"Creata cartella OAuth: {oauth_folder}")
            except Exception as e:
                self.log(f"Errore nella creazione della cartella OAuth: {str(e)}")
                return
        
        # Se siamo in modalità dry-run, usiamo un approccio diverso
        if is_dry_run:
            self._simulate_schedule(selected_days)
            return
                    
        try:
            self.log("Pianificazione allenamenti in corso...")
            
            # Crea un oggetto args simulato
            class Args:
                pass
                    
            args = Args()
            args.oauth_folder = self.oauth_folder.get()
            args.training_plan = self.training_plan.get()
            args.race_day = self.race_day.get()
            args.workout_days = ",".join([str(day) for day in selected_days])
            args.dry_run = False
            
            # Se è specificata una data di inizio, aggiungila
            if hasattr(self, 'start_day') and self.start_day.get():
                args.start_day = self.start_day.get()
            else:
                args.start_day = None
            
            # Esegui la funzione direttamente
            cmd_schedule_workouts(args)
            
            self.log("Pianificazione completata con successo")
            messagebox.showinfo("Successo", "Pianificazione completata con successo")
            
            # Aggiorna il calendario
            self.refresh_calendar()
            
        except Exception as e:
            self.log(f"Errore durante la pianificazione: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore durante la pianificazione: {str(e)}")

            
    def _parse_schedule_output(self, output):
            """Estrai le informazioni sugli allenamenti pianificati dall'output del comando"""
            try:
                workouts = []
                # Cerca linee come "Scheduling workout XXX (YYY) on ZZZZ-MM-DD"
                pattern = r"Scheduling workout (.*?) \((.*?)\) on (\d{4}-\d{2}-\d{2})"
                matches = re.findall(pattern, output)
                
                for match in matches:
                    workout_name, workout_id, date = match
                    workouts.append({
                        'date': date,
                        'title': workout_name,
                        'id': workout_id
                    })
                    
                return workouts
            except Exception as e:
                self.log(f"Errore nel parsing dell'output di pianificazione: {str(e)}")
                return None

    def _display_simulated_workouts(self, workouts):
        """Visualizza gli allenamenti simulati nella tabella del calendario"""
        try:
            # Pulisci la tabella del calendario
            for item in self.step4_calendar_tree.get_children():
                self.step4_calendar_tree.delete(item)
                    
            # Aggiungi gli allenamenti simulati
            for workout in workouts:
                # Debug per vedere esattamente quali allenamenti stiamo visualizzando
                self.log(f"DEBUG: Visualizzo allenamento simulato: {workout['date']} - {workout['title']}")
                
                # Estrai informazioni e aggiungi (SIMULAZIONE) al titolo
                date = workout['date']
                title = workout['title'] + " (SIMULAZIONE)"
                
                # Inserisci nella tabella
                self.step4_calendar_tree.insert("", "end", values=(date, title))
                    
            self.log(f"Visualizzati {len(workouts)} allenamenti simulati nel calendario")
            
            # Mostra un messaggio all'utente
            if workouts:
                messagebox.showinfo("Simulazione completata", 
                                 f"Simulazione pianificazione completata.\n"
                                 f"Sono stati pianificati {len(workouts)} allenamenti.\n"
                                 f"Questi allenamenti sono visualizzati nella tabella con l'etichetta (SIMULAZIONE).")
            else:
                messagebox.showinfo("Simulazione", "Nessun allenamento da pianificare con i criteri specificati.")
        
        except Exception as e:
            self.log(f"Errore nella visualizzazione degli allenamenti simulati: {str(e)}")
            import traceback
            traceback.print_exc()


    def _do_import_for_schedule(self):
        """Esegue l'importazione degli allenamenti e attende il completamento."""
        self.log("Importazione degli allenamenti in corso...")
        
        # Assicurati che la cartella OAuth esista
        oauth_folder = self.oauth_folder.get()
        if not os.path.exists(oauth_folder):
            os.makedirs(oauth_folder, exist_ok=True)
            self.log(f"Creata cartella OAuth: {oauth_folder}")
        
        try:
            # Crea un oggetto args simulato
            class Args:
                pass
                    
            args = Args()
            args.oauth_folder = oauth_folder
            args.workouts_file = self.import_file.get()
            args.name_filter = None
            args.replace = self.import_replace.get()
            args.dry_run = False
            args.treadmill = False
            
            # Esegui la funzione direttamente
            cmd_import_workouts(args)
            
            self.log("Importazione completata con successo")
            
            # Aggiorna la lista degli allenamenti
            self.refresh_workouts()
            
            messagebox.showinfo("Importazione completata", 
                              "Gli allenamenti sono stati importati con successo.\n"
                              "Ora puoi procedere con la pianificazione.")
            
            return True
        except Exception as e:
            self.log(f"Errore durante l'importazione: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            messagebox.showerror("Errore", f"Errore durante l'importazione: {str(e)}")
            return False

    def _refresh_workouts_silent(self):
        """Versione silenziosa di refresh_workouts che non mostra messaggi all'utente"""
        try:
            # Crea un client Garmin direttamente
            client = GarminClient(self.oauth_folder.get())
            
            # Ottieni la lista degli allenamenti
            workouts = client.list_workouts()
            
            # Salva gli allenamenti nella cache
            self.save_workouts_to_cache(workouts)
            self.log(f"Aggiornati {len(workouts)} allenamenti nella cache")
        
        except Exception as e:
            self.log(f"Errore durante l'aggiornamento silenzioso: {str(e)}")
            
    def validate_name_prefix(self, prefix):
        """
        Verifica che il prefisso del nome termini con uno spazio.
        
        Args:
            prefix: Il name_prefix da validare
            
        Returns:
            Il prefisso corretto con uno spazio alla fine
        """
        if prefix:
            # Se c'è un prefisso, assicurati che termini con uno spazio
            if not prefix.endswith(' '):
                return prefix.rstrip() + ' '
        return prefix

        

class TextHandler(logging.Handler):
    """Handler avanzato per redirezionare i log al widget Text con formattazione e buffer limitato"""
    
    def __init__(self, text_widget, max_lines=500):
        # Inizializzare con un livello specifico (INFO)
        logging.Handler.__init__(self, level=logging.INFO)
        self.text_widget = text_widget
        self.max_lines = max_lines
        self.line_count = 0
        
        # Formattazione colorata per diversi livelli di log
        self.text_widget.tag_configure("INFO", foreground="black")
        self.text_widget.tag_configure("WARNING", foreground="orange")
        self.text_widget.tag_configure("ERROR", foreground="red")
        self.text_widget.tag_configure("DEBUG", foreground="blue")
        self.text_widget.tag_configure("CRITICAL", foreground="red", font=("", 0, "bold"))
        self.text_widget.tag_configure("TIMESTAMP", foreground="gray")
        
    def emit(self, record):
        msg = self.format(record)
        
        def append():
            self.text_widget.configure(state='normal')
            
            # Timestamp formattato
            timestamp = datetime.now().strftime("%H:%M:%S")
            
            # Aggiungi timestamp e messaggio
            self.text_widget.insert(tk.END, f"[{timestamp}] ", "TIMESTAMP")
            
            # Determina il tag appropriato in base al livello del record
            level_tag = record.levelname if record.levelname in ["INFO", "WARNING", "ERROR", "DEBUG", "CRITICAL"] else "INFO"
            
            # Aggiungi il messaggio con il tag appropriato
            self.text_widget.insert(tk.END, f"{msg}\n", level_tag)
            
            # Incrementa il conteggio delle righe
            self.line_count += 1
            
            # Se abbiamo superato il numero massimo di righe, rimuovi le più vecchie
            if self.line_count > self.max_lines:
                # Cancella le prime 100 righe quando il buffer è pieno
                self.text_widget.delete("1.0", "100.0")
                self.line_count = max(0, self.line_count - 100)
            
            self.text_widget.configure(state='disabled')
            self.text_widget.yview(tk.END)  # Auto-scroll to end
            
        # Tkinter is not thread-safe, so we need to schedule the update on the main thread
        self.text_widget.after(0, append)


if __name__ == "__main__":
    app = GarminPlannerGUI()
    app.mainloop()