#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel to YAML Converter for Garmin Planner with Scheduling Support

Questo script converte un file Excel strutturato in un file YAML compatibile con garmin-planner.
Include supporto per colonna Date con pianificazione degli allenamenti.
"""

import pandas as pd
import yaml
import re
import os
import sys
from datetime import datetime
import argparse

def excel_to_yaml(excel_file, output_file=None):
    """
    Converte un file Excel strutturato in un file YAML compatibile con garmin-planner.
    
    Args:
        excel_file: Percorso del file Excel di input
        output_file: Percorso del file YAML di output (opzionale)
    """
    # Se non viene specificato un file di output, creiamo uno con lo stesso nome ma estensione .yaml
    if output_file is None:
        output_file = os.path.splitext(excel_file)[0] + '.yaml'
    
    print(f"Convertendo {excel_file} in {output_file}...")
    
    # Carica il file Excel
    xls = pd.ExcelFile(excel_file)
    
    # Dizionario che conterrà il piano completo
    plan = {'config': {
        'heart_rates': {},
        'paces': {},
        'margins': {
            'faster': '0:03',
            'slower': '0:03',
            'hr_up': 5,
            'hr_down': 5
        },
        'name_prefix': ''
    }}
    
    # Estrai le informazioni di configurazione
    if 'Config' in xls.sheet_names:
        config_df = pd.read_excel(xls, 'Config', header=0)
        
        # Estrai il prefisso del nome (se presente)
        name_prefix_rows = config_df[config_df.iloc[:, 0] == 'name_prefix']
        if not name_prefix_rows.empty:
            plan['config']['name_prefix'] = str(name_prefix_rows.iloc[0, 1]).strip()
        
        # Estrai i margini (se presenti)
        margins_rows = config_df[config_df.iloc[:, 0] == 'margins']
        if not margins_rows.empty:
            # Controlla se ci sono valori per i margini
            if pd.notna(margins_rows.iloc[0, 1]):
                plan['config']['margins']['faster'] = str(margins_rows.iloc[0, 1]).strip()
            if pd.notna(margins_rows.iloc[0, 2]):
                plan['config']['margins']['slower'] = str(margins_rows.iloc[0, 2]).strip()
            if pd.notna(margins_rows.iloc[0, 3]):
                plan['config']['margins']['hr_up'] = int(margins_rows.iloc[0, 3])
            if pd.notna(margins_rows.iloc[0, 4]):
                plan['config']['margins']['hr_down'] = int(margins_rows.iloc[0, 4])
    
    # Estrai i ritmi
    if 'Paces' in xls.sheet_names:
        paces_df = pd.read_excel(xls, 'Paces', header=0)
        
        for _, row in paces_df.iterrows():
            # Assicurati che ci siano sia il nome che il valore
            if pd.notna(row[0]) and pd.notna(row[1]):
                name = str(row[0]).strip()
                value = str(row[1]).strip()
                plan['config']['paces'][name] = value
    
    # Estrai le frequenze cardiache
    if 'HeartRates' in xls.sheet_names:
        hr_df = pd.read_excel(xls, 'HeartRates', header=0)
        
        for _, row in hr_df.iterrows():
            # Assicurati che ci siano sia il nome che il valore
            if pd.notna(row[0]) and pd.notna(row[1]):
                name = str(row[0]).strip()
                value = str(row[1]).strip()
                plan['config']['heart_rates'][name] = value
    
    # Ora processiamo gli allenamenti in base alla struttura del file Excel
    
    # Controlla se esiste un foglio "Workouts" che contiene tutti gli allenamenti
    if 'Workouts' in xls.sheet_names:
        # Formato a foglio singolo - tutti gli allenamenti sono in un unico foglio
        workouts_df = pd.read_excel(xls, 'Workouts', header=0)
        
        # Identifica le colonne dal nome
        week_col = None
        session_col = None
        desc_col = None
        steps_col = None
        
        for i, col_name in enumerate(workouts_df.columns):
            col_name_lower = str(col_name).lower()
            if col_name_lower == 'week':
                week_col = i
            elif col_name_lower == 'session':
                session_col = i
            elif col_name_lower == 'description':
                desc_col = i
            elif col_name_lower == 'steps':
                steps_col = i
        
        # Verifica che le colonne necessarie esistano
        if None in (week_col, session_col, desc_col, steps_col):
            raise ValueError("Il foglio 'Workouts' deve contenere le colonne: 'Week', 'Session', 'Description', 'Steps'")
        
        # Identifica se c'è una colonna Date e la sua posizione
        date_col = None
        for i, col_name in enumerate(workouts_df.columns):
            if str(col_name).lower() == 'date':
                date_col = i
                break
        
        # Processa gli allenamenti
        for _, row in workouts_df.iterrows():
            # Verifica che ci siano le informazioni minime necessarie
            if pd.isna(row[week_col]) or pd.isna(row[session_col]) or pd.isna(row[desc_col]) or pd.isna(row[steps_col]):
                continue
            
            # Estrai settimana, sessione, descrizione e passi
            week = str(int(row[week_col])).zfill(2)  # Converte in int per rimuovere decimali e poi formatta
            session = str(int(row[session_col])).zfill(2)
            description = str(row[desc_col]).strip()
            
            # Crea il nome completo dell'allenamento con il formato WxxSxx
            full_name = f"W{week}S{session} {description}"
            
            # Estrai i passi dell'allenamento
            steps_str = str(row[steps_col]).strip()
            
            # Prepara la lista dei passi
            workout_steps = parse_workout_steps(steps_str, full_name)
            
            # Se c'è una data, aggiungila ai commenti dell'allenamento
            if date_col is not None and pd.notna(row[date_col]):
                date_str = ""
                if isinstance(row[date_col], datetime):
                    date_str = row[date_col].strftime("%d/%m/%Y")
                else:
                    date_str = str(row[date_col])
                
                # Modifica il nome per includere la data nel commento (sarà estratto dal parser YAML)
                full_name = f"{full_name} (Data: {date_str})"
            
            # Aggiungi l'allenamento al piano
            plan[full_name] = workout_steps
    else:
        # Formato a fogli multipli - ogni settimana ha il suo foglio
        for sheet_name in xls.sheet_names:
            # Cerca fogli che corrispondono al pattern "Week01", "Week02", ecc.
            if re.match(r'^Week\d+$', sheet_name):
                week_number = sheet_name.replace('Week', '')
                
                # Leggi il foglio della settimana
                week_df = pd.read_excel(xls, sheet_name, header=0)
                
                # Processa gli allenamenti in questa settimana
                for _, row in week_df.iterrows():
                    # Ignora le righe senza nome di allenamento
                    if pd.isna(row[0]) or not str(row[0]).strip():
                        continue
                    
                    # Estrai il nome e la descrizione dell'allenamento
                    session_name = str(row[0]).strip()
                    description = str(row[1]).strip() if pd.notna(row[1]) else ""
                    
                    # Crea il nome completo dell'allenamento con il formato WxxSxx
                    full_name = f"W{week_number}S{session_name.zfill(2)} {description}"
                    
                    # Ignora se non ci sono passi dell'allenamento
                    if pd.isna(row[2]):
                        continue
                    
                    # Estrai i passi dell'allenamento
                    steps_str = str(row[2]).strip()
                    
                    # Prepara la lista dei passi
                    workout_steps = parse_workout_steps(steps_str, full_name)
                    
                    # Aggiungi l'allenamento al piano
                    plan[full_name] = workout_steps
    
    # Salva il piano in formato YAML
    with open(output_file, 'w') as f:
        # Aggiungi i commenti per le descrizioni degli allenamenti
        yaml_content = yaml.dump(plan, default_flow_style=False, sort_keys=False)
        
        # Modifica il YAML per aggiungere i commenti per le descrizioni
        for workout_name, steps in plan.items():
            if workout_name != 'config' and ' ' in workout_name:
                # Trova la descrizione dopo lo spazio nel nome dell'allenamento
                workout_id, description = workout_name.split(' ', 1)
                
                # Sostituzione nel YAML per aggiungere il commento
                workout_line = f"{workout_name}:"
                workout_line_with_comment = f"{workout_name}: # {description}"
                yaml_content = yaml_content.replace(workout_line, workout_line_with_comment)
        
        f.write(yaml_content)
    
    print(f"Conversione completata! File YAML salvato in: {output_file}")
    return plan

def parse_workout_steps(steps_str, workout_name):
    """
    Analizza una stringa di passi e restituisce una lista strutturata.
    Supporta sia il separatore newline ('\n') che il punto e virgola (';').
    
    Args:
        steps_str: Stringa contenente i passi dell'allenamento
        workout_name: Nome dell'allenamento (per messaggi di errore)
    
    Returns:
        Lista di passi dell'allenamento
    """
    # Prepara la lista dei passi
    workout_steps = []
    
    # Sostituisci i separatori punto e virgola con newline per un'elaborazione uniforme
    steps_str = steps_str.replace(';', '\n')
    
    # Suddividi i passi in base alle righe
    step_lines = steps_str.split('\n')
    i = 0
    while i < len(step_lines):
        step_str = step_lines[i].strip()
        if not step_str:
            i += 1
            continue
        
        # Identifica il tipo di passo e i dettagli
        step_parts = step_str.split(':')
        if len(step_parts) < 2:
            print(f"ATTENZIONE: formato passo non valido in {workout_name}: {step_str}")
            i += 1
            continue
        
        step_type = step_parts[0].strip().lower()
        step_details = ':'.join(step_parts[1:]).strip()
        
        # Controlla se questo è un passo di ripetizione
        repeat_match = re.match(r'^repeat\s+(\d+)$', step_type)
        if repeat_match:
            iterations = int(repeat_match.group(1))
            
            # Estrai i passi all'interno della ripetizione
            substeps = []
            i += 1  # Passa alla riga successiva
            
            # Raccogli tutti i passi indentati dopo la ripetizione
            while i < len(step_lines):
                substep_str = step_lines[i].strip()
                if not substep_str:
                    i += 1
                    continue
                
                # Se la riga non inizia con spazi o tabulazioni, non fa parte della ripetizione
                if not step_lines[i].startswith((' ', '\t')):
                    break
                
                substep_parts = substep_str.split(':')
                if len(substep_parts) < 2:
                    print(f"ATTENZIONE: formato sotto-passo non valido in {workout_name}: {substep_str}")
                    i += 1
                    continue
                
                substep_type = substep_parts[0].strip().lower()
                substep_details = ':'.join(substep_parts[1:]).strip()
                
                substeps.append({substep_type: substep_details})
                i += 1
            
            # Aggiungi il passo di ripetizione con i suoi sotto-passi
            repeat_step = {f"repeat {iterations}": substeps}
            workout_steps.append(repeat_step)
        else:
            # Aggiungi un passo normale
            workout_steps.append({step_type: step_details})
            i += 1
    
    return workout_steps

def auto_adjust_column_widths(worksheet):
    """
    Regola automaticamente la larghezza delle colonne in base al contenuto
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
        
        adjusted_width = max(max_length + 2, 8)  # Aggiungi un po' di spazio extra
        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 60)  # Limita a 60 per evitare colonne troppo larghe


def create_sample_excel(output_file='sample_training_plan.xlsx'):
    """
    Crea un file Excel di esempio con la struttura prevista per il piano di allenamento.
    Include supporto per la colonna Date e nome atleta.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: La libreria openpyxl non è installata.")
        print("Installa openpyxl con: pip install openpyxl")
        return None
    
    wb = openpyxl.Workbook()
    
    # Foglio Config
    config_sheet = wb.active
    config_sheet.title = 'Config'
    
    config_sheet['A1'] = 'Parameter'
    config_sheet['B1'] = 'Value'
    config_sheet['C1'] = 'Slower'
    config_sheet['D1'] = 'HR Up'
    config_sheet['E1'] = 'HR Down'
    
    config_sheet['A2'] = 'name_prefix'
    config_sheet['B2'] = 'MYRUN_'
    
    config_sheet['A3'] = 'margins'
    config_sheet['B3'] = '0:03'  # faster
    config_sheet['C3'] = '0:03'  # slower
    config_sheet['D3'] = 5       # hr_up
    config_sheet['E3'] = 5       # hr_down
    
    # Formatta l'intestazione
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col in ['A', 'B', 'C', 'D', 'E']:
        config_sheet[f'{col}1'].font = Font(bold=True)
        config_sheet[f'{col}1'].fill = header_fill
    
    # Foglio Paces (Zone Z1-Z5)
    paces_sheet = wb.create_sheet(title='Paces')
    
    paces_sheet['A1'] = 'Name'
    paces_sheet['B1'] = 'Value'
    
    paces_sheet['A2'] = 'Z1'
    paces_sheet['B2'] = '6:30'
    
    paces_sheet['A3'] = 'Z2'
    paces_sheet['B3'] = '6:20'
    
    paces_sheet['A4'] = 'Z3'
    paces_sheet['B4'] = '6:00'
    
    paces_sheet['A5'] = 'Z4'
    paces_sheet['B5'] = '5:20'
    
    paces_sheet['A6'] = 'Z5'
    paces_sheet['B6'] = '4:50'
    
    # Formatta l'intestazione
    for col in ['A', 'B']:
        paces_sheet[f'{col}1'].font = Font(bold=True)
        paces_sheet[f'{col}1'].fill = header_fill
    
    # Foglio HeartRates (Zone Z1-Z5)
    hr_sheet = wb.create_sheet(title='HeartRates')
    
    hr_sheet['A1'] = 'Name'
    hr_sheet['B1'] = 'Value'
    
    # Esempio di uso di max_hr con percentuali
    hr_sheet['A2'] = 'max_hr'
    hr_sheet['B2'] = '198'
    
    hr_sheet['A3'] = 'Z1'
    hr_sheet['B3'] = '62-76% max_hr'
    
    hr_sheet['A4'] = 'Z2'
    hr_sheet['B4'] = '76-85% max_hr'
    
    hr_sheet['A5'] = 'Z3'
    hr_sheet['B5'] = '85-91% max_hr'
    
    hr_sheet['A6'] = 'Z4'
    hr_sheet['B6'] = '91-95% max_hr'
    
    hr_sheet['A7'] = 'Z5'
    hr_sheet['B7'] = '95-100% max_hr'
    
    # Formatta l'intestazione
    for col in ['A', 'B']:
        hr_sheet[f'{col}1'].font = Font(bold=True)
        hr_sheet[f'{col}1'].fill = header_fill
    
    # Foglio unico Workouts per tutti gli allenamenti
    workouts_sheet = wb.create_sheet(title='Workouts')
    
    # Aggiungi la riga per il nome dell'atleta
    # Creiamo una riga unita per il nome dell'atleta
    workouts_sheet.merge_cells('A1:E1')
    athlete_cell = workouts_sheet['A1']
    athlete_cell.value = "Atleta: "  # Predisposta per essere compilata
    athlete_cell.alignment = Alignment(horizontal='center', vertical='center')
    athlete_cell.font = Font(size=12, bold=True)
    
    # Intestazioni nella riga 2
    workouts_sheet['A2'] = 'Week'
    workouts_sheet['B2'] = 'Date'  # Aggiungiamo la colonna Date
    workouts_sheet['C2'] = 'Session'
    workouts_sheet['D2'] = 'Description'
    workouts_sheet['E2'] = 'Steps'
    
    # Formatta l'intestazione
    for col in ['A', 'B', 'C', 'D', 'E']:
        workouts_sheet[f'{col}2'].font = Font(bold=True)
        workouts_sheet[f'{col}2'].fill = header_fill
    
    # Aggiungi alcuni esempi di allenamenti
    workouts = [
        # Week, Session, Description, Steps
        (1, 1, 'Corsa facile', 'warmup: 10min @ Z1; steady: 30min @ Z2; cooldown: 5min @ Z1'),
        (1, 2, 'Intervalli brevi', 'warmup: 15min @ Z1; repeat 5:\n  interval: 400m @ Z5\n  recovery: 2min @ Z1; cooldown: 10min @ Z1'),
        (1, 3, 'Lungo lento', 'warmup: 10min @ Z1; steady: 45min @ Z2; cooldown: 5min @ Z1'),
        (2, 1, 'Corsa rigenerativa', 'steady: 30min @ Z1'),
        (2, 2, 'Threshold run', 'warmup: 15min @ Z1; steady: 20min @ Z4; cooldown: 10min @ Z1'),
        (2, 3, 'Lungo progressivo', 'warmup: 10min @ Z1; steady: 30min @ Z2; steady: 20min @ Z3; cooldown: 10min @ Z1')
    ]
    
    # Definisci colori alternati per le settimane
    week_colors = [
        "FFF2CC",  # Giallo chiaro
        "DAEEF3",  # Azzurro chiaro
        "E2EFDA",  # Verde chiaro
        "FCE4D6",  # Arancione chiaro
        "EAD1DC",  # Rosa chiaro
        "D9D9D9",  # Grigio chiaro
    ]
    
    # Bordo sottile per tutte le celle
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Aggiungi gli allenamenti al foglio
    current_week = None
    current_color_index = 0
    
    for i, (week, session, description, steps) in enumerate(workouts, start=3):  # Inizia dalla riga 3 (dopo l'intestazione e la riga dell'atleta)
        # Se cambia la settimana, cambia il colore
        if week != current_week:
            current_week = week
            current_color_index = (current_color_index + 1) % len(week_colors)
            
        # Colore di sfondo per la riga corrente
        row_fill = PatternFill(start_color=week_colors[current_color_index], 
                              end_color=week_colors[current_color_index], 
                              fill_type="solid")
        
        # Assegna valori alle celle
        workouts_sheet[f'A{i}'] = week
        workouts_sheet[f'B{i}'] = None  # Colonna Date vuota da riempire con funzione Pianifica
        workouts_sheet[f'C{i}'] = session
        workouts_sheet[f'D{i}'] = description
        workouts_sheet[f'E{i}'] = steps
        
        # Applica il colore di sfondo e il bordo a tutte le celle della riga
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = workouts_sheet[f'{col}{i}']
            cell.fill = row_fill
            cell.border = thin_border
            
            # Impostazione del testo a capo e allineamento
            cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Calcola l'altezza appropriata della riga in base al contenuto
        # Conta le righe di testo nei passi (sia \n che ;)
        num_lines = 1 + steps.count('\n') + steps.count(';')
        
        # Considera anche l'indentazione per le ripetizioni
        if 'repeat' in steps and '\n' in steps:
            # Conta le righe indentate dopo repeat
            lines_after_repeat = steps.split('repeat')[1].count('\n')
            if lines_after_repeat > 0:
                num_lines += lines_after_repeat - 1  # -1 perché la riga con 'repeat' è già contata
        
        # Altezza minima più altezza per ogni riga di testo (circa 15 punti per riga)
        row_height = max(15, 12 * num_lines)
        workouts_sheet.row_dimensions[i].height = row_height
    
    # Imposta le dimensioni delle colonne
    workouts_sheet.column_dimensions['A'].width = 10  # Week
    workouts_sheet.column_dimensions['B'].width = 15  # Date
    workouts_sheet.column_dimensions['C'].width = 10  # Session
    workouts_sheet.column_dimensions['D'].width = 25  # Description
    workouts_sheet.column_dimensions['E'].width = 60  # Steps
    
    # Regola automaticamente la larghezza delle colonne nei fogli Config, Paces e HR
    auto_adjust_column_widths(config_sheet)
    auto_adjust_column_widths(paces_sheet)
    auto_adjust_column_widths(hr_sheet)
    
    # Salva il file
    wb.save(output_file)
    print(f"File Excel di esempio creato: {output_file}")
    return output_file


def main():
    # Definisci argomenti da linea di comando
    parser = argparse.ArgumentParser(description='Converte un file Excel in un file YAML per garmin-planner')
    parser.add_argument('--excel', '-e', help='Percorso del file Excel di input', default='')
    parser.add_argument('--output', '-o', help='Percorso del file YAML di output (opzionale)')
    parser.add_argument('--create-sample', '-s', action='store_true', help='Crea un file Excel di esempio')
    parser.add_argument('--sample-name', help='Nome per il file Excel di esempio', default='sample_training_plan.xlsx')
    
    args = parser.parse_args()
    
    # Crea un file di esempio se richiesto
    if args.create_sample:
        sample_file = create_sample_excel(args.sample_name)
        if sample_file:
            # Se specificato --excel, converti subito il file di esempio
            if args.excel == '':
                args.excel = sample_file
    
    # Verifica che sia specificato un file di input
    if not args.excel:
        print("ERROR: Devi specificare un file Excel di input (--excel)")
        print("Usa --create-sample per creare un file di esempio")
        parser.print_help()
        return
    
    # Verifica che il file Excel esista
    if not os.path.exists(args.excel):
        print(f"ERROR: Il file {args.excel} non esiste")
        return
    
    # Converti il file Excel in YAML
    try:
        excel_to_yaml(args.excel, args.output)
        print("Operazione completata con successo!")
    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()