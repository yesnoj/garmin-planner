#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel to YAML Converter for Garmin Planner with Scheduling Support

This module converts a structured Excel file into a YAML file compatible with garmin-planner.
It includes support for a Date column with workout scheduling.
"""

import pandas as pd
import yaml
import re
import os
import sys
import copy
import openpyxl  # Aggiungi questa riga
from datetime import datetime
import argparse
import logging
import random
import string

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Customize YAML dumper to avoid references/aliases
class NoAliasDumper(yaml.SafeDumper):
    """Custom YAML dumper that ignores aliases"""
    def ignore_aliases(self, data):
        return True

# Valid step types supported by garmin-planner
VALID_STEP_TYPES = {"warmup", "cooldown", "interval", "recovery", "rest", "repeat", "other"}


def yaml_to_excel(yaml_data, excel_file, create_new=False):
    """
    Converti i dati YAML in un file Excel.
    
    Args:
        yaml_data: Dizionario con i dati YAML
        excel_file: Percorso del file Excel di output
        create_new: Se True, crea un nuovo file. Se False, aggiorna un file esistente.
    
    Returns:
        True se la conversione è riuscita, False altrimenti
    """
    try:
        # Importa openpyxl all'interno della funzione
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Estrai la configurazione
        config = yaml_data.get('config', {})
        sport_type = config.get('sport_type', 'running')
        
        # Se il file non esiste o vogliamo crearne uno nuovo, crealo
        if create_new or not os.path.exists(excel_file):
            # Crea un nuovo file Excel con la struttura di base
            create_sample_excel(excel_file, sport_type)
            wb = openpyxl.load_workbook(excel_file)
        else:
            # Altrimenti carica il file esistente per preservare la struttura
            wb = openpyxl.load_workbook(excel_file)
            
            # Verifica quali fogli esistono nel file originale
            original_sheets = wb.sheetnames
            
            # Se nel file originale esiste Speeds ma non Paces e stiamo passando a running
            # oppure esiste Paces ma non Speeds e stiamo passando a cycling
            # creiamo il foglio mancante
            if 'Speeds' in original_sheets and 'Paces' not in original_sheets and sport_type == 'running':
                wb.create_sheet('Paces')
            elif 'Paces' in original_sheets and 'Speeds' not in original_sheets and sport_type == 'cycling':
                wb.create_sheet('Speeds')
        
        # Aggiorna la configurazione se esiste un foglio Config
        if 'Config' in wb.sheetnames:
            update_config_sheet(wb['Config'], config)
        
        # Aggiorna i ritmi/velocità in base al tipo di sport
        if sport_type == 'running' and 'Paces' in wb.sheetnames:
            update_paces_sheet(wb['Paces'], config.get('paces', {}))
        elif sport_type == 'cycling' and 'Speeds' in wb.sheetnames:
            update_speeds_sheet(wb['Speeds'], config.get('speeds', {}))
        
        # Aggiorna le frequenze cardiache se esiste il foglio HeartRates
        if 'HeartRates' in wb.sheetnames:
            update_heart_rates_sheet(wb['HeartRates'], config.get('heart_rates', {}))
        
        # Definisci i colori per le settimane
        week_colors = [
            "FFF2CC",  # Light yellow
            "DAEEF3",  # Light blue
            "E2EFDA",  # Light green
            "FCE4D6",  # Light orange
            "EAD1DC",  # Light pink
            "D9D9D9",  # Light gray
        ]
        
        # Bordo per le celle
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aggiorna gli allenamenti nel foglio Workouts se esiste
        if 'Workouts' in wb.sheetnames:
            workouts_sheet = wb['Workouts']
            
            # Mantieni le prime due righe (intestazione e nome atleta)
            if workouts_sheet.max_row > 2:
                for row in range(workouts_sheet.max_row, 2, -1):
                    workouts_sheet.delete_rows(row)
            
            # Imposta il nome dell'atleta se presente
            if 'athlete_name' in config:
                workouts_sheet['A1'] = f"Atleta: {config['athlete_name']}"
            
            # Ottieni la lista di allenamenti (escluso 'config')
            workouts = []
            for name, steps in yaml_data.items():
                if name != 'config' and isinstance(steps, list):
                    # Estrai informazioni dall'allenamento
                    match = re.match(r'W(\d+)S(\d+)\s+(.*)', name)
                    if match:
                        week = int(match.group(1))
                        session = int(match.group(2))
                        description = match.group(3)
                        
                        # Estrai la data se presente
                        workout_date = ""
                        workout_sport_type = sport_type  # Default al tipo di sport principale
                        
                        # Filtra i passi effettivi (escludendo metadati)
                        actual_steps = []
                        for step in steps:
                            if isinstance(step, dict):
                                if 'sport_type' in step:
                                    workout_sport_type = step['sport_type']
                                elif 'date' in step:
                                    workout_date = step['date']
                                else:
                                    actual_steps.append(step)
                        
                        # Converti i passi in formato leggibile e ben formattato
                        steps_text = format_steps_for_excel(actual_steps, workout_sport_type)
                        
                        workouts.append((week, session, workout_date, description, steps_text, workout_sport_type))
            
            # Ordina gli allenamenti per settimana e sessione
            workouts.sort(key=lambda x: (x[0], x[1]))
            
            # Aggiungi gli allenamenti al foglio
            current_week = None
            current_color_index = 0
            
            row = 3  # Prima riga di dati (dopo intestazione e atleta)
            for week, session, workout_date, description, steps_text, workout_sport_type in workouts:
                # Se la settimana cambia, cambia il colore
                if week != current_week:
                    current_week = week
                    current_color_index = (week - 1) % len(week_colors)
                
                # Colore di sfondo per la riga corrente
                row_fill = PatternFill(start_color=week_colors[current_color_index],
                                     end_color=week_colors[current_color_index],
                                     fill_type="solid")
                
                # Aggiungi valori alle celle
                workouts_sheet.cell(row=row, column=1, value=week)
                workouts_sheet.cell(row=row, column=2, value=workout_date)
                workouts_sheet.cell(row=row, column=3, value=session)
                workouts_sheet.cell(row=row, column=4, value=description)
                workouts_sheet.cell(row=row, column=5, value=steps_text)
                
                # Applica colore di sfondo e bordo a tutte le celle della riga
                for col in range(1, 6):  # Colonne A-E
                    cell = workouts_sheet.cell(row=row, column=col)
                    cell.fill = row_fill
                    cell.border = thin_border
                    
                    # Imposta testo a capo e allineamento
                    cell.alignment = Alignment(wrapText=True, vertical='top')
                
                # Calcola l'altezza appropriata della riga in base al contenuto
                # Conta le linee di testo negli step (sia \n che ;)
                num_lines = 1 + steps_text.count('\n') + steps_text.count(';')
                
                # Considera l'indentazione per i repeat
                if 'repeat' in steps_text and '\n' in steps_text:
                    # Conta le linee indentate dopo repeat
                    lines_after_repeat = steps_text.split('repeat')[1].count('\n')
                    if lines_after_repeat > 0:
                        num_lines += lines_after_repeat - 1  # -1 perché la linea con 'repeat' è già contata
                
                # Altezza minima più altezza per ogni linea di testo (circa 15 punti per linea)
                row_height = max(20, 15 * num_lines)  # Altezza minima aumentata
                workouts_sheet.row_dimensions[row].height = row_height
                
                row += 1
            
            # Assicurati che le colonne abbiano la giusta larghezza
            workouts_sheet.column_dimensions['A'].width = 10  # Week
            workouts_sheet.column_dimensions['B'].width = 15  # Date
            workouts_sheet.column_dimensions['C'].width = 10  # Session
            workouts_sheet.column_dimensions['D'].width = 25  # Description
            workouts_sheet.column_dimensions['E'].width = 60  # Steps
            
        # Salva il file Excel
        wb.save(excel_file)
        return True
    
    except Exception as e:
        logging.error(f"Errore nella conversione YAML to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def format_steps_for_excel(steps):
    """
    Formatta i passi per il foglio Excel con la corretta indentazione.
    
    Args:
        steps: Lista di passi dell'allenamento
        
    Returns:
        Testo formattato dei passi
    """
    formatted_steps = []
    
    for step in steps:
        if 'repeat' in step and 'steps' in step:
            # Passo di tipo repeat
            iterations = step['repeat']
            substeps = step['steps']
            
            # Formatta il passo di repeat
            formatted_steps.append(f"repeat {iterations}:")
            
            # Formatta i substeps con indentazione
            for substep in substeps:
                if isinstance(substep, dict) and len(substep) == 1:
                    substep_type = list(substep.keys())[0]
                    substep_detail = substep[substep_type]
                    # Usa l'indentazione con due spazi per i substep
                    formatted_steps.append(f"  {substep_type}: {substep_detail}")
        
        elif isinstance(step, dict) and len(step) == 1:
            # Passo normale
            step_type = list(step.keys())[0]
            step_detail = step[step_type]
            formatted_steps.append(f"{step_type}: {step_detail}")
    
    return "\n".join(formatted_steps)



def excel_to_yaml(excel_file, output_file=None, sport_type="running"):
    """
    Converte un file Excel strutturato in un file YAML compatibile con garmin-planner.
    Include supporto per estrarre le date degli allenamenti e la data della gara.
    
    Args:
        excel_file: Percorso del file Excel di input
        output_file: Percorso del file YAML di output (opzionale)
        sport_type: Tipo di sport ('running' o 'cycling')
    """
    # Se non viene specificato un file di output, creiamo uno con lo stesso nome ma estensione .yaml
    if output_file is None:
        output_file = os.path.splitext(excel_file)[0] + '.yaml'
    
    print(f"Convertendo {excel_file} in {output_file}...")
    
    # Carica il file Excel
    try:
        # Leggi esplicitamente con le intestazioni nella seconda riga (header=1)
        df = pd.read_excel(excel_file, sheet_name='Workouts', header=1)
        
        # Verifica che ci siano le colonne richieste
        required_cols = ['Week', 'Session', 'Description', 'Steps']
        if all(col in df.columns for col in required_cols):
            print("Foglio 'Workouts' trovato con intestazioni nella seconda riga.")
        else:
            # Verifica se le colonne esistono ma con case diverso
            df_cols_lower = [col.lower() for col in df.columns]
            missing = []
            
            for req_col in required_cols:
                if req_col.lower() not in df_cols_lower:
                    missing.append(req_col)
            
            if missing:
                raise ValueError(f"Colonne mancanti nel foglio 'Workouts': {', '.join(missing)}")
            else:
                # Rinomina le colonne per uniformarle
                rename_map = {}
                for col in df.columns:
                    for req_col in required_cols:
                        if col.lower() == req_col.lower():
                            rename_map[col] = req_col
                
                df = df.rename(columns=rename_map)
                print("Colonne rinominate per uniformità.")
        
        # Ora puoi continuare con la lettura del resto del file
        xls = pd.ExcelFile(excel_file)
        
    except Exception as e:
        raise ValueError(f"Errore nel caricamento del foglio 'Workouts': {str(e)}")
    
    # Dizionario che conterrà il piano completo
    plan = {'config': {
        'heart_rates': {},
        'paces': {},
        'speeds': {},  # Aggiungi sezione speeds per il ciclismo
        'margins': {
            'faster': '0:03',
            'slower': '0:03',
            'faster_spd': '2.0',  # Margini per velocità in km/h
            'slower_spd': '2.0',
            'hr_up': 5,
            'hr_down': 5
        },
        'name_prefix': '',
    }}
    
    # Estrai il nome atleta dalla prima riga se presente
    try:
        athlete_row = pd.read_excel(excel_file, sheet_name='Workouts', header=None, nrows=1)
        athlete_text = str(athlete_row.iloc[0, 0])
        
        if athlete_text and athlete_text.strip().startswith("Atleta:"):
            athlete_name = athlete_text.replace("Atleta:", "").strip()
            if athlete_name:
                # Aggiungi il nome dell'atleta alla configurazione
                plan['config']['athlete_name'] = athlete_name
                print(f"Nome atleta estratto: {athlete_name}")
    except Exception as e:
        print(f"Nota: impossibile estrarre il nome dell'atleta: {str(e)}")
    
    # Estrai la data della gara SOLO dal foglio Config
    race_day = None
    try:
        if 'Config' in xls.sheet_names:
            config_df = pd.read_excel(excel_file, sheet_name='Config')
            race_day_rows = config_df[config_df.iloc[:, 0] == 'race_day']
            
            # Check for sport_type in Config sheet
            sport_type_rows = config_df[config_df.iloc[:, 0] == 'sport_type']
            if not sport_type_rows.empty and pd.notna(sport_type_rows.iloc[0, 1]):
                extracted_sport_type = str(sport_type_rows.iloc[0, 1]).strip().lower()
                if extracted_sport_type in ["running", "cycling"]:
                    sport_type = extracted_sport_type
                    print(f"Tipo di sport trovato nel foglio Config: {sport_type}")
            
            if not race_day_rows.empty and pd.notna(race_day_rows.iloc[0, 1]):
                race_day_value = race_day_rows.iloc[0, 1]
                
                # Gestisci diversi formati di data
                if isinstance(race_day_value, datetime):
                    race_day = race_day_value.strftime("%Y-%m-%d")
                elif isinstance(race_day_value, str):
                    try:
                        # Prova a interpretare il formato
                        if len(race_day_value) == 10 and race_day_value[4] == '-' and race_day_value[7] == '-':
                            # Già in formato YYYY-MM-DD
                            race_day = race_day_value
                        else:
                            # Prova altre interpretazioni comuni
                            try:
                                date_obj = datetime.strptime(race_day_value, "%d/%m/%Y").date()
                                race_day = date_obj.strftime("%Y-%m-%d")
                            except ValueError:
                                try:
                                    date_obj = datetime.strptime(race_day_value, "%m/%d/%Y").date()
                                    race_day = date_obj.strftime("%Y-%m-%d")
                                except ValueError:
                                    print(f"Impossibile interpretare la data: {race_day_value}")
                    except:
                        print(f"Errore nel parsing della data: {race_day_value}")
                else:
                    # Gestisci altri tipi di dato, come date numeriche
                    try:
                        race_day = pd.to_datetime(race_day_value).strftime("%Y-%m-%d")
                    except:
                        print(f"Impossibile convertire il valore in data: {race_day_value}")
                
                if race_day:
                    plan['config']['race_day'] = race_day
                    print(f"Data della gara trovata nel foglio Config: {race_day}")
            else:
                print("Campo 'race_day' non trovato nel foglio Config o valore mancante")
                # Qui potresti lanciare un'eccezione se la data della gara è obbligatoria
                raise ValueError("Data della gara (race_day) mancante nel foglio Config")
        else:
            print("Foglio Config non trovato nel file Excel")
            # Qui potresti lanciare un'eccezione se la data della gara è obbligatoria
            raise ValueError("Foglio Config mancante nel file Excel, impossibile determinare la data della gara")
    except Exception as e:
        print(f"Errore nell'estrazione della data della gara: {str(e)}")
        # Se la data della gara è considerata obbligatoria, rilancia l'eccezione
        raise ValueError(f"Impossibile determinare la data della gara: {str(e)}")
    
    # Estrai le informazioni di configurazione
    if 'Config' in xls.sheet_names:
        config_df = pd.read_excel(xls, 'Config', header=0)
        
        # Estrai il prefisso del nome (se presente)
        name_prefix_rows = config_df[config_df.iloc[:, 0] == 'name_prefix']
        if not name_prefix_rows.empty:
            # Assicurati che il prefisso termini con uno spazio
            prefix = str(name_prefix_rows.iloc[0, 1]).strip()
            # Aggiungi uno spazio alla fine se non c'è già
            if prefix and not prefix.endswith(' '):
                prefix = prefix + ' '
            plan['config']['name_prefix'] = prefix
        
        # Estrai i margini (se presenti)
        margins_rows = config_df[config_df.iloc[:, 0] == 'margins']
        if not margins_rows.empty:
            # Controlla se ci sono valori per i margini
            if pd.notna(margins_rows.iloc[0, 1]):
                if sport_type == "cycling":
                    plan['config']['margins']['faster_spd'] = str(margins_rows.iloc[0, 1]).strip()
                else:
                    plan['config']['margins']['faster'] = str(margins_rows.iloc[0, 1]).strip()
            if pd.notna(margins_rows.iloc[0, 2]):
                if sport_type == "cycling":
                    plan['config']['margins']['slower_spd'] = str(margins_rows.iloc[0, 2]).strip()
                else:
                    plan['config']['margins']['slower'] = str(margins_rows.iloc[0, 2]).strip()
            if pd.notna(margins_rows.iloc[0, 3]):
                plan['config']['margins']['hr_up'] = int(margins_rows.iloc[0, 3])
            if pd.notna(margins_rows.iloc[0, 4]):
                plan['config']['margins']['hr_down'] = int(margins_rows.iloc[0, 4])
    
    # Estrai i ritmi per corsa o velocità per ciclismo
    if sport_type == "running" and 'Paces' in xls.sheet_names:
        paces_df = pd.read_excel(xls, 'Paces', header=0)
        
        for _, row in paces_df.iterrows():
            # Assicurati che ci siano sia il nome che il valore
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                name = str(row.iloc[0]).strip()
                value = str(row.iloc[1]).strip()
                plan['config']['paces'][name] = value
    elif sport_type == "cycling" and 'Speeds' in xls.sheet_names:
        speeds_df = pd.read_excel(xls, 'Speeds', header=0)
        
        for _, row in speeds_df.iterrows():
            # Assicurati che ci siano sia il nome che il valore
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                name = str(row.iloc[0]).strip()
                value = str(row.iloc[1]).strip()
                plan['config']['speeds'][name] = value
    
    # Estrai le frequenze cardiache
    if 'HeartRates' in xls.sheet_names:
        hr_df = pd.read_excel(xls, 'HeartRates', header=0)
        
        for _, row in hr_df.iterrows():
            # Assicurati che ci siano sia il nome che il valore
            if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                name = str(row.iloc[0]).strip()
                value = row.iloc[1]
                
                # Converti i valori numerici in interi
                if isinstance(value, (int, float)) and not pd.isna(value):
                    value = int(value)
                elif isinstance(value, str) and value.strip().isdigit():
                    value = int(value.strip())
                else:
                    value = str(value).strip()
                    
                plan['config']['heart_rates'][name] = value
    
    # Dictionary to store workout descriptions for comments
    workout_descriptions = {}
    
    # Processa gli allenamenti dal DataFrame
    for _, row in df.iterrows():
        # Verifica che ci siano i dati necessari
        if pd.isna(row['Week']) or pd.isna(row['Session']) or pd.isna(row['Description']) or pd.isna(row['Steps']):
            continue
        
        # Estrai i dati
        week = str(int(row['Week'])).zfill(2)  # Formatta come 01, 02, ecc.
        session = str(int(row['Session'])).zfill(2)
        description = str(row['Description']).strip()
        
        # Crea il nome completo dell'allenamento (senza includere la data)
        full_name = f"W{week}S{session} {description}"
        
        # Memorizza la descrizione per i commenti
        workout_descriptions[full_name] = description
        
        # Estrai i passi dell'allenamento
        steps_str = str(row['Steps']).strip()
        
        # Prepara la lista dei passi
        workout_steps = parse_workout_steps(steps_str, full_name, sport_type)
        
        # Aggiungi metadati del tipo di sport come primo elemento 
        sport_type_meta = {"sport_type": sport_type}
        workout_steps.insert(0, sport_type_meta)
        
        # Aggiungi la data come secondo elemento se disponibile
        if 'Date' in df.columns and pd.notna(row['Date']):
            date_value = row['Date']
            if isinstance(date_value, str):
                formatted_date = date_value
            else:
                # Se è un oggetto datetime o date, formattalo come stringa
                formatted_date = date_value.strftime("%Y-%m-%d") if hasattr(date_value, 'strftime') else str(date_value)
            
            # Aggiungi la data come secondo elemento dei passi
            date_step = {"date": formatted_date}
            workout_steps.insert(1, date_step)
        
        # Aggiungi l'allenamento al piano (senza la data nel nome)
        plan[full_name] = workout_steps
    
    # Salva il piano in formato YAML
    with open(output_file, 'w', encoding='utf-8') as f:
        # Usa NoAliasDumper per evitare riferimenti YAML
        yaml.dump(plan, f, default_flow_style=False, sort_keys=False, Dumper=NoAliasDumper)
    
    print(f"Conversione completata! File YAML salvato in: {output_file}")
    
    # Ora aggiungi i commenti al file
    add_comments_to_yaml(output_file, workout_descriptions)
    
    return plan


def update_workouts_sheet(sheet, yaml_data):
    """
    Aggiorna il foglio Workouts con i dati dagli allenamenti YAML.
    
    Args:
        sheet: Foglio Excel Workouts
        yaml_data: Dizionario con i dati YAML
    """
    # Mantieni le prime due righe (intestazione e atleta)
    for row in range(sheet.max_row, 2, -1):
        sheet.delete_rows(row)
    
    # Ottieni la lista di allenamenti (escluso 'config')
    workouts = []
    for name, steps in yaml_data.items():
        if name != 'config' and isinstance(steps, list):
            # Estrai informazioni dall'allenamento
            match = re.match(r'W(\d+)S(\d+)\s+(.*)', name)
            if match:
                week = int(match.group(1))
                session = int(match.group(2))
                description = match.group(3)
                
                # Estrai la data se presente
                workout_date = ""
                sport_type = "running"
                
                # Filtra i passi effettivi (escludendo metadati)
                actual_steps = []
                for step in steps:
                    if isinstance(step, dict):
                        if 'sport_type' in step:
                            sport_type = step['sport_type']
                        elif 'date' in step:
                            workout_date = step['date']
                        else:
                            actual_steps.append(step)
                
                # Converti i passi in formato leggibile
                steps_text = format_steps_for_excel(actual_steps)
                
                workouts.append((week, session, workout_date, description, steps_text, sport_type))
    
    # Ordina gli allenamenti per settimana e sessione
    workouts.sort(key=lambda x: (x[0], x[1]))
    
    # Aggiungi gli allenamenti al foglio
    row = 3  # Prima riga di dati (dopo intestazione e atleta)
    for week, session, workout_date, description, steps_text, sport_type in workouts:
        sheet.cell(row=row, column=1).value = week
        sheet.cell(row=row, column=2).value = workout_date
        sheet.cell(row=row, column=3).value = session
        sheet.cell(row=row, column=4).value = description
        sheet.cell(row=row, column=5).value = steps_text
        row += 1


def format_steps_for_excel(steps, sport_type="running"):
    """
    Formatta i passi per il foglio Excel con la corretta indentazione.
    
    Args:
        steps: Lista di passi dell'allenamento
        sport_type: Tipo di sport ('running' o 'cycling')
        
    Returns:
        Testo formattato dei passi
    """
    formatted_steps = []
    
    for step in steps:
        if 'repeat' in step and 'steps' in step:
            # Passo di tipo repeat
            iterations = step['repeat']
            substeps = step['steps']
            
            # Formatta il passo di repeat
            formatted_steps.append(f"repeat {iterations}:")
            
            # Formatta i substeps con indentazione
            for substep in substeps:
                if isinstance(substep, dict) and len(substep) == 1:
                    substep_type = list(substep.keys())[0]
                    substep_detail = substep[substep_type]
                    
                    # Gestisci i diversi formati in base al tipo di sport
                    if sport_type == "cycling":
                        # Per il ciclismo, assicurati che @ diventi @spd nei passi che hanno zona di ritmo
                        if '@' in substep_detail and '@spd' not in substep_detail and '@hr' not in substep_detail:
                            substep_detail = substep_detail.replace('@', '@spd ')
                    else:  # running
                        # Per la corsa, assicurati che @spd diventi @ nei passi che hanno zona di velocità
                        if '@spd' in substep_detail:
                            substep_detail = substep_detail.replace('@spd', '@')
                    
                    # Usa l'indentazione con due spazi per i substep
                    formatted_steps.append(f"  {substep_type}: {substep_detail}")
        
        elif isinstance(step, dict) and len(step) == 1:
            # Passo normale
            step_type = list(step.keys())[0]
            step_detail = step[step_type]
            
            # Gestisci i diversi formati in base al tipo di sport
            if sport_type == "cycling":
                # Per il ciclismo, assicurati che @ diventi @spd nei passi che hanno zona di ritmo
                if '@' in step_detail and '@spd' not in step_detail and '@hr' not in step_detail:
                    step_detail = step_detail.replace('@', '@spd ')
            else:  # running
                # Per la corsa, assicurati che @spd diventi @ nei passi che hanno zona di velocità
                if '@spd' in step_detail:
                    step_detail = step_detail.replace('@spd', '@')
            
            formatted_steps.append(f"{step_type}: {step_detail}")
    
    return "\n".join(formatted_steps)

def update_heart_rates_sheet(sheet, heart_rates):
    """
    Aggiorna il foglio HeartRates con i dati dalle frequenze cardiache YAML.
    
    Args:
        sheet: Foglio Excel HeartRates
        heart_rates: Dizionario con le frequenze cardiache
    """
    # Cancella le righe esistenti (tranne l'intestazione)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Aggiungi le nuove righe
    row = 2
    for name, value in heart_rates.items():
        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=2).value = value
        row += 1


def update_speeds_sheet(sheet, speeds):
    """
    Aggiorna il foglio Speeds con i dati dalle velocità YAML.
    
    Args:
        sheet: Foglio Excel Speeds
        speeds: Dizionario con le velocità
    """
    # Cancella le righe esistenti (tranne l'intestazione)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Aggiungi le nuove righe
    row = 2
    for name, value in speeds.items():
        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=2).value = value
        row += 1

def update_paces_sheet(sheet, paces):
    """
    Aggiorna il foglio Paces con i dati dai ritmi YAML.
    
    Args:
        sheet: Foglio Excel Paces
        paces: Dizionario con i ritmi
    """
    # Cancella le righe esistenti (tranne l'intestazione)
    for row in range(sheet.max_row, 1, -1):
        sheet.delete_rows(row)
    
    # Aggiungi le nuove righe
    row = 2
    for name, value in paces.items():
        sheet.cell(row=row, column=1).value = name
        sheet.cell(row=row, column=2).value = value
        row += 1

def update_config_sheet(sheet, config):
    """
    Aggiorna il foglio Config con i dati dalla configurazione YAML.
    
    Args:
        sheet: Foglio Excel Config
        config: Dizionario con la configurazione
    """
    # Mappa delle righe da aggiornare
    config_rows = {}
    
    # Trova le righe esistenti e mappa le chiavi alle righe
    for row in range(1, sheet.max_row + 1):
        key = sheet.cell(row=row, column=1).value
        if key:
            config_rows[key] = row
    
    # Liste delle chiavi da gestire
    priority_keys = ['name_prefix', 'sport_type', 'margins', 'race_day', 'preferred_days', 'athlete_name']
    
    # Prima gestisci le chiavi prioritarie
    for key in priority_keys:
        if key in config:
            if key in config_rows:
                row_index = config_rows[key]
            else:
                row_index = sheet.max_row + 1
                sheet.cell(row=row_index, column=1).value = key
            
            if key == 'margins':
                # Aggiorna i margini
                margins = config.get('margins', {})
                if 'faster' in margins and margins['faster'] is not None:
                    sheet.cell(row=row_index, column=2).value = margins['faster']
                if 'slower' in margins and margins['slower'] is not None:
                    sheet.cell(row=row_index, column=3).value = margins['slower']
                if 'hr_up' in margins and margins['hr_up'] is not None:
                    sheet.cell(row=row_index, column=4).value = margins['hr_up']
                if 'hr_down' in margins and margins['hr_down'] is not None:
                    sheet.cell(row=row_index, column=5).value = margins['hr_down']
            elif key == 'preferred_days':
                # Gestisci preferred_days come lista o stringa
                preferred_days = config[key]
                if preferred_days is not None:
                    if isinstance(preferred_days, list):
                        sheet.cell(row=row_index, column=2).value = str(preferred_days)
                    else:
                        sheet.cell(row=row_index, column=2).value = str(preferred_days)
            else:
                # Gestisci altre chiavi normali
                value = config[key]
                if value is not None:
                    # Converti dizionari, liste o altri tipi complessi in stringhe
                    if isinstance(value, (dict, list, tuple, set)):
                        if value:  # Se non è vuoto
                            sheet.cell(row=row_index, column=2).value = str(value)
                        else:
                            sheet.cell(row=row_index, column=2).value = ""  # Dizionario vuoto -> stringa vuota
                    else:
                        sheet.cell(row=row_index, column=2).value = value
    
    # Poi gestisci altre chiavi che potrebbero essere presenti nel config
    for key, value in config.items():
        if key not in priority_keys:
            if key in config_rows:
                row_index = config_rows[key]
            else:
                row_index = sheet.max_row + 1
                sheet.cell(row=row_index, column=1).value = key
            
            # Gestisci tutti gli altri parametri, convertendo tipi complessi in stringhe
            if value is not None:
                if isinstance(value, (dict, list, tuple, set)):
                    if value:  # Se non è vuoto
                        sheet.cell(row=row_index, column=2).value = str(value)
                    else:
                        sheet.cell(row=row_index, column=2).value = ""  # Dizionario vuoto -> stringa vuota
                else:
                    sheet.cell(row=row_index, column=2).value = value


def are_required_columns_present(df, required_cols):
    """
    Check if all required columns are present in the DataFrame.
    
    Args:
        df: DataFrame to check
        required_cols: List of required column names
        
    Returns:
        True if all required columns are present, False otherwise
    """
    return all(col in df.columns for col in required_cols)

def handle_missing_columns(excel_file, required_cols):
    """
    Handle missing columns by trying different header positions or case-insensitive matching.
    
    Args:
        excel_file: Path to the Excel file
        required_cols: List of required column names
        
    Returns:
        DataFrame with corrected columns
        
    Raises:
        ValueError: If required columns cannot be found
    """
    try:
        # Try reading with header in the first row
        df = pd.read_excel(excel_file, sheet_name='Workouts', header=0)
        
        if are_required_columns_present(df, required_cols):
            logging.info("'Workouts' sheet found with headers in the first row.")
            return df
            
        # Try case-insensitive matching
        df_cols_lower = [col.lower() for col in df.columns]
        missing = []
        
        for req_col in required_cols:
            if req_col.lower() not in df_cols_lower:
                missing.append(req_col)
        
        if missing:
            raise ValueError(f"Missing columns in 'Workouts' sheet: {', '.join(missing)}")
        
        # Rename columns for consistency
        rename_map = {}
        for col in df.columns:
            for req_col in required_cols:
                if col.lower() == req_col.lower():
                    rename_map[col] = req_col
        
        df = df.rename(columns=rename_map)
        logging.info("Columns renamed for consistency.")
        return df
        
    except Exception as e:
        raise ValueError(f"Error finding required columns: {str(e)}")

def extract_config(xls, plan):
    """
    Extract configuration information from the Config sheet.
    
    Args:
        xls: ExcelFile object
        plan: Plan dictionary to update
        
    Returns:
        Updated plan dictionary
    """
    if 'Config' in xls.sheet_names:
        try:
            config_df = pd.read_excel(xls, 'Config', header=0)
            
            # Extract name prefix (if present)
            name_prefix_rows = config_df[config_df.iloc[:, 0] == 'name_prefix']
            if not name_prefix_rows.empty:
                # Ensure the prefix ends with a space
                prefix = str(name_prefix_rows.iloc[0, 1]).strip()
                # Add a space at the end if not already there
                if prefix and not prefix.endswith(' '):
                    prefix = prefix + ' '
                plan['config']['name_prefix'] = prefix
            
            # Extract margins (if present)
            margins_rows = config_df[config_df.iloc[:, 0] == 'margins']
            if not margins_rows.empty:
                # Check if there are values for the margins
                if pd.notna(margins_rows.iloc[0, 1]):
                    plan['config']['margins']['faster'] = str(margins_rows.iloc[0, 1]).strip()
                if pd.notna(margins_rows.iloc[0, 2]):
                    plan['config']['margins']['slower'] = str(margins_rows.iloc[0, 2]).strip()
                if pd.notna(margins_rows.iloc[0, 3]):
                    plan['config']['margins']['hr_up'] = int(margins_rows.iloc[0, 3])
                if pd.notna(margins_rows.iloc[0, 4]):
                    plan['config']['margins']['hr_down'] = int(margins_rows.iloc[0, 4])
        except Exception as e:
            logging.warning(f"Error extracting configuration: {str(e)}")
    
    return plan

def extract_paces(xls, plan):
    """
    Extract pace information from the Paces sheet.
    
    Args:
        xls: ExcelFile object
        plan: Plan dictionary to update
        
    Returns:
        Updated plan dictionary
    """
    if 'Paces' in xls.sheet_names:
        try:
            paces_df = pd.read_excel(xls, 'Paces', header=0)
            
            for _, row in paces_df.iterrows():
                # Ensure both name and value are present
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                    name = str(row.iloc[0]).strip()
                    value = str(row.iloc[1]).strip()
                    plan['config']['paces'][name] = value
        except Exception as e:
            logging.warning(f"Error extracting paces: {str(e)}")
    
    return plan

def extract_heart_rates(xls, plan):
    """
    Extract heart rate information from the HeartRates sheet.
    
    Args:
        xls: ExcelFile object
        plan: Plan dictionary to update
        
    Returns:
        Updated plan dictionary
    """
    if 'HeartRates' in xls.sheet_names:
        try:
            hr_df = pd.read_excel(xls, 'HeartRates', header=0)
            
            for _, row in hr_df.iterrows():
                # Ensure both name and value are present
                if pd.notna(row.iloc[0]) and pd.notna(row.iloc[1]):
                    name = str(row.iloc[0]).strip()
                    value = row.iloc[1]
                    
                    # Convert numeric values to integers
                    if isinstance(value, (int, float)) and not pd.isna(value):
                        value = int(value)
                    elif isinstance(value, str) and value.strip().isdigit():
                        value = int(value.strip())
                    else:
                        value = str(value).strip()
                        
                    plan['config']['heart_rates'][name] = value
        except Exception as e:
            logging.warning(f"Error extracting heart rates: {str(e)}")
    
    return plan

def add_comments_to_yaml(yaml_file, descriptions):
    """
    Add comments to the YAML file for workout descriptions.
    
    Args:
        yaml_file: Path to the YAML file
        descriptions: Dictionary with workout names and their descriptions
    """
    try:
        with open(yaml_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Add comments for each workout
        for workout_name, description in descriptions.items():
            # Find the line with the workout name
            pattern = f"^{re.escape(workout_name)}:"
            content = re.sub(pattern, f"{workout_name}: # {description}", content, flags=re.MULTILINE)
        
        # Write the updated content
        with open(yaml_file, 'w', encoding='utf-8') as f:
            f.write(content)
            
        logging.info("Comments added to YAML file")
    except Exception as e:
        logging.warning(f"Error adding comments to YAML: {str(e)}")

def parse_workout_steps(steps_str, workout_name, sport_type="running"):
    """
    Parse a string of steps into a structured list.
    
    Args:
        steps_str: String containing the workout steps
        workout_name: Name of the workout (for error messages)
        sport_type: Type of sport ('running' or 'cycling')
        
    Returns:
        List of structured workout steps
    """
    # Prepare the list of steps
    workout_steps = []
    
    # Replace semicolons with newlines for uniform processing
    # Ma prima correggi il formato "repeat X:" se è seguito da altre istruzioni sulla stessa riga
    steps_str = re.sub(r'(repeat\s+\d+):(.*?);', r'\1:\n\2;', steps_str)
    steps_str = re.sub(r';(\s*repeat\s+\d+:)', r';\n\1', steps_str)
    steps_str = steps_str.replace(';', '\n')
    
    # Manage steps with @spd notation for cycling
    if sport_type == "cycling":
        steps_str = steps_str.replace(' @ ', ' @spd ')
    
    # Split steps by lines
    step_lines = steps_str.split('\n')
    i = 0
    
    # Process each line
    while i < len(step_lines):
        step_str = step_lines[i].strip()
        if not step_str:
            i += 1
            continue
        
        # Check specifically for repeat pattern
        repeat_match = re.match(r'^repeat\s+(\d+):?$', step_str)
        if repeat_match:
            iterations = int(repeat_match.group(1))
            
            # Extract steps within the repeat
            substeps = []
            i += 1  # Move to the next line
            
            # Collect all indented steps after the repeat
            while i < len(step_lines):
                substep_str = step_lines[i].strip()
                if not substep_str:
                    i += 1
                    continue
                
                # If line doesn't look like a main step (doesn't contain a colon or is indented),
                # consider it part of the repeat block
                if not re.match(r'^(warmup|cooldown|interval|recovery|rest|repeat|other):', substep_str) or step_lines[i].startswith((' ', '\t')):
                    # Identify the substep type and details
                    substep_parts = substep_str.split(':')
                    if len(substep_parts) < 2:
                        logging.warning(f"Invalid substep format in {workout_name}: {substep_str}")
                        i += 1
                        continue
                    
                    substep_type = substep_parts[0].strip().lower()
                    substep_details = ':'.join(substep_parts[1:]).strip()
                    
                    # Verify substep type is valid
                    if substep_type not in VALID_STEP_TYPES:
                        if substep_type == "steady":
                            logging.warning(f"'steady' not supported in garmin-planner. Converted to 'interval' in {workout_name}")
                            substep_type = "interval"
                        else:
                            logging.warning(f"Substep type '{substep_type}' not recognized in {workout_name}, converted to 'other'")
                            substep_type = "other"
                    
                    # Handle cooldown inside repeat (likely an error)
                    if substep_type == "cooldown":
                        logging.warning(f"'cooldown' found inside a repeat in {workout_name}. Moved outside.")
                        # Save the cooldown for later
                        cooldown_details = substep_details
                        i += 1
                        # Add the repeat step with its substeps
                        # Correct format for garmin-planner: 'repeat' key and iterations value
                        repeat_step = {"repeat": iterations, "steps": copy.deepcopy(substeps)}
                        workout_steps.append(repeat_step)
                        # Add the cooldown as a separate step
                        workout_steps.append({"cooldown": cooldown_details})
                        break  # Exit the substep loop
                    
                    substeps.append({substep_type: substep_details})
                    i += 1
                else:
                    # This is a new main step, exit the repeat substeps loop
                    break
            
            # If we collected substeps, add the repeat step
            if substeps:
                # Correct format for garmin-planner: 'repeat' key and iterations value
                repeat_step = {"repeat": iterations, "steps": copy.deepcopy(substeps)}
                workout_steps.append(repeat_step)
            else:
                # If no substeps were found, add a simple repeat
                workout_steps.append({"repeat": iterations, "steps": []})
                logging.warning(f"No substeps found for repeat in {workout_name}")
        else:
            # If not a repeat, it's a normal step
            # Identify step type and details
            step_parts = step_str.split(':')
            if len(step_parts) < 2:
                logging.warning(f"Invalid step format in {workout_name}: {step_str}")
                i += 1
                continue
            
            step_type = step_parts[0].strip().lower()
            step_details = ':'.join(step_parts[1:]).strip()
            
            # Verify step type is valid for garmin-planner
            if step_type not in VALID_STEP_TYPES:
                if step_type == "steady":
                    logging.warning(f"'steady' is not supported in garmin-planner. Converted to 'interval' in {workout_name}")
                    step_type = "interval"
                else:
                    logging.warning(f"Step type '{step_type}' not recognized in {workout_name}, converted to 'other'")
                    step_type = "other"
            
            # Add a normal step
            workout_steps.append({step_type: step_details})
            i += 1
    
    return workout_steps


def auto_adjust_column_widths(worksheet):
    """
    Automatically adjust column widths based on content.
    
    Args:
        worksheet: openpyxl worksheet object
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
        
        adjusted_width = max(max_length + 2, 8)  # Add some extra space
        worksheet.column_dimensions[column_letter].width = min(adjusted_width, 60)  # Limit to 60 to avoid too wide columns

def create_sample_excel(output_file='sample_training_plan.xlsx', sport_type="running"):
    """
    Create a sample Excel file with the expected structure for the training plan.
    Includes support for a Date column.
    
    Args:
        output_file: Path for the output Excel file
        sport_type: Type of sport ('running' or 'cycling')
        
    Returns:
        Path to the created Excel file, or None if there was an error
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logging.error("ERROR: openpyxl library is not installed.")
        logging.error("Install openpyxl with: pip install openpyxl")
        return None
    
    logging.info(f"Creating sample Excel file for {sport_type}: {output_file}")
    
    wb = openpyxl.Workbook()
    
    # Define a thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    prefix = f"MY{'RUN' if sport_type == 'running' else 'BIKE'}_{random_suffix}_"
    
    # Config sheet
    config_sheet = wb.active
    config_sheet.title = 'Config'
    
    # Config sheet headers
    config_sheet['A1'] = 'Parameter'
    config_sheet['B1'] = 'Value'
    config_sheet['C1'] = 'Slower'
    config_sheet['D1'] = 'HR Up'
    config_sheet['E1'] = 'HR Down'
    
    # Config sheet values
    config_sheet['A2'] = 'name_prefix'
    config_sheet['B2'] = prefix
    
    # Aggiungi il tipo di sport alla configurazione
    config_sheet['A3'] = 'sport_type'
    config_sheet['B3'] = sport_type
    
    # Imposta i margini appropriati in base al tipo di sport
    config_sheet['A4'] = 'margins'
    if sport_type == "cycling":
        config_sheet['B4'] = '2.0'    # faster_spd in km/h
        config_sheet['C4'] = '2.0'    # slower_spd in km/h
    else:  # running
        config_sheet['B4'] = '0:03'   # faster in min:sec
        config_sheet['C4'] = '0:03'   # slower in min:sec
    config_sheet['D4'] = 5        # hr_up
    config_sheet['E4'] = 5        # hr_down
    
    # Format header
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for col in ['A', 'B', 'C', 'D', 'E']:
        config_sheet[f'{col}1'].font = Font(bold=True)
        config_sheet[f'{col}1'].fill = header_fill
    
    # Crea il foglio appropriato in base al tipo di sport
    if sport_type == "running":
        # Paces sheet per running
        zones_sheet = wb.create_sheet(title='Paces')
        zones_sheet['A1'] = 'Name'
        zones_sheet['B1'] = 'Value'
        
        zones_sheet['A2'] = 'Z1'
        zones_sheet['B2'] = '6:30'
        
        zones_sheet['A3'] = 'Z2'
        zones_sheet['B3'] = '6:00'
        
        zones_sheet['A4'] = 'Z3'
        zones_sheet['B4'] = '5:30'
        
        zones_sheet['A5'] = 'Z4'
        zones_sheet['B5'] = '5:00'
        
        zones_sheet['A6'] = 'Z5'
        zones_sheet['B6'] = '4:30'
    else:  # cycling
        # Speeds sheet per cycling
        zones_sheet = wb.create_sheet(title='Speeds')
        zones_sheet['A1'] = 'Name'
        zones_sheet['B1'] = 'Value'
        
        zones_sheet['A2'] = 'Z1'
        zones_sheet['B2'] = '15.0'
        
        zones_sheet['A3'] = 'Z2'
        zones_sheet['B3'] = '20.0'
        
        zones_sheet['A4'] = 'Z3'
        zones_sheet['B4'] = '25.0'
        
        zones_sheet['A5'] = 'Z4'
        zones_sheet['B5'] = '30.0'
        
        zones_sheet['A6'] = 'Z5'
        zones_sheet['B6'] = '35.0'
    
    # Format header
    for col in ['A', 'B']:
        zones_sheet[f'{col}1'].font = Font(bold=True)
        zones_sheet[f'{col}1'].fill = header_fill
    
    # HeartRates sheet (Z1-Z5 zones)
    hr_sheet = wb.create_sheet(title='HeartRates')
    
    hr_sheet['A1'] = 'Name'
    hr_sheet['B1'] = 'Value'
    
    # Example of using max_hr with percentages
    hr_sheet['A2'] = 'max_hr'
    hr_sheet['B2'] = 180  # Use an integer instead of a string
    
    hr_sheet['A3'] = 'Z1_HR'
    hr_sheet['B3'] = '62-76% max_hr'
    
    hr_sheet['A4'] = 'Z2_HR'
    hr_sheet['B4'] = '76-85% max_hr'
    
    hr_sheet['A5'] = 'Z3_HR'
    hr_sheet['B5'] = '85-91% max_hr'
    
    hr_sheet['A6'] = 'Z4_HR'
    hr_sheet['B6'] = '91-95% max_hr'
    
    hr_sheet['A7'] = 'Z5_HR'
    hr_sheet['B7'] = '95-100% max_hr'
    
    # Format header
    for col in ['A', 'B']:
        hr_sheet[f'{col}1'].font = Font(bold=True)
        hr_sheet[f'{col}1'].fill = header_fill
    
    # Single Workouts sheet for all workouts
    workouts_sheet = wb.create_sheet(title='Workouts')
    
    # Add a row for the athlete's name
    # Create a merged cell for the athlete's name
    workouts_sheet.merge_cells('A1:E1')
    athlete_cell = workouts_sheet['A1']
    athlete_cell.value = "Athlete: "  # Prepared to be filled in
    athlete_cell.alignment = Alignment(horizontal='center', vertical='center')
    athlete_cell.font = Font(size=12, bold=True)
    # Add border to the athlete cell
    athlete_cell.border = thin_border

    # Headers in row 2
    workouts_sheet['A2'] = 'Week'
    workouts_sheet['B2'] = 'Date'
    workouts_sheet['C2'] = 'Session'
    workouts_sheet['D2'] = 'Description'
    workouts_sheet['E2'] = 'Steps'

    # Format header
    for col in ['A', 'B', 'C', 'D', 'E']:
        cell = workouts_sheet[f'{col}2']
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = thin_border  # Add border to all header cells
    
    # Scegli gli esempi di workout in base al tipo di sport
    if sport_type == "running":
        workouts = [
            # Week, Session, Description, Steps
            (1, 1, 'Easy run', 'warmup: 10min @ Z1_HR\ninterval: 30min @ Z2\ncooldown: 5min @ Z1_HR'),
            (1, 2, 'Short intervals', 'warmup: 15min @ Z1_HR\nrepeat 5:\n  interval: 400m @ Z5\n  recovery: 2min @ Z1_HR\ncooldown: 10min @ Z1_HR'),
            (1, 3, 'Long slow run', 'warmup: 10min @ Z1_HR\ninterval: 45min @ Z2\ncooldown: 5min @ Z1_HR'),
            (2, 1, 'Recovery run', 'interval: 30min @ Z1_HR'),
            (2, 2, 'Threshold run', 'warmup: 15min @ Z1_HR\ninterval: 20min @ Z4\ncooldown: 10min @ Z1_HR'),
            (2, 3, 'Progressive long run', 'warmup: 10min @ Z1_HR\ninterval: 30min @ Z2\ninterval: 20min @ Z3\ncooldown: 10min @ Z1_HR')
        ]
    else:  # cycling
        workouts = [
            # Week, Session, Description, Steps
            (1, 1, 'Easy ride', 'warmup: 10min @hr Z1_HR\ninterval: 30min @spd Z2\ncooldown: 5min @hr Z1_HR'),
            (1, 2, 'Short intervals', 'warmup: 15min @hr Z1_HR\nrepeat 5:\n  interval: 2min @spd Z5\n  recovery: 3min @hr Z1_HR\ncooldown: 10min @hr Z1_HR'),
            (1, 3, 'Long endurance ride', 'warmup: 10min @hr Z1_HR\ninterval: 60min @spd Z2\ncooldown: 5min @hr Z1_HR'),
            (2, 1, 'Recovery ride', 'interval: 30min @hr Z1_HR'),
            (2, 2, 'Threshold ride', 'warmup: 15min @hr Z1_HR\ninterval: 20min @spd Z4\ncooldown: 10min @hr Z1_HR'),
            (2, 3, 'Progressive ride', 'warmup: 10min @hr Z1_HR\ninterval: 30min @spd Z2\ninterval: 20min @spd Z3\ncooldown: 10min @hr Z1_HR')
        ]
    
    # Define alternating colors for weeks
    week_colors = [
        "FFF2CC",  # Light yellow
        "DAEEF3",  # Light blue
        "E2EFDA",  # Light green
        "FCE4D6",  # Light orange
        "EAD1DC",  # Light pink
        "D9D9D9",  # Light gray
    ]
    
    # Add workouts to the sheet
    current_week = None
    current_color_index = 0
    
    for i, (week, session, description, steps) in enumerate(workouts, start=3):  # Start from row 3 (after header and athlete row)
        # If the week changes, change the color
        if week != current_week:
            current_week = week
            current_color_index = (current_color_index + 1) % len(week_colors)
            
        # Background color for the current row
        row_fill = PatternFill(start_color=week_colors[current_color_index], 
                              end_color=week_colors[current_color_index], 
                              fill_type="solid")
        
        # Assign values to cells
        workouts_sheet[f'A{i}'] = week
        workouts_sheet[f'B{i}'] = None  # Empty Date column to be filled by Plan function
        workouts_sheet[f'C{i}'] = session
        workouts_sheet[f'D{i}'] = description
        workouts_sheet[f'E{i}'] = steps
        
        # Apply background color and border to all cells in the row
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = workouts_sheet[f'{col}{i}']
            cell.fill = row_fill
            cell.border = thin_border
            
            # Set text wrapping and alignment
            cell.alignment = Alignment(wrapText=True, vertical='top')
        
        # Calculate appropriate row height based on content
        # Count lines of text in steps (both \n and ;)
        num_lines = 1 + steps.count('\n') + steps.count(';')
        
        # Consider indentation for repeats
        if 'repeat' in steps and '\n' in steps:
            # Count indented lines after repeat
            lines_after_repeat = steps.split('repeat')[1].count('\n')
            if lines_after_repeat > 0:
                num_lines += lines_after_repeat - 1  # -1 because the line with 'repeat' is already counted
        
        # Minimum height plus height for each line of text (about 15 points per line)
        row_height = max(20, 15 * num_lines)  # Increased minimum height
        workouts_sheet.row_dimensions[i].height = row_height
    
    # Set column widths
    workouts_sheet.column_dimensions['A'].width = 10  # Week
    workouts_sheet.column_dimensions['B'].width = 15  # Date
    workouts_sheet.column_dimensions['C'].width = 10  # Session
    workouts_sheet.column_dimensions['D'].width = 25  # Description
    workouts_sheet.column_dimensions['E'].width = 60  # Steps
    
    # Automatically adjust column widths in Config and HR sheets
    auto_adjust_column_widths(config_sheet)
    auto_adjust_column_widths(zones_sheet)
    auto_adjust_column_widths(hr_sheet)
    
    # Save the file
    wb.save(output_file)
    logging.info(f"Sample Excel file created for {sport_type}: {output_file}")
    return output_file


def main():
    """Main function for command line use"""
    # Define command line arguments
    parser = argparse.ArgumentParser(description='Convert an Excel file to a YAML file for garmin-planner')
    parser.add_argument('--excel', '-e', help='Path to the input Excel file', default='')
    parser.add_argument('--output', '-o', help='Path to the output YAML file (optional)')
    parser.add_argument('--create-sample', '-s', action='store_true', help='Create a sample Excel file')
    parser.add_argument('--sample-name', help='Name for the sample Excel file', default='sample_training_plan.xlsx')
    parser.add_argument('--sport-type', help='Type of sport (running or cycling)', choices=['running', 'cycling'], default='running')
    
    args = parser.parse_args()
    
    # Create a sample file if requested
    if args.create_sample:
        sample_file = create_sample_excel(args.sample_name, args.sport_type)
        if sample_file:
            # If specified --excel, immediately convert the sample file
            if args.excel == '':
                args.excel = sample_file
    
    # Verify that an input file is specified
    if not args.excel:
        logging.error("ERROR: You must specify an input Excel file (--excel)")
        logging.info("Use --create-sample to create a sample file")
        parser.print_help()
        return
    
    # Verify that the Excel file exists
    if not os.path.exists(args.excel):
        logging.error(f"ERROR: File {args.excel} does not exist")
        return
    
    # Convert the Excel file to YAML
    try:
        excel_to_yaml(args.excel, args.output, args.sport_type)
        logging.info("Operation completed successfully!")
    except Exception as e:
        logging.error(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()