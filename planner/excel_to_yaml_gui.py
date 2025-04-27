#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel to YAML Converter GUI for Garmin Planner with Scheduling

Interfaccia grafica per convertire file Excel in YAML compatibili con garmin-planner.
Include la funzionalità di pianificazione automatica degli allenamenti.
"""

import sys
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
from datetime import timedelta
import calendar
import pandas as pd
from tkcalendar import DateEntry  # Aggiungiamo il componente calendario
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta

# Verifica che le librerie necessarie siano installate
try:
    import pandas as pd
    import yaml
    import openpyxl
    from tkcalendar import DateEntry
except ImportError as e:
    # Crea una finestra di errore minima
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Errore di importazione", 
                       f"Manca una libreria necessaria: {e}\n\n"
                       f"Installa le librerie richieste con:\n"
                       f"pip install pandas pyyaml openpyxl tkcalendar")
    sys.exit(1)

# Importa il convertitore
from planner.excel_to_yaml_converter import excel_to_yaml, create_sample_excel

class ScheduleDialog(tk.Toplevel):
    """Finestra di dialogo per pianificare gli allenamenti"""
    def __init__(self, parent, excel_file):
        super().__init__(parent)
        
        self.parent = parent
        self.excel_file = excel_file
        self.result = None
        self.sport_type = "running"  # Default
        
        self.title("Pianifica Allenamenti")
        self.geometry("570x630")  # Un po' più alta per il selettore di sport
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()
        
        # Centra la finestra
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (parent.winfo_screenwidth() // 2) - (width // 2)
        y = (parent.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
        
        # Estrai il tipo di sport dal file Excel
        self.extract_sport_type()
        
        # Crea l'interfaccia della finestra di dialogo
        self.create_widgets()
        
        # Carica i dati del file Excel se esiste
        if os.path.exists(excel_file):
            try:
                self.load_excel_info()
            except Exception as e:
                messagebox.showerror("Errore", f"Non è stato possibile caricare il file Excel: {str(e)}")
        
        # Attende che la finestra venga chiusa
        self.wait_window()
    

    def extract_sport_type(self):
        """Estrai il tipo di sport dal file Excel"""
        try:
            xls = pd.ExcelFile(self.excel_file)
            if 'Config' in xls.sheet_names:
                config_df = pd.read_excel(self.excel_file, sheet_name='Config')
                sport_type_rows = config_df[config_df.iloc[:, 0] == 'sport_type']
                
                if not sport_type_rows.empty and pd.notna(sport_type_rows.iloc[0, 1]):
                    extracted_sport_type = str(sport_type_rows.iloc[0, 1]).strip().lower()
                    if extracted_sport_type in ["running", "cycling"]:
                        self.sport_type = extracted_sport_type
                        print(f"Sport type extracted from Excel: {self.sport_type}")
                        return
            
            # Se arriviamo qui, non abbiamo trovato un tipo di sport esplicito
            # Proviamo a dedurlo dalla presenza di fogli Paces o Speeds
            if 'Speeds' in xls.sheet_names and 'Paces' not in xls.sheet_names:
                self.sport_type = "cycling"
                print("Sport type inferred as 'cycling' from Speeds sheet presence")
            elif 'Paces' in xls.sheet_names and 'Speeds' not in xls.sheet_names:
                self.sport_type = "running"
                print("Sport type inferred as 'running' from Paces sheet presence")
            
        except Exception as e:
            print(f"Error extracting sport type: {str(e)}")
            # In caso di errore, manteniamo il valore predefinito

    def create_custom_date_picker(self, parent, date_var):
        """Create a custom date picker with separate widgets for year, month, and day"""
        frame = ttk.Frame(parent)
        
        # Get current date from variable or use today
        try:
            current_date = datetime.strptime(date_var.get(), "%Y-%m-%d")
        except (ValueError, TypeError):
            current_date = datetime.now()
        
        # Year picker
        year_values = list(range(current_date.year, current_date.year + 10))
        year_var = tk.StringVar(value=str(current_date.year))
        ttk.Label(frame, text="Anno:").grid(row=0, column=0, padx=(0, 5))
        year_combo = ttk.Combobox(frame, textvariable=year_var, values=year_values, width=6, state="readonly")
        year_combo.grid(row=0, column=1, padx=(0, 10))
        
        # Month picker
        month_names = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", 
                      "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
        month_var = tk.StringVar(value=month_names[current_date.month-1])
        ttk.Label(frame, text="Mese:").grid(row=0, column=2, padx=(0, 5))
        month_combo = ttk.Combobox(frame, textvariable=month_var, values=month_names, width=10, state="readonly")
        month_combo.grid(row=0, column=3, padx=(0, 10))
        
        # Day picker - days will be adjusted based on month/year
        day_var = tk.StringVar(value=str(current_date.day))
        ttk.Label(frame, text="Giorno:").grid(row=0, column=4, padx=(0, 5))
        day_combo = ttk.Combobox(frame, textvariable=day_var, width=5, state="readonly")
        day_combo.grid(row=0, column=5)
        
        def update_days(*args):
            """Update the available days based on the selected month and year"""
            try:
                year = int(year_var.get())
                month = month_names.index(month_var.get()) + 1
                
                # Get the number of days in the month
                _, num_days = calendar.monthrange(year, month)
                
                # Update the day values
                day_values = [str(i) for i in range(1, num_days + 1)]
                day_combo['values'] = day_values
                
                # Adjust the day if it's out of range
                current_day = int(day_var.get()) if day_var.get() else 1
                if current_day > num_days:
                    day_var.set(str(num_days))
                elif not day_var.get():
                    day_var.set("1")
                    
            except (ValueError, TypeError):
                # Default to 31 days if there's an error
                day_combo['values'] = [str(i) for i in range(1, 32)]
        
        # Initial update of days
        update_days()
        
        # Function to update the main date variable
        def update_date(*args):
            try:
                year = int(year_var.get())
                month = month_names.index(month_var.get()) + 1
                day = int(day_var.get())
                
                # Validate the date
                date_obj = datetime(year, month, day)
                date_var.set(date_obj.strftime("%Y-%m-%d"))
            except (ValueError, TypeError):
                # Invalid date, don't update
                pass
        
        # Bind the update function to the variables
        year_var.trace_add("write", update_date)
        month_var.trace_add("write", lambda *args: [update_days(), update_date()])
        day_var.trace_add("write", update_date)
        
        return frame

    
    def create_widgets(self):
        """Crea i widget dell'interfaccia"""
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Titolo
        ttk.Label(main_frame, text="Pianifica Allenamenti", font=("", 14, "bold")).pack(pady=10)
        
        # Descrizione
        if self.sport_type == "cycling":
            activity_type = "ciclismo"
        else:
            activity_type = "corsa"
            
        ttk.Label(main_frame, text=f"Imposta la data di inizio e seleziona i giorni della settimana per gli allenamenti di {activity_type}.",
                 wraplength=450).pack(pady=5)
        
        # Frame per il tipo di sport
        sport_frame = ttk.LabelFrame(main_frame, text="Tipo di Sport")
        sport_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.sport_type_var = tk.StringVar(value=self.sport_type)
        ttk.Label(sport_frame, text="Tipo di allenamento:").pack(side=tk.LEFT, padx=5, pady=5)
        
        # Rendere il tipo di sport read-only per evitare confusione
        sport_display = ttk.Label(sport_frame, text=self.sport_type, font=("", 10, "bold"))
        sport_display.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Aggiunta di una nota informativa
        ttk.Label(sport_frame, text="(determinato dal file Excel)", foreground="gray").pack(side=tk.LEFT, padx=5, pady=5)
        
        # Frame per nome atleta
        athlete_frame = ttk.LabelFrame(main_frame, text="Atleta")
        athlete_frame.pack(fill=tk.X, padx=5, pady=10)
        
        ttk.Label(athlete_frame, text="Nome dell'atleta:").pack(side=tk.LEFT, padx=5, pady=5)
        self.athlete_name_var = tk.StringVar()
        ttk.Entry(athlete_frame, textvariable=self.athlete_name_var, width=30).pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        date_frame = ttk.LabelFrame(main_frame, text="Data di Inizio")
        date_frame.pack(fill=tk.X, padx=5, pady=10)
        
        ttk.Label(date_frame, text="Seleziona la data di inizio del piano:").pack(pady=5)
        
        # Data di default: oggi + 7 giorni
        default_start = datetime.now().date() + timedelta(days=7)
        # Variabile per memorizzare la data
        self.start_date_var = tk.StringVar(value=default_start.strftime("%Y-%m-%d"))
        
        # Usa il selettore di data personalizzato invece di DateEntry
        start_date_picker = self.create_custom_date_picker(date_frame, self.start_date_var)
        start_date_picker.pack(pady=5)
        
        # Frame per i giorni della settimana
        days_frame = ttk.LabelFrame(main_frame, text="Giorni della Settimana")
        days_frame.pack(fill=tk.X, padx=5, pady=10)
        
        # Crea checkbox per ogni giorno della settimana
        self.days_vars = []
        days = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
        
        days_grid = ttk.Frame(days_frame)
        days_grid.pack(padx=5, pady=5)
        
        # Prima riga: checkbox
        for i, day in enumerate(days):
            var = tk.IntVar(value=0)  # Deselezionato di default
            self.days_vars.append(var)
            ttk.Checkbutton(days_grid, text=day, variable=var).grid(row=0, column=i, padx=5, pady=5)
        
        # Aggiungi bottoni per selezionare gruppi comuni
        buttons_frame = ttk.Frame(days_frame)
        buttons_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Preseleziona i giorni in base al tipo di sport
        if self.sport_type == "cycling":
            # Pattern comuni di allenamento per ciclismo
            ttk.Button(buttons_frame, text="Weekend", 
                     command=lambda: self.select_days([5, 6])).pack(side=tk.LEFT, padx=5)
            ttk.Button(buttons_frame, text="Mar-Gio-Sab", 
                     command=lambda: self.select_days([1, 3, 5])).pack(side=tk.LEFT, padx=5)
            ttk.Button(buttons_frame, text="Lun-Mer-Ven-Dom", 
                     command=lambda: self.select_days([0, 2, 4, 6])).pack(side=tk.LEFT, padx=5)
        else:  # running
            # Pattern comuni di allenamento per corsa
            ttk.Button(buttons_frame, text="Mar-Gio-Sab", 
                     command=lambda: self.select_days([1, 3, 5])).pack(side=tk.LEFT, padx=5)
            ttk.Button(buttons_frame, text="Lun-Mer-Ven", 
                     command=lambda: self.select_days([0, 2, 4])).pack(side=tk.LEFT, padx=5)
            ttk.Button(buttons_frame, text="Mar-Mer-Ven-Dom", 
                     command=lambda: self.select_days([1, 2, 4, 6])).pack(side=tk.LEFT, padx=5)
    
        # Frame per le sessioni per settimana
        sessions_frame = ttk.LabelFrame(main_frame, text="Sessioni per Settimana")
        sessions_frame.pack(fill=tk.X, padx=5, pady=10)
        
        ttk.Label(sessions_frame, text="Quante sessioni di allenamento vuoi fare ogni settimana?").pack(pady=5)
        
        sessions_values = [str(i) for i in range(1, 8)]
        self.sessions_var = tk.StringVar(value="3")  # Default: 3 sessioni a settimana
        sessions_combo = ttk.Combobox(sessions_frame, textvariable=self.sessions_var, values=sessions_values, width=5, state="readonly")
        sessions_combo.pack(pady=5)
        
        # Informazioni sugli allenamenti attuali
        self.workouts_info = ttk.Label(main_frame, text="", wraplength=450)
        self.workouts_info.pack(pady=10)
        
        # Pulsanti OK e Annulla
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill=tk.X, padx=5, pady=10)
        
        ttk.Button(buttons_frame, text="Annulla", command=self.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(buttons_frame, text="Pianifica Date", command=self.schedule_dates).pack(side=tk.RIGHT, padx=5)

    
    def select_days(self, indices):
        """Seleziona i giorni specificati"""
        # Prima azzera tutti
        for var in self.days_vars:
            var.set(0)
        
        # Poi seleziona quelli specificati
        for i in indices:
            self.days_vars[i].set(1)
    
    def load_excel_info(self):
        """Carica le informazioni dal file Excel"""
        try:
            # Verifica se il file esiste
            if not os.path.exists(self.excel_file):
                return
                
            # Prova prima a caricare assumendo che le intestazioni siano nella riga 1
            try:
                df = pd.read_excel(self.excel_file, sheet_name='Workouts')
                if 'Week' in df.columns:
                    # Le intestazioni sono correttamente nella riga 1
                    self._process_workout_data(df)
                    return
            except Exception as e:
                print(f"Tentativo 1 fallito: {str(e)}")
                
            # Se il primo tentativo fallisce, prova assumendo che le intestazioni siano nella riga 2
            try:
                df = pd.read_excel(self.excel_file, sheet_name='Workouts', header=1)  # header=1 significa seconda riga
                if 'Week' in df.columns:
                    # Le intestazioni sono nella riga 2
                    self._process_workout_data(df)
                    return
            except Exception as e:
                print(f"Tentativo 2 fallito: {str(e)}")
                
            # Se siamo qui, nessuno dei tentativi ha funzionato
            raise Exception("Non è stato possibile trovare le intestazioni corrette nel file Excel")
                
        except Exception as e:
            raise Exception(f"Errore nel caricamento del file Excel: {str(e)}")
            
    def _process_workout_data(self, df):
        """Elabora i dati degli allenamenti"""
        if not df.empty:
            # Converte 'Week' in numerico se necessario
            if 'Week' in df.columns:
                df['Week'] = pd.to_numeric(df['Week'], errors='coerce')
                
            unique_weeks = df['Week'].nunique()
            max_sessions = df.groupby('Week').size().max()
            
            # Aggiorna il messaggio informativo con il tipo di sport
            sport_display = "ciclismo" if self.sport_type == "cycling" else "corsa"
            self.workouts_info.config(text=f"Il piano di {sport_display} contiene {unique_weeks} settimane con un massimo di {max_sessions} sessioni per settimana.")
            
            # Imposta il numero di sessioni
            self.sessions_var.set(str(max_sessions))
            
            # Preseleziona i giorni in base al numero di sessioni e al tipo di sport
            if self.sport_type == "cycling":
                # Pattern per ciclismo
                if max_sessions == 1:
                    self.select_days([5])  # Sabato
                elif max_sessions == 2:
                    self.select_days([3, 5])  # Giovedì, Sabato
                elif max_sessions == 3:
                    self.select_days([1, 3, 5])  # Martedì, Giovedì, Sabato
                elif max_sessions == 4:
                    self.select_days([0, 2, 4, 6])  # Lun, Mer, Ven, Dom
                elif max_sessions >= 5:
                    self.select_days([0, 1, 3, 4, 6])  # L, M, G, V, D
            else:
                # Pattern per corsa
                if max_sessions == 1:
                    self.select_days([2])  # Mercoledì
                elif max_sessions == 2:
                    self.select_days([1, 4])  # Martedì, Venerdì
                elif max_sessions == 3:
                    self.select_days([1, 3, 5])  # Martedì, Giovedì, Sabato
                elif max_sessions == 4:
                    self.select_days([1, 3, 5, 6])  # Martedì, Giovedì, Sabato, Domenica
                elif max_sessions >= 5:
                    self.select_days([0, 1, 3, 4, 6])  # L, M, G, V, D



    def schedule_dates(self):
        """Assegna le date in base ai giorni selezionati"""
        try:
            selected_days = [i for i, var in enumerate(self.days_vars) if var.get() == 1]
            if not selected_days:
                messagebox.showerror("Errore", "Seleziona almeno un giorno della settimana.")
                return
                
            # Ottieni la data di inizio dal selettore personalizzato
            start_date_str = self.start_date_var.get()
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            except ValueError:
                messagebox.showerror("Errore", f"Data non valida: {start_date_str}")
                return
                
            athlete_name = self.athlete_name_var.get().strip()
            
            # Carica il file Excel
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb['Workouts']
            
            # Layout fisso:
            # Riga 1: Atleta
            # Riga 2: Intestazioni (Week, Session, ecc.)
            
            # Aggiorna il nome dell'atleta
            if athlete_name:
                athlete_cell = ws.cell(row=1, column=1)
                athlete_cell.value = f"Atleta: {athlete_name}"
            
            # Verifichiamo che la seconda riga contenga "Week" nella prima colonna
            if ws.cell(row=2, column=1).value != "Week":
                messagebox.showerror("Errore", "Il file Excel non ha la struttura attesa. La seconda riga dovrebbe contenere 'Week' nella prima colonna.")
                return
            
            # Ottieni gli indici delle colonne dalla seconda riga
            col_indices = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=2, column=col).value
                if header:
                    col_indices[header] = col
            
            # Verifica che ci siano le colonne necessarie
            for col_name in ["Week", "Session", "Description", "Steps"]:
                if col_name not in col_indices:
                    messagebox.showerror("Errore", f"Colonna '{col_name}' non trovata nel file Excel.")
                    return
            
            # Aggiungi la colonna Date se non esiste
            if "Date" not in col_indices:
                # Inserisci dopo Week
                week_col = col_indices["Week"]
                ws.insert_cols(week_col + 1)
                
                # Aggiorna gli indici delle colonne
                for key in list(col_indices.keys()):
                    if col_indices[key] > week_col:
                        col_indices[key] += 1
                
                # Aggiungi l'intestazione
                date_col = week_col + 1
                ws.cell(row=2, column=date_col, value="Date")
                
                # Formatta l'intestazione
                header_cell = ws.cell(row=2, column=date_col)
                header_cell.font = openpyxl.styles.Font(bold=True)
                header_cell.fill = openpyxl.styles.PatternFill(
                    fill_type="solid",
                    start_color="DDEBF7",
                    end_color="DDEBF7"
                )
                header_cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    bottom=openpyxl.styles.Side(style="thin")
                )
                
                # Aggiungi alla mappa
                col_indices["Date"] = date_col
            
            # Raccogli i dati degli allenamenti
            workouts = []
            week_col = col_indices["Week"]
            session_col = col_indices["Session"]
            
            for row in range(3, ws.max_row + 1):  # Inizia dalla terza riga (dopo intestazioni)
                week = ws.cell(row=row, column=week_col).value
                session = ws.cell(row=row, column=session_col).value
                
                if week is not None and session is not None:
                    # Converti in numeri se necessario
                    if isinstance(week, str) and week.strip().isdigit():
                        week = int(week.strip())
                    if isinstance(session, str) and session.strip().isdigit():
                        session = int(session.strip())
                    
                    # Ottieni il colore di sfondo
                    cell = ws.cell(row=row, column=week_col)
                    color = cell.fill.start_color.index if cell.fill and cell.fill.start_color else None
                    
                    workouts.append((week, session, row, color))
            
            # Se non ci sono allenamenti, mostra un errore
            if not workouts:
                messagebox.showerror("Errore", "Non sono stati trovati allenamenti validi nel foglio.")
                return
            
            # Ordina gli allenamenti
            workouts.sort(key=lambda x: (x[0], x[1]))
            
            # Assegna le date
            date_col = col_indices["Date"]
            current_date = start_date
            weekday = current_date.weekday()
            
            # Vai al primo giorno selezionato
            while weekday not in selected_days:
                current_date += timedelta(days=1)
                weekday = current_date.weekday()
            
            current_week = workouts[0][0]
            
            # Per ogni allenamento...
            for week, session, row, color in workouts:
                # Se cambia la settimana, vai alla settimana successiva
                if week > current_week:
                    current_week = week
                    # Vai a lunedì della settimana successiva
                    days_to_monday = 7 - current_date.weekday()
                    current_date += timedelta(days=days_to_monday)
                    weekday = 0  # lunedì
                    
                    # Vai al primo giorno selezionato della nuova settimana
                    while weekday not in selected_days:
                        current_date += timedelta(days=1)
                        weekday = current_date.weekday()
                
                # Assegna la data
                date_cell = ws.cell(row=row, column=date_col)
                date_cell.value = current_date
                date_cell.number_format = "YYYY-MM-DD"
                
                # Mantieni lo stile
                if color:
                    date_cell.fill = openpyxl.styles.PatternFill(
                        fill_type="solid",
                        start_color=color,
                        end_color=color
                    )
                
                date_cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style="thin"),
                    right=openpyxl.styles.Side(style="thin"),
                    top=openpyxl.styles.Side(style="thin"),
                    bottom=openpyxl.styles.Side(style="thin")
                )
                
                date_cell.alignment = openpyxl.styles.Alignment(
                    horizontal="center",
                    vertical="center",
                    wrapText=True
                )
                
                # Trova il prossimo giorno selezionato
                current_idx = selected_days.index(weekday)
                next_idx = (current_idx + 1) % len(selected_days)
                next_day = selected_days[next_idx]
                
                # Calcola quanti giorni aggiungere
                days_to_add = (next_day - weekday) % 7
                if days_to_add == 0:
                    days_to_add = 7  # Stesso giorno della settimana successiva
                
                current_date += timedelta(days=days_to_add)
                weekday = current_date.weekday()
            
            # Adatta la larghezza della colonna Date
            date_col_letter = openpyxl.utils.get_column_letter(date_col)
            ws.column_dimensions[date_col_letter].width = 15
            
            # Salva il file
            wb.save(self.excel_file)
            
            # Aggiorna il tipo di sport se necessario
            if 'Config' in wb.sheetnames:
                config_sheet = wb['Config']
                sport_type_found = False
                
                # Cerca la riga con il tipo di sport
                for row in range(1, config_sheet.max_row + 1):
                    if config_sheet.cell(row=row, column=1).value == 'sport_type':
                        sport_type_found = True
                        if config_sheet.cell(row=row, column=2).value != self.sport_type:
                            config_sheet.cell(row=row, column=2).value = self.sport_type
                            print(f"Sport type updated to {self.sport_type} in Config sheet")
                        break
                
                # Se non esiste, aggiungiamo una riga per il tipo di sport
                if not sport_type_found:
                    next_row = config_sheet.max_row + 1
                    config_sheet.cell(row=next_row, column=1).value = 'sport_type'
                    config_sheet.cell(row=next_row, column=2).value = self.sport_type
                    print(f"Sport type added as {self.sport_type} in Config sheet")
                    
                    # Formattazione
                    config_sheet.cell(row=next_row, column=1).border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style="thin"),
                        right=openpyxl.styles.Side(style="thin"),
                        top=openpyxl.styles.Side(style="thin"),
                        bottom=openpyxl.styles.Side(style="thin")
                    )
                    config_sheet.cell(row=next_row, column=2).border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style="thin"),
                        right=openpyxl.styles.Side(style="thin"),
                        top=openpyxl.styles.Side(style="thin"),
                        bottom=openpyxl.styles.Side(style="thin")
                    )
                
                # Salva nuovamente il file dopo le modifiche al foglio Config
                wb.save(self.excel_file)
            
            messagebox.showinfo("Successo", f"Date pianificate con successo per {len(workouts)} allenamenti di {self.sport_type}.")
            self.result = True
            self.destroy()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Errore", f"Errore durante la pianificazione: {str(e)}")

    def adjust_row_heights(worksheet, start_row, steps_col):
        """
        Regola l'altezza delle righe in base al contenuto della colonna dei passi
        
        Args:
            worksheet: foglio di lavoro
            start_row: riga di inizio (dopo l'intestazione)
            steps_col: indice della colonna con i passi dell'allenamento
        """
        for row in range(start_row, worksheet.max_row + 1):
            steps = worksheet.cell(row=row, column=steps_col).value
            if steps:
                # Conta le righe di testo nei passi (sia \n che ;)
                num_lines = 1 + steps.count('\n') + steps.count(';')
                
                # Considera anche l'indentazione per le ripetizioni
                if 'repeat' in steps and '\n' in steps:
                    # Conta le righe indentate dopo repeat
                    lines_after_repeat = steps.split('repeat')[1].count('\n')
                    if lines_after_repeat > 0:
                        num_lines += lines_after_repeat - 1  # -1 perché la riga con 'repeat' è già contata
                
                # Altezza minima più altezza per ogni riga di testo
                row_height = max(15, 12 * num_lines)
                worksheet.row_dimensions[row].height = row_height

class ExcelToYamlGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        
        self.title("Excel to YAML Converter for Garmin Planner")
        self.geometry("700x600")  # Leggermente più alta per aggiungere il selettore di sport
        self.minsize(600, 450)
        
        # Imposta variabili
        self.excel_file = tk.StringVar()
        self.yaml_file = tk.StringVar()
        self.sport_type = tk.StringVar(value="running")  # Default a running
        
        # Crea l'interfaccia
        self.create_widgets()
        
        # Centra la finestra
        self.center_window()
    
    def center_window(self):
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def create_widgets(self):
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Titolo
        title_label = ttk.Label(main_frame, text="Convertitore Excel-YAML per Garmin Planner", 
                             font=("", 14, "bold"))
        title_label.pack(pady=10)
        
        # Descrizione
        description = (
            "Questo strumento converte un file Excel strutturato in un file YAML compatibile con garmin-planner.\n"
            "Supporta la pianificazione automatica degli allenamenti in base ai giorni della settimana."
        )
        desc_label = ttk.Label(main_frame, text=description, wraplength=600, justify="center")
        desc_label.pack(pady=5)
        
        # Frame per i file e opzioni
        options_frame = ttk.LabelFrame(main_frame, text="Opzioni", padding="10")
        options_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Sport type selector
        sport_type_frame = ttk.Frame(options_frame)
        sport_type_frame.pack(fill=tk.X, pady=5)
        
        
        # Aggiungi una descrizione chiara
        sport_desc = "running = corsa, cycling = ciclismo"
        ttk.Label(sport_type_frame, text=sport_desc, foreground="gray").grid(row=0, column=2, padx=5, sticky=tk.W)
        
        # Frame per i file
        file_frame = ttk.LabelFrame(main_frame, text="File", padding="10")
        file_frame.pack(fill=tk.X, padx=5, pady=10)
        
        # Input Excel
        excel_label = ttk.Label(file_frame, text="File Excel di input:")
        excel_label.grid(row=0, column=0, sticky=tk.W, pady=5)
        
        excel_entry = ttk.Entry(file_frame, textvariable=self.excel_file, width=60)
        excel_entry.grid(row=0, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        excel_button = ttk.Button(file_frame, text="Sfoglia...", command=self.browse_excel)
        excel_button.grid(row=0, column=2, padx=5, pady=5)
        
        # Output YAML
        yaml_label = ttk.Label(file_frame, text="File YAML di output:")
        yaml_label.grid(row=1, column=0, sticky=tk.W, pady=5)
        
        yaml_entry = ttk.Entry(file_frame, textvariable=self.yaml_file, width=60)
        yaml_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W+tk.E)
        
        yaml_button = ttk.Button(file_frame, text="Sfoglia...", command=self.browse_yaml)
        yaml_button.grid(row=1, column=2, padx=5, pady=5)
        
        file_frame.columnconfigure(1, weight=1)
        
        # Area di log
        log_frame = ttk.LabelFrame(main_frame, text="Log", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=10)
        
        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Pulsanti azione
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=10)
        
        create_button = ttk.Button(button_frame, text="Crea file Excel di esempio", 
                                 command=self.create_sample)
        create_button.pack(side=tk.LEFT, padx=5)
        
        schedule_button = ttk.Button(button_frame, text="Pianifica Allenamenti", 
                                   command=self.schedule_workouts)
        schedule_button.pack(side=tk.LEFT, padx=5)
        
        convert_button = ttk.Button(button_frame, text="Converti Excel in YAML", 
                                  command=self.convert_file)
        convert_button.pack(side=tk.RIGHT, padx=5)
        
        # Versione
        version_label = ttk.Label(main_frame, 
                                text=f"v1.3.0 - {datetime.now().year}", 
                                foreground="gray")
        version_label.pack(side=tk.RIGHT, padx=10, pady=5)
        
        # Aggiungi un messaggio iniziale al log
        self.log("Benvenuto nel convertitore Excel-YAML per Garmin Planner!")
        self.log("Seleziona un file Excel o crea un file di esempio per iniziare.")
        self.log("Supporta sia allenamenti di corsa (running) che di ciclismo (cycling).")
    
    def log(self, message):
        """Aggiunge un messaggio all'area di log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")
        self.update_idletasks()
    
    def browse_excel(self):
        """Apre il dialogo per selezionare il file Excel di input"""
        file_path = filedialog.askopenfilename(
            title="Seleziona file Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file.set(file_path)
            self.log(f"File Excel selezionato: {file_path}")
            
            # Se il file YAML non è stato specificato, proponi lo stesso nome con estensione yaml
            if not self.yaml_file.get():
                yaml_path = os.path.splitext(file_path)[0] + ".yaml"
                self.yaml_file.set(yaml_path)
            
            # Esegui l'estrazione del tipo di sport se presente nel file
            self.extract_sport_type_from_excel(file_path)
    


    def extract_sport_type_from_excel(self, excel_path):
        """Estrai il tipo di sport dal foglio Config se esiste"""
        try:
            if os.path.exists(excel_path):
                # Carica il foglio Config
                xls = pd.ExcelFile(excel_path)
                if 'Config' in xls.sheet_names:
                    config_df = pd.read_excel(excel_path, sheet_name='Config')
                    sport_type_rows = config_df[config_df.iloc[:, 0] == 'sport_type']
                    
                    if not sport_type_rows.empty and pd.notna(sport_type_rows.iloc[0, 1]):
                        extracted_sport_type = str(sport_type_rows.iloc[0, 1]).strip().lower()
                        if extracted_sport_type in ["running", "cycling"]:
                            self.sport_type.set(extracted_sport_type)
                            self.log(f"Tipo di sport rilevato nel file: {extracted_sport_type}")
                            return
                
                # Se arriviamo qui, non abbiamo trovato un tipo di sport esplicito
                # Proviamo a dedurlo dalla presenza di fogli Paces o Speeds
                if 'Speeds' in xls.sheet_names and 'Paces' not in xls.sheet_names:
                    self.sport_type.set("cycling")
                    self.log("Dedotto tipo di sport 'cycling' dalla presenza del foglio Speeds")
                elif 'Paces' in xls.sheet_names and 'Speeds' not in xls.sheet_names:
                    self.sport_type.set("running")
                    self.log("Dedotto tipo di sport 'running' dalla presenza del foglio Paces")
                else:
                    # Usa il valore predefinito attuale
                    self.log(f"Tipo di sport non rilevato nel file, usiamo il default: {self.sport_type.get()}")
        except Exception as e:
            self.log(f"Errore nell'estrazione del tipo di sport: {str(e)}")


    def browse_yaml(self):
        """Apre il dialogo per selezionare il file YAML di output"""
        file_path = filedialog.asksaveasfilename(
            title="Salva file YAML",
            filetypes=[("YAML files", "*.yaml *.yml"), ("All files", "*.*")],
            defaultextension=".yaml"
        )
        if file_path:
            self.yaml_file.set(file_path)
            self.log(f"File YAML di output impostato: {file_path}")
    
    def schedule_workouts(self):
        """Apre il dialogo per pianificare gli allenamenti"""
        excel_path = self.excel_file.get()
        
        if not excel_path:
            messagebox.showerror("Errore", "Seleziona prima un file Excel.")
            return
        
        if not os.path.exists(excel_path):
            if messagebox.askyesno("File non trovato", 
                                  f"Il file {excel_path} non esiste.\n\nVuoi creare un file di esempio?"):
                sport_type = self.sport_type.get()
                excel_path = create_sample_excel(excel_path, sport_type)
                self.excel_file.set(excel_path)
                self.log(f"Creato file Excel di esempio per {sport_type} in: {excel_path}")
            else:
                return
        
        try:
            # Apri il dialogo di pianificazione
            dialog = ScheduleDialog(self, excel_path)
            
            # Se la pianificazione è stata completata con successo
            if dialog.result:
                self.log("Pianificazione degli allenamenti completata con successo.")
                self.log("Le date sono state aggiunte al file Excel.")
        except Exception as e:
            self.log(f"Errore durante la pianificazione: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore durante la pianificazione:\n{str(e)}")

    
    def create_sample(self):
        """Crea un file Excel di esempio"""
        sport_type = self.sport_type.get()
        file_path = filedialog.asksaveasfilename(
            title=f"Salva file Excel di esempio per {sport_type}",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            defaultextension=".xlsx"
        )
        if not file_path:
            return
        
        try:
            self.log(f"Creazione file Excel di esempio per {sport_type}: {file_path}")
            
            # Esegui la creazione in un thread separato per non bloccare l'interfaccia
            threading.Thread(target=self._create_sample_thread, args=(file_path, sport_type)).start()
        except Exception as e:
            self.log(f"Errore: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore durante la creazione del file:\n{str(e)}")
    
    
    def _create_sample_thread(self, file_path, sport_type):
        """Thread per la creazione del file di esempio"""
        try:
            create_sample_excel(file_path, sport_type)
            
            # Aggiorna l'interfaccia nel thread principale
            self.after(0, lambda: self._sample_created(file_path, sport_type))
        except Exception as e:
            self.after(0, lambda: self._sample_error(str(e)))
    
    def _sample_created(self, file_path, sport_type):
        """Chiamato quando il file di esempio è stato creato"""
        self.log(f"File Excel di esempio creato con successo per {sport_type}!")
        self.excel_file.set(file_path)
        
        # Proponi un file YAML con lo stesso nome
        yaml_path = os.path.splitext(file_path)[0] + ".yaml"
        self.yaml_file.set(yaml_path)
        
        messagebox.showinfo("Successo", 
                          f"File Excel di esempio per {sport_type} creato con successo:\n{file_path}\n\n"
                          f"Ora puoi utilizzare il pulsante 'Pianifica Allenamenti' per assegnare date agli allenamenti.")
  
    
    def _sample_error(self, error_msg):
        """Chiamato in caso di errore nella creazione del file di esempio"""
        self.log(f"Errore nella creazione del file di esempio: {error_msg}")
        messagebox.showerror("Errore", f"Si è verificato un errore durante la creazione del file:\n{error_msg}")
    
    def convert_file(self):
        """Converte il file Excel in YAML"""
        excel_path = self.excel_file.get()
        yaml_path = self.yaml_file.get()
        sport_type = self.sport_type.get()
        
        if not excel_path:
            messagebox.showerror("Errore", "Seleziona un file Excel da convertire")
            return
        
        if not os.path.exists(excel_path):
            messagebox.showerror("Errore", f"Il file Excel non esiste: {excel_path}")
            return
        
        if not yaml_path:
            # Usa lo stesso nome del file Excel ma con estensione yaml
            yaml_path = os.path.splitext(excel_path)[0] + ".yaml"
            self.yaml_file.set(yaml_path)
        
        try:
            self.log(f"Conversione di {excel_path} in {yaml_path} per {sport_type}...")
            
            # Esegui la conversione in un thread separato per non bloccare l'interfaccia
            threading.Thread(target=self._convert_thread, args=(excel_path, yaml_path, sport_type)).start()
        except Exception as e:
            self.log(f"Errore: {str(e)}")
            messagebox.showerror("Errore", f"Si è verificato un errore durante la conversione:\n{str(e)}")

    
    def _convert_thread(self, excel_path, yaml_path, sport_type):
        """Thread per la conversione"""
        try:
            plan = excel_to_yaml(excel_path, yaml_path, sport_type)
            
            # Aggiorna l'interfaccia nel thread principale
            self.after(0, lambda: self._conversion_success(yaml_path, len(plan) - 1, sport_type))  # -1 per escludere config
        except Exception as e:
            # Cattura l'errore e memorizza il messaggio
            error_msg = str(e)
            # Usa una variabile locale invece di riferirsi a 'e' nella lambda
            self.after(0, lambda msg=error_msg: self._conversion_error(msg))
    
    def _conversion_success(self, yaml_path, workout_count, sport_type):
        """Chiamato quando la conversione è stata completata con successo"""
        self.log(f"Conversione completata con successo!")
        self.log(f"File YAML salvato: {yaml_path}")
        self.log(f"Piano di allenamento per {sport_type} con {workout_count} allenamenti")
        
        messagebox.showinfo("Successo", 
                          f"Conversione completata con successo!\n\n"
                          f"File YAML salvato: {yaml_path}\n"
                          f"Tipo di sport: {sport_type}\n"
                          f"Allenamenti creati: {workout_count}")
    
    def _conversion_error(self, error_msg):
        """Chiamato in caso di errore nella conversione"""
        self.log(f"Errore nella conversione: {error_msg}")
        messagebox.showerror("Errore", f"Si è verificato un errore durante la conversione:\n{error_msg}")

if __name__ == "__main__":
    app = ExcelToYamlGUI()
    app.mainloop()